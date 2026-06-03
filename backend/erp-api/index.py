"""
ERP API — универсальный endpoint для всех операций системы.
Роуты: /deals, /projects, /procurement, /payments, /kcompany, /dashboard, /clients, /staff, /employees, /reports
"""
import json
import os
import base64
import random
import string
import logging
import traceback
import psycopg2
import boto3
from datetime import date, datetime, timedelta
from decimal import Decimal

SCHEMA = "t_p60494808_erp_system_creation"

# Логирование: пишет в CloudWatch / stdout облачной функции
logger = logging.getLogger("erp-api")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    h = logging.StreamHandler()
    h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(h)

CORS_HEADERS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type, X-User-Id, X-User-Role, X-Auth-Token",
    "Content-Type": "application/json",
}

def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])

def json_serial(obj):
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError(f"Type {type(obj)} not serializable")

def ok(data, status=200):
    return {
        "statusCode": status,
        "headers": CORS_HEADERS,
        "body": json.dumps(data, default=json_serial, ensure_ascii=False),
    }

def err(msg, status=400):
    return {
        "statusCode": status,
        "headers": CORS_HEADERS,
        "body": json.dumps({"error": msg}, ensure_ascii=False),
    }

def next_code(cur, table, prefix, col="code"):
    # Используем MAX по числовому суффиксу + FOR UPDATE чтобы избежать race condition
    cur.execute(f"""
        SELECT COALESCE(MAX(CAST(SPLIT_PART({col}, '-', 2) AS INTEGER)), 0) + 1
        FROM {SCHEMA}.{table}
        WHERE {col} LIKE %s
    """, (f"{prefix}-%",))
    n = cur.fetchone()[0]
    return f"{prefix}-{n:04d}"

# ─── DEALS ───────────────────────────────────────────────────────────────────

MIN_SLOT_LAG_DAYS = 10  # Минимальное плечо: слот не раньше чем через 10 дней от сегодня

def get_free_slots(cur, signed_date_str: str = None):
    """
    Возвращает свободные слоты.
    Минимальное плечо — 10 дней от сегодня.
    signed_date_str: если передана дата подписания — дополнительно фильтрует
    по signed_date + 15 дней.
    Архивные слоты не попадают в список и не влияют на загрузку.
    """
    min_start = date.today() + timedelta(days=MIN_SLOT_LAG_DAYS)
    if signed_date_str:
        try:
            signed = date.fromisoformat(signed_date_str)
            candidate = signed + timedelta(days=15)
            if candidate > min_start:
                min_start = candidate
        except Exception:
            pass

    cur.execute(f"""
        SELECT
            s.id, s.year, s.month, s.start_date, s.status, s.monthly_limit,
            COUNT(occupied.id) AS occupied_count
        FROM {SCHEMA}.slots s
        LEFT JOIN {SCHEMA}.slots occupied
            ON occupied.year = s.year
            AND occupied.month = s.month
            AND occupied.status IN ('booked', 'busy')
        WHERE s.status = 'free'
          AND s.start_date >= %s
        GROUP BY s.id, s.year, s.month, s.start_date, s.status, s.monthly_limit
        ORDER BY s.start_date
    """, (min_start,))
    cols = [desc[0] for desc in cur.description]
    slots = []
    for row in cur.fetchall():
        d = dict(zip(cols, row))
        d["available"] = int(d["monthly_limit"]) - int(d["occupied_count"]) > 0
        d["occupied_count"] = int(d["occupied_count"])
        slots.append(d)
    return slots

def get_slot_plan(cur, show_archived: bool = False):
    """Полная картина слот-плана по месяцам + детальные слоты.
    Архивные слоты не учитываются в загрузке (прогресс-бар).
    """
    # Месяцы — только не-архивные слоты для расчёта загрузки
    cur.execute(f"""
        SELECT
            s.year, s.month, s.monthly_limit,
            COUNT(*) FILTER (WHERE s.status = 'free')     AS free_count,
            COUNT(*) FILTER (WHERE s.status = 'booked')   AS booked_count,
            COUNT(*) FILTER (WHERE s.status = 'busy')     AS busy_count,
            COUNT(*) FILTER (WHERE s.status = 'archived') AS archived_count
        FROM {SCHEMA}.slots s
        WHERE s.status != 'archived'
        GROUP BY s.year, s.month, s.monthly_limit
        ORDER BY s.year, s.month
    """)
    cols = [desc[0] for desc in cur.description]
    months = []
    for row in cur.fetchall():
        d = dict(zip(cols, row))
        total_occupied = int(d["booked_count"]) + int(d["busy_count"])
        d["total_occupied"] = total_occupied
        d["load_pct"] = round(total_occupied / int(d["monthly_limit"]) * 100) if d["monthly_limit"] else 0
        d["overloaded"] = total_occupied >= int(d["monthly_limit"])
        months.append(d)

    # Детальный список слотов — только активные проекты (не archived/completed)
    where_archived = "" if show_archived else "WHERE s.status != 'archived'"
    cur.execute(f"""
        SELECT
            s.id, s.year, s.month, s.start_date, s.status, s.monthly_limit,
            s.deal_id,
            d.code AS deal_code,
            c.name AS client_name,
            p.id   AS project_id,
            p.code AS project_code,
            p.status AS project_status
        FROM {SCHEMA}.slots s
        LEFT JOIN {SCHEMA}.deals d ON d.id = s.deal_id AND d.is_archived = FALSE
        LEFT JOIN {SCHEMA}.clients c ON c.id = d.client_id
        LEFT JOIN {SCHEMA}.projects p ON p.slot_id = s.id AND p.status NOT IN ('archived', 'completed')
        {where_archived}
        ORDER BY s.start_date
    """)
    scols = [desc[0] for desc in cur.description]
    slots = [dict(zip(scols, r)) for r in cur.fetchall()]

    return {"months": months, "slots": slots}

def create_slots(cur, year: int, month: int, count: int, monthly_limit: int):
    """
    Создаёт слоты на выбранный месяц (каждую неделю).
    Если слот на эту дату уже существует — пропускает.
    """
    import calendar
    # Генерируем даты: первый понедельник месяца + каждые 7 дней, итого count штук
    first_day = date(year, month, 1)
    # Находим первый понедельник (или 1-е число если понедельник)
    weekday = first_day.weekday()
    if weekday != 0:
        first_monday = first_day + timedelta(days=(7 - weekday))
    else:
        first_monday = first_day

    created = []
    current = first_monday
    for i in range(count):
        if current.month != month:
            break
        # Проверяем — есть ли уже слот на эту дату
        cur.execute(f"""
            SELECT id FROM {SCHEMA}.slots WHERE start_date = %s
        """, (current,))
        if not cur.fetchone():
            cur.execute(f"""
                INSERT INTO {SCHEMA}.slots (year, month, start_date, status, monthly_limit)
                VALUES (%s, %s, %s, 'free', %s)
                RETURNING id
            """, (year, month, current, monthly_limit))
            new_id = cur.fetchone()[0]
            created.append({"id": new_id, "start_date": current.isoformat()})
        # Обновляем monthly_limit для всех слотов этого месяца
        cur.execute(f"""
            UPDATE {SCHEMA}.slots SET monthly_limit = %s
            WHERE year = %s AND month = %s
        """, (monthly_limit, year, month))
        current += timedelta(days=7)

    return {"created": created, "count": len(created)}


def delete_slot(cur, slot_id: int):
    """Удаляет слот если он свободен."""
    cur.execute(f"""
        SELECT id, status FROM {SCHEMA}.slots WHERE id = %s
    """, (slot_id,))
    row = cur.fetchone()
    if not row:
        return None, "Слот не найден"
    if row[1] != 'free':
        return None, "Нельзя удалить зарезервированный или занятый слот"
    cur.execute(f"DELETE FROM {SCHEMA}.slots WHERE id = %s", (slot_id,))
    return {"deleted": slot_id}, None


def get_user_context(event: dict):
    """Извлекает (user_id, user_role) из заголовков запроса.
    Заголовки case-insensitive — нормализуем в lowercase.
    """
    headers = event.get("headers") or {}
    norm = {k.lower(): v for k, v in headers.items() if isinstance(k, str)}
    role = norm.get("x-user-role") or norm.get("x-userrole") or ""
    uid_raw = norm.get("x-user-id") or norm.get("x-userid") or ""
    try:
        uid = int(uid_raw) if uid_raw not in ("", None) else None
    except (TypeError, ValueError):
        uid = None
    return uid, (role or "").strip()


def get_verified_role(cur, user_id, claimed_role=""):
    """Возвращает РЕАЛЬНУЮ роль пользователя из БД по user_id.
    Заголовок X-User-Role не является источником истины — его легко подделать.
    Для критичных операций (деньги, смена статуса) роль берём ТОЛЬКО отсюда.
    Если user_id не передан или сотрудник не найден — возвращаем "" (нет прав).
    Демо-режим: если user_id отсутствует, но claimed_role задан — доверяем заголовку
    (нужно для витрины без авторизации; критичные операции всё равно потребуют user_id).
    """
    if user_id is None:
        return (claimed_role or "").strip()
    cur.execute(
        f"SELECT role FROM {SCHEMA}.staff WHERE id = %s AND is_active = TRUE",
        (int(user_id),),
    )
    row = cur.fetchone()
    return (row[0] if row else "").strip()


def require_role(cur, user_id, claimed_role, allowed_roles):
    """Проверяет, что РЕАЛЬНАЯ роль пользователя входит в allowed_roles.
    Бросает PermissionError, если нет прав. Используется для критичных операций.
    """
    real_role = get_verified_role(cur, user_id, claimed_role)
    if real_role not in allowed_roles:
        raise PermissionError(
            f"Недостаточно прав для этой операции (требуется одна из ролей: {', '.join(allowed_roles)})"
        )
    return real_role


def require_deal_owner(cur, deal_id, user_id, claimed_role):
    """Owner-check: менеджер/риэлтор может менять ТОЛЬКО свою сделку.
    Директор / коммерческий / демо-режим (роль не определена) — любую.
    Бросает PermissionError если пользователь пытается изменить чужую сделку.
    """
    real_role = get_verified_role(cur, user_id, claimed_role)
    # Полный доступ — менять любую сделку
    if real_role in DEALS_FULL_ACCESS or real_role == "":
        return
    if real_role not in ("crm_manager", "realtor"):
        # Прочие роли к сделкам не имеют доступа
        raise PermissionError("Недостаточно прав для изменения сделки")
    # Менеджер/риэлтор — только свои
    cur.execute(
        f"SELECT manager_id, realtor_id FROM {SCHEMA}.deals WHERE id=%s",
        (int(deal_id),),
    )
    row = cur.fetchone()
    if not row:
        raise ValueError("Сделка не найдена")
    manager_id, realtor_id = row
    owner_id = manager_id if real_role == "crm_manager" else realtor_id
    if user_id is None or owner_id != user_id:
        raise PermissionError("Можно изменять только свои сделки")


# Роли, которым видны все сделки без фильтра
DEALS_FULL_ACCESS = {"director", "commercial", "general_director", "commercial_director"}
# Роли, которым сделки в принципе не положены
DEALS_NO_ACCESS = {
    "construction_director", "supply_director", "finance_director",
    "foreman", "supplier", "mechanic", "accountant", "client",
    "project_manager", "quality",
}


def get_deals(cur, archived=False, user_id=None, user_role=""):
    """Список сделок с фильтрацией по роли пользователя.
    - crm_manager → только свои (manager_id = user_id)
    - realtor → только свои (realtor_id = user_id)
    - director / commercial → все
    - остальные → пустой список
    """
    role = (user_role or "").strip()

    # Базовый фильтр по архивности
    base_where = "d.is_archived = TRUE" if archived else "d.is_archived = FALSE"

    # Роли без доступа к сделкам — сразу пустой список
    if role in DEALS_NO_ACCESS:
        return []

    role_where = ""
    params: tuple = ()
    if role in DEALS_FULL_ACCESS or role == "":
        # Полный доступ. Пустая роль = демо/админ — без фильтра.
        role_where = ""
    elif role == "crm_manager":
        # Только сделки где manager_id = текущий пользователь.
        # Если user_id не передан — пустой список (безопаснее чем показывать всё).
        if not user_id:
            return []
        role_where = " AND d.manager_id = %s"
        params = (int(user_id),)
    elif role == "realtor":
        # Только сделки где realtor_id = текущий пользователь.
        if not user_id:
            return []
        role_where = " AND d.realtor_id = %s"
        params = (int(user_id),)
    else:
        # Неизвестная роль — ничего не показываем
        return []

    cur.execute(f"""
        SELECT d.id, d.code, d.stage, d.budget, d.start_date, d.source, d.notes, d.created_at,
               c.name as client_name, c.phone as client_phone,
               sm.name as manager_name, sr.name as realtor_name,
               d.manager_id, d.realtor_id,
               sl.id as slot_id, sl.year as slot_year, sl.month as slot_month,
               sl.start_date as slot_start_date, sl.status as slot_status,
               d.project_id, d.project_type,
               sp.name as serial_project_name,
               cfg.name as configuration_name,
               cfg.duration_days as configuration_duration,
               cfg.price_coefficient,
               d.selected_stages, d.signed_date, d.buffer_days,
               d.kp_notes, d.address, d.planned_start_date,
               d.serial_project_id, d.configuration_id,
               COALESCE(d.contract_status, 'none') as contract_status,
               d.last_reject_reason,
               d.is_archived,
               COALESCE(d.kp_slot_id, 0) as kp_slot_id,
               COALESCE(d.payment_confirmed, false) as payment_confirmed,
               COALESCE(d.contract_signed, false) as contract_signed,
               ksl.start_date as kp_slot_start_date, ksl.year as kp_slot_year, ksl.month as kp_slot_month,
               d.client_token,
               d.commission_rate, d.commission_amount, d.closed_at
        FROM {SCHEMA}.deals d
        LEFT JOIN {SCHEMA}.clients c ON c.id = d.client_id
        LEFT JOIN {SCHEMA}.staff sm ON sm.id = d.manager_id
        LEFT JOIN {SCHEMA}.staff sr ON sr.id = d.realtor_id
        LEFT JOIN {SCHEMA}.slots sl ON sl.id = d.slot_id
        LEFT JOIN {SCHEMA}.slots ksl ON ksl.id = d.kp_slot_id
        LEFT JOIN {SCHEMA}.serial_projects sp ON sp.id = d.serial_project_id
        LEFT JOIN {SCHEMA}.configurations cfg ON cfg.id = d.configuration_id
        WHERE {base_where}{role_where}
        ORDER BY d.created_at DESC
        LIMIT 100
    """, params)
    cols = [desc[0] for desc in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]

def get_stage_durations(cur):
    """Загружает нормативы этапов из БД (единый источник правды). Исключает деактивированные."""
    cur.execute(f"""
        SELECT stage_num, stage_name, duration_days, parallel_group, depends_on
        FROM {SCHEMA}.stage_durations
        WHERE stage_name NOT LIKE '_DISABLED_%%'
          AND duration_days > 0
        ORDER BY sort_order, stage_num
    """)
    result = []
    for row in cur.fetchall():
        result.append({
            "stage_num": row[0],
            "name": row[1],
            "duration": row[2],
            "parallel_group": row[3],
            "depends_on": list(row[4]) if row[4] else [],
        })
    return result

def recalc_active_projects_stages(cur) -> dict:
    """
    Пересчитывает planned_start / planned_end / duration_days во всех активных проектах
    (status IN ('active', 'planning')) на основе актуальных нормативов из stage_durations.

    Логика:
    - берём start_date проекта как точку отсчёта
    - для каждого этапа в project_stages обновляем duration_days из stage_durations
    - пересчитываем planned_start / planned_end через _build_plan_from_norms
    - этапы с actual_start (уже начаты) не сдвигаем по дате начала, но обновляем длительность
    - обновляем deadline проекта по максимальной planned_end
    """
    stage_norms = get_stage_durations(cur)
    norm_map = {s["stage_num"]: s for s in stage_norms}

    # Находим все активные проекты
    cur.execute(f"""
        SELECT id, start_date, status
        FROM {SCHEMA}.projects
        WHERE status IN ('active', 'planning')
        ORDER BY id
    """)
    projects = cur.fetchall()
    updated_projects = 0

    for proj_id, proj_start, _ in projects:
        if proj_start is None:
            continue
        start_dt = proj_start if isinstance(proj_start, date) else date.fromisoformat(str(proj_start))

        # Получаем этапы проекта
        cur.execute(f"""
            SELECT id, stage_num, order_num, actual_start, status
            FROM {SCHEMA}.project_stages
            WHERE project_id = %s
            ORDER BY order_num
        """, (proj_id,))
        stages = cur.fetchall()
        if not stages:
            continue

        # Собираем нормативы только для тех stage_num, которые есть в проекте
        included_nums = [s[1] for s in stages if s[1] is not None]
        if not included_nums:
            continue

        filtered_norms = [norm_map[n] for n in included_nums if n in norm_map]
        if not filtered_norms:
            continue

        # Строим новый Гант-план
        plan = _build_plan_from_norms(start_dt, filtered_norms)
        plan_map = {p[0]: p for p in plan}  # stage_num -> (snum, name, duration, ps, pe, ...)

        max_end = start_dt

        for stage_row in stages:
            stage_id, stage_num, order_num, actual_start, stage_status = stage_row
            if stage_num not in plan_map:
                continue

            p = plan_map[stage_num]
            new_duration = p[2]
            new_ps = p[3]
            new_pe = p[4]

            # Если этап уже идёт — сдвигаем только дату окончания (не трогаем начало)
            if actual_start is not None and stage_status in ('in_progress', 'done'):
                actual_dt = actual_start if isinstance(actual_start, date) else date.fromisoformat(str(actual_start))
                new_pe = actual_dt + timedelta(days=new_duration - 1)
                new_ps = actual_dt

            cur.execute(f"""
                UPDATE {SCHEMA}.project_stages
                SET duration_days = %s,
                    planned_start = %s,
                    planned_end   = %s,
                    updated_at    = now()
                WHERE id = %s
            """, (new_duration, new_ps, new_pe, stage_id))

            if new_pe > max_end:
                max_end = new_pe

        # Обновляем deadline проекта
        cur.execute(f"""
            UPDATE {SCHEMA}.projects
            SET deadline = %s, updated_at = now()
            WHERE id = %s
        """, (max_end, proj_id))
        updated_projects += 1

    return {"recalculated_projects": updated_projects}

def calc_duration_for_stages(stage_norms: list, included_nums: list, buffer_days: int = 7) -> int:
    """Считает итоговую длительность для выбранного набора этапов + буфер."""
    filtered = [s for s in stage_norms if s["stage_num"] in included_nums]
    if not filtered:
        return buffer_days
    # Строим план дат, берём максимальную дату завершения
    start = date(2000, 1, 1)
    plan = _build_plan_from_norms(start, filtered)
    if not plan:
        return buffer_days
    max_end = max(p[4] for p in plan)
    return (max_end - start).days + 1 + buffer_days

def _build_plan_from_norms(start_date: date, stage_norms: list) -> list:
    """Внутренняя: строит даты по нормативам из БД."""
    stage_map = {s["stage_num"]: s for s in stage_norms}
    end_dates = {}
    result = []
    for snum in sorted(stage_map.keys()):
        s = stage_map[snum]
        deps = [d for d in s["depends_on"] if d in end_dates]
        if deps:
            s_date = max(end_dates[d] for d in deps) + timedelta(days=1)
        else:
            s_date = start_date
        # Параллельная группа — все стартуют вместе
        if s["parallel_group"] is not None:
            group_starts = [r[3] for r in result if r[5] == s["parallel_group"]]
            if group_starts:
                s_date = min(group_starts)
        e_date = s_date + timedelta(days=s["duration"] - 1)
        end_dates[snum] = e_date
        result.append((snum, s["name"], s["duration"], s_date, e_date, s["parallel_group"], s["depends_on"]))
    return result

def create_deal(cur, body, user_id=None, user_role=""):
    """Создание нового лида — только базовые данные: клиент, тип проекта, источник.
    Если создаёт риэлтор — realtor_id принудительно = текущему пользователю,
    чтобы он не мог записать сделку на чужое имя.
    """
    client_id    = int(body["client_id"])
    # manager_id может быть null (когда сделку создаёт риэлтор без менеджера)
    raw_mgr      = body.get("manager_id")
    manager_id   = int(raw_mgr) if raw_mgr not in (None, "", "null") else None
    realtor_id   = body.get("realtor_id")
    # Если запрос от риэлтора — принудительно фиксируем его как realtor_id
    if (user_role or "").strip() == "realtor" and user_id:
        realtor_id = user_id
    source       = body.get("source", "")
    notes        = body.get("notes", "")
    project_type = body.get("project_type", "serial")
    sp_id        = body.get("serial_project_id")
    address      = body.get("address", "")
    slot_id      = body.get("slot_id")

    code = next_code(cur, "deals", "ЛД")
    realtor_val = int(realtor_id) if realtor_id not in (None, "", "null") else None
    sp_val      = int(sp_id) if sp_id else None
    slot_val    = int(slot_id) if slot_id else None

    cur.execute(f"""
        INSERT INTO {SCHEMA}.deals
            (code, client_id, manager_id, realtor_id, source, notes,
             stage, project_type, serial_project_id, address, start_date, slot_id)
        VALUES (%s, %s, %s, %s, %s, %s, 'lead', %s, %s, %s, CURRENT_DATE, %s)
        RETURNING id, code
    """, (code, client_id, manager_id, realtor_val, source, notes,
          project_type, sp_val, address, slot_val))

    deal_id, deal_code = cur.fetchone()

    # Резервируем слот атомарно — привязываем к deal_id сразу
    if slot_val:
        cur.execute(f"""
            UPDATE {SCHEMA}.slots SET status='booked', deal_id=%s
            WHERE id=%s AND status='free'
            RETURNING id
        """, (deal_id, slot_val))
        if not cur.fetchone():
            return None, "Выбранный слот уже занят — пожалуйста, выберите другой"

    # Для индивидуального — сразу создаём карточку проектирования
    if project_type == 'individual':
        desired_area = float(body.get("desired_area", 0))
        spec_req = body.get("special_requests", "")
        cur.execute(f"""
            INSERT INTO {SCHEMA}.individual_project_requests
                (deal_id, client_id, desired_area, special_requests, status)
            VALUES (%s, %s, %s, %s, 'awaiting_design')
        """, (deal_id, client_id, desired_area, spec_req))

    return {"id": deal_id, "code": deal_code, "project_type": project_type}, None

def update_deal_kp(cur, deal_id, body):
    """
    Заполнение данных на стадии КП:
    - серийный проект, комплектация / выбранные этапы, бюджет, буфер
    """
    cfg_id          = body.get("configuration_id")
    selected_stages = body.get("selected_stages")  # кастомный набор этапов
    budget          = body.get("budget")
    kp_notes        = body.get("kp_notes", "")
    buffer_days     = int(body.get("buffer_days", 7))
    sp_id           = body.get("serial_project_id")

    sets, vals = [], []
    if cfg_id is not None:
        sets.append("configuration_id=%s"); vals.append(int(cfg_id))
    if sp_id is not None:
        sets.append("serial_project_id=%s"); vals.append(int(sp_id))
    if selected_stages is not None:
        sets.append("selected_stages=%s"); vals.append(selected_stages)
    if budget is not None:
        sets.append("budget=%s"); vals.append(float(budget))
    if kp_notes:
        sets.append("kp_notes=%s"); vals.append(kp_notes)
    sets.append("buffer_days=%s"); vals.append(buffer_days)
    sets.append("stage='kp'")
    sets.append("updated_at=now()")
    vals.append(deal_id)

    cur.execute(f"UPDATE {SCHEMA}.deals SET {', '.join(sets)} WHERE id=%s RETURNING id, stage, budget, contractor_id", vals)
    row = cur.fetchone()
    if not row:
        return None, "Сделка не найдена"
    # Получаем имя клиента для заголовка документа
    cur.execute(f"""
        SELECT d.code, c.name as client_name, d.contractor_id
        FROM {SCHEMA}.deals d
        LEFT JOIN {SCHEMA}.clients c ON c.id = d.client_id
        WHERE d.id = %s
    """, (deal_id,))
    drow = cur.fetchone()
    if drow:
        auto_create_deal_documents(cur, deal_id, "kp", {
            "code": drow[0], "client_name": drow[1],
            "budget": row[2], "contractor_id": drow[2],
        })
    return {"id": row[0], "stage": row[1]}, None

def update_deal_contract(cur, deal_id, body):
    """
    Подписание договора:
    - слот, плановая дата начала, адрес, подтверждение бюджета
    - автоматически создаёт проект с этапами
    """
    slot_id         = body.get("slot_id")
    planned_start   = body.get("planned_start_date")
    address         = body.get("address", "")
    budget          = body.get("budget")
    signed_date     = body.get("signed_date") or date.today().isoformat()

    # Получаем сделку
    cur.execute(f"""
        SELECT id, client_id, project_type, configuration_id, selected_stages, buffer_days
        FROM {SCHEMA}.deals WHERE id=%s
    """, (deal_id,))
    deal_row = cur.fetchone()
    if not deal_row:
        return None, "Сделка не найдена"
    did, client_id, project_type, cfg_id, sel_stages, buf_days = deal_row
    buf_days = buf_days or 7

    sets, vals = [], []

    # Обрабатываем слот (только для серийных) — защита от двойного бронирования через SELECT FOR UPDATE
    actual_start = planned_start
    if project_type == 'serial' and slot_id:
        slot_id = int(slot_id)
        # SELECT FOR UPDATE блокирует строку до конца транзакции — защита от race condition
        cur.execute(f"""
            SELECT id, year, month, start_date, status, monthly_limit
            FROM {SCHEMA}.slots WHERE id=%s FOR UPDATE
        """, (slot_id,))
        slot_row = cur.fetchone()
        if not slot_row:
            return None, "Слот не найден"
        s_id, s_year, s_month, s_start_date, s_status, s_limit = slot_row
        if s_status != 'free':
            return None, "Слот уже занят или зарезервирован — выберите другой"
        cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.slots WHERE year=%s AND month=%s AND status IN ('booked','busy')", (s_year, s_month))
        if int(cur.fetchone()[0]) + 1 > s_limit:
            return None, f"Месяц перегружен (лимит {s_limit}). Выберите другой слот"
        # start_date = дата слота (без буфера — буфер только для display)
        slot_start = s_start_date if isinstance(s_start_date, date) else date.fromisoformat(str(s_start_date))
        actual_start = slot_start.isoformat()
        sets.append("slot_id=%s"); vals.append(slot_id)
        cur.execute(f"UPDATE {SCHEMA}.slots SET status='booked', deal_id=%s WHERE id=%s AND status='free' RETURNING id", (deal_id, slot_id))
        if not cur.fetchone():
            return None, "Слот был занят параллельным запросом — выберите другой"

    if actual_start:
        sets.append("planned_start_date=%s"); vals.append(actual_start)
    if address:
        sets.append("address=%s"); vals.append(address)
    if budget:
        sets.append("budget=%s"); vals.append(float(budget))
    sets.append("signed_date=%s");   vals.append(signed_date)
    sets.append("stage='contract'")  # Промежуточный статус — до создания проекта
    sets.append("updated_at=now()")
    vals.append(deal_id)

    cur.execute(f"UPDATE {SCHEMA}.deals SET {', '.join(sets)} WHERE id=%s", vals)

    # Автосоздание проекта
    start_for_project = actual_start or date.today().isoformat()
    project_id, perr = create_project_from_deal(cur, deal_id, client_id, start_for_project, slot_id)
    if perr:
        return None, perr

    # Если проект создан — переводим сделку в planning
    if project_id:
        cur.execute(f"""
            UPDATE {SCHEMA}.deals SET stage='planning', updated_at=now() WHERE id=%s
        """, (deal_id,))

    # Генерируем client_token если ещё нет
    cur.execute(f"SELECT client_token FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
    row = cur.fetchone()
    if not row or not row[0]:
        token = "CL-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=4)) + "-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=4))
        cur.execute(f"UPDATE {SCHEMA}.deals SET client_token=%s WHERE id=%s", (token, deal_id))
    else:
        token = row[0]

    return {"id": deal_id, "stage": "planning" if project_id else "contract", "project_id": project_id, "client_token": token}, None

def save_kp_slot(cur, deal_id: int, body: dict):
    """
    Шаг 1 КП: менеджер выбирает слот.
    Слот НЕ бронируется — только сохраняется как kp_slot_id.
    """
    kp_slot_id = body.get("kp_slot_id")
    if not kp_slot_id:
        return None, "Укажите слот"

    slot_id = int(kp_slot_id)
    # Проверяем что слот свободен
    cur.execute(f"SELECT id, status FROM {SCHEMA}.slots WHERE id=%s", (slot_id,))
    slot_row = cur.fetchone()
    if not slot_row:
        return None, "Слот не найден"
    if slot_row[1] not in ('free',):
        return None, "Слот уже занят или зарезервирован — выберите другой"

    cur.execute(f"""
        UPDATE {SCHEMA}.deals SET kp_slot_id=%s, updated_at=now() WHERE id=%s RETURNING id
    """, (slot_id, deal_id))
    if not cur.fetchone():
        return None, "Сделка не найдена"
    return {"deal_id": deal_id, "kp_slot_id": slot_id}, None


def confirm_kp_contract(cur, deal_id: int):
    """
    Шаг 2 КП: менеджер отмечает что договор подписан.
    """
    cur.execute(f"""
        UPDATE {SCHEMA}.deals
        SET contract_signed=true, contract_signed_at=now(), signed_date=CURRENT_DATE, updated_at=now()
        WHERE id=%s AND kp_slot_id IS NOT NULL
        RETURNING id
    """, (deal_id,))
    if not cur.fetchone():
        return None, "Сначала выберите слот"
    return {"deal_id": deal_id, "contract_signed": True}, None


def confirm_kp_payment(cur, deal_id: int):
    """
    Шаг 3 КП: менеджер подтверждает оплату аванса.
    """
    cur.execute(f"""
        UPDATE {SCHEMA}.deals
        SET payment_confirmed=true, updated_at=now()
        WHERE id=%s AND contract_signed=true
        RETURNING id
    """, (deal_id,))
    if not cur.fetchone():
        return None, "Сначала отметьте договор как подписанный"
    return {"deal_id": deal_id, "payment_confirmed": True}, None


def move_to_planning(cur, deal_id: int, body: dict):
    """
    Шаг 4 КП: кнопка «Перевести в планирование».
    - слот (kp_slot_id) → booked (SELECT FOR UPDATE)
    - создаётся проект со статусом planning
    - сделка → stage=planning, slot_id=kp_slot_id
    Условие: contract_signed=true AND payment_confirmed=true
    """
    # Читаем сделку
    cur.execute(f"""
        SELECT d.id, d.client_id, d.project_type, d.kp_slot_id,
               d.contract_signed, d.payment_confirmed,
               d.configuration_id, d.selected_stages, d.buffer_days,
               d.address, d.budget
        FROM {SCHEMA}.deals d
        WHERE d.id=%s
    """, (deal_id,))
    row = cur.fetchone()
    if not row:
        return None, "Сделка не найдена"
    (did, client_id, project_type, kp_slot_id,
     contract_signed, payment_confirmed,
     cfg_id, sel_stages, buf_days, address, budget) = row

    if not contract_signed:
        return None, "Договор не отмечен как подписанный"
    if not payment_confirmed:
        return None, "Оплата не подтверждена"
    if not kp_slot_id:
        return None, "Слот не выбран"

    buf_days = buf_days or 7

    # Блокируем слот через SELECT FOR UPDATE
    cur.execute(f"""
        SELECT id, year, month, start_date, status, monthly_limit
        FROM {SCHEMA}.slots WHERE id=%s FOR UPDATE
    """, (kp_slot_id,))
    slot_row = cur.fetchone()
    if not slot_row:
        return None, "Слот не найден"
    s_id, s_year, s_month, s_start_date, s_status, s_limit = slot_row

    if s_status != 'free':
        return None, "Слот уже занят — выберите другой в настройках КП"

    # Проверяем лимит месяца
    cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.slots WHERE year=%s AND month=%s AND status IN ('booked','busy')", (s_year, s_month))
    if int(cur.fetchone()[0]) + 1 > s_limit:
        return None, f"Месяц перегружен (лимит {s_limit})"

    # Бронируем слот (атомарный UPDATE — FOR UPDATE держит блокировку)
    cur.execute(f"UPDATE {SCHEMA}.slots SET status='booked', deal_id=%s WHERE id=%s AND status='free' RETURNING id", (deal_id, kp_slot_id))
    if not cur.fetchone():
        return None, "Слот был занят параллельным запросом — выберите другой"

    slot_start = s_start_date if isinstance(s_start_date, date) else date.fromisoformat(str(s_start_date))
    start_for_project = slot_start.isoformat()

    # Обновляем сделку: slot_id = kp_slot_id, planned_start_date
    cur.execute(f"""
        UPDATE {SCHEMA}.deals
        SET slot_id=%s, planned_start_date=%s, stage='contract', updated_at=now()
        WHERE id=%s
    """, (kp_slot_id, start_for_project, deal_id))

    # Создаём проект
    project_id, perr = create_project_from_deal(cur, deal_id, client_id, start_for_project, kp_slot_id)
    if perr:
        return None, perr

    # Переводим в planning
    if project_id:
        cur.execute(f"""
            UPDATE {SCHEMA}.deals SET stage='planning', updated_at=now() WHERE id=%s
        """, (deal_id,))

    return {"deal_id": deal_id, "stage": "planning", "project_id": project_id, "slot_id": kp_slot_id}, None


def calc_commission_rate(closed_count: int) -> float:
    """Шкала комиссии риэлтора по количеству ранее закрытых сделок:
    0–4 → 3.0% (новичок)
    5–8 → 4.5% (в теме)
    9+  → 5.5% (профи)
    """
    if closed_count >= 9:
        return 5.5
    if closed_count >= 5:
        return 4.5
    return 3.0


def qualification_for_count(count: int) -> str:
    if count >= 9:
        return "pro"
    if count >= 5:
        return "inTopic"
    return "novice"


def close_deal_with_commission(cur, deal_id):
    """Перевод сделки в 'closed' с фиксацией комиссии риэлтора.
    - Идемпотентно: если сделка уже closed — комиссию не пересчитываем.
    - Считаем процент по числу ранее закрытых сделок риэлтора (без учёта текущей).
    - Инкрементируем staff.closed_deals_count и обновляем qualification.
    """
    cur.execute(f"""
        SELECT id, stage, realtor_id, COALESCE(budget, 0)
        FROM {SCHEMA}.deals WHERE id=%s
    """, (deal_id,))
    row = cur.fetchone()
    if not row:
        return None, "Сделка не найдена"
    _, current_stage, realtor_id, budget = row

    if current_stage == "closed":
        cur.execute(f"""
            UPDATE {SCHEMA}.deals SET updated_at=now() WHERE id=%s
            RETURNING id, stage, commission_rate, commission_amount
        """, (deal_id,))
        r = cur.fetchone()
        return {"id": r[0], "stage": r[1],
                "commission_rate": float(r[2]) if r[2] is not None else None,
                "commission_amount": float(r[3]) if r[3] is not None else None}, None

    rate = None
    amount = None
    if realtor_id:
        cur.execute(f"""
            SELECT COALESCE(closed_deals_count, 0) FROM {SCHEMA}.staff WHERE id=%s
        """, (realtor_id,))
        srow = cur.fetchone()
        prev_count = int(srow[0]) if srow else 0
        rate = calc_commission_rate(prev_count)
        amount = round(float(budget or 0) * rate / 100.0, 2)

        new_count = prev_count + 1
        new_qual = qualification_for_count(new_count)
        cur.execute(f"""
            UPDATE {SCHEMA}.staff
            SET closed_deals_count = %s, qualification = %s
            WHERE id = %s
        """, (new_count, new_qual, realtor_id))

    cur.execute(f"""
        UPDATE {SCHEMA}.deals
        SET stage='closed',
            commission_rate = %s,
            commission_amount = %s,
            closed_at = now(),
            updated_at = now()
        WHERE id=%s
        RETURNING id, stage
    """, (rate, amount, deal_id))
    r = cur.fetchone()
    return {"id": r[0], "stage": r[1],
            "commission_rate": rate,
            "commission_amount": amount}, None


def update_deal_stage(cur, deal_id, new_stage, body=None):
    """Обобщённое изменение стадии (для lost и других простых переходов)."""
    body = body or {}

    # Защита от обратного перехода из терминальных статусов.
    # Закрытая/сданная сделка не может вернуться в активные стадии —
    # иначе при повторном закрытии повторно начислится комиссия (двойной KPI).
    cur.execute(f"SELECT stage FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
    cur_row = cur.fetchone()
    if not cur_row:
        return None, "Сделка не найдена"
    current_stage = cur_row[0]
    TERMINAL = {"closed", "done"}
    if current_stage in TERMINAL and new_stage != current_stage:
        return None, "Сделка уже закрыта — изменение её статуса запрещено"

    if new_stage == "kp":
        return update_deal_kp(cur, deal_id, body)
    if new_stage == "contract":
        return update_deal_contract(cur, deal_id, body)
    if new_stage == "closed":
        return close_deal_with_commission(cur, deal_id)

    cur.execute(f"""
        UPDATE {SCHEMA}.deals SET stage=%s, updated_at=now() WHERE id=%s
        RETURNING id, stage
    """, (new_stage, deal_id))
    row = cur.fetchone()
    if not row:
        return None, "Сделка не найдена"
    return {"id": row[0], "stage": row[1]}, None

# ─── PROJECTS ────────────────────────────────────────────────────────────────

def create_project_from_deal(cur, deal_id, client_id, start_date, slot_id):
    """Создаёт проект и этапы из сделки. Нормативы берутся из stage_durations."""
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)

    # Проверить — нет ли уже проекта
    cur.execute(f"SELECT id FROM {SCHEMA}.projects WHERE deal_id=%s", (deal_id,))
    ex = cur.fetchone()
    if ex:
        return ex[0], None

    # Получаем данные сделки
    cur.execute(f"""
        SELECT d.project_type, d.configuration_id, d.selected_stages,
               d.buffer_days, d.planned_start_date, d.address,
               cfg.included_stages as cfg_stages,
               d.serial_project_id
        FROM {SCHEMA}.deals d
        LEFT JOIN {SCHEMA}.configurations cfg ON cfg.id = d.configuration_id
        WHERE d.id = %s
    """, (deal_id,))
    deal_row = cur.fetchone()
    if not deal_row:
        return None, "Сделка не найдена"

    project_type, cfg_id, sel_stages, buf_days, planned_start, address, cfg_stages, sp_id = deal_row
    buf_days = buf_days or 7

    # Для индивидуального проекта без сметы — не создаём автоматически
    if project_type == 'individual':
        return None, None

    # Определяем итоговый набор этапов
    # Приоритет: кастомный выбор → комплектация → все этапы
    if sel_stages:
        included_nums = list(sel_stages)
    elif cfg_stages:
        included_nums = list(cfg_stages)
    else:
        included_nums = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]

    # Нормативы из БД — единый источник правды
    stage_norms = get_stage_durations(cur)
    filtered_norms = [s for s in stage_norms if s["stage_num"] in included_nums]

    # Плановая дата старта
    if planned_start:
        start_date = planned_start if isinstance(planned_start, date) else date.fromisoformat(str(planned_start))

    # Строим Гант-план
    stage_plan = _build_plan_from_norms(start_date, filtered_norms)
    if stage_plan:
        total_duration = (max(p[4] for p in stage_plan) - start_date).days + 1 + buf_days
    else:
        total_duration = 62 + buf_days

    deadline = start_date + timedelta(days=total_duration)
    code = next_code(cur, "projects", "ДОМ")

    cur.execute(f"""
        INSERT INTO {SCHEMA}.projects (code, deal_id, client_id, start_date, deadline, status, address, slot_id)
        VALUES (%s, %s, %s, %s, %s, 'planning', %s, %s)
        RETURNING id
    """, (code, deal_id, client_id, start_date, deadline, address or "", slot_id))
    project_id = cur.fetchone()[0]

    # Разворачиваем этапы
    for order_num, (snum, name, duration, ps, pe, par_group, deps) in enumerate(stage_plan, 1):
        cur.execute(f"""
            INSERT INTO {SCHEMA}.project_stages
                (project_id, name, order_num, stage_num, duration_days,
                 planned_start, planned_end, parallel_group, depends_on, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending')
        """, (project_id, name, order_num, snum, duration, ps, pe,
              par_group, deps if deps else None))

    # Слот остаётся 'booked' — станет 'busy' только после утверждения директором по строительству

    # Привязать project_id к сделке
    cur.execute(f"""
        UPDATE {SCHEMA}.deals SET project_id=%s, updated_at=now()
        WHERE id=%s
    """, (project_id, deal_id))

    return project_id, None


def approve_project(cur, project_id: int):
    """
    Директор по строительству берёт проект в производство:
    - статус проекта: planning → active
    - статус слота: booked → busy
    - связанная сделка: planning → closed (с начислением комиссии риэлтору)
    Используем FOR UPDATE для защиты от race condition (двойное нажатие).
    """
    cur.execute(f"""
        SELECT p.id, p.status, p.slot_id, p.deal_id, s.status as slot_status
        FROM {SCHEMA}.projects p
        LEFT JOIN {SCHEMA}.slots s ON s.id = p.slot_id
        WHERE p.id = %s
        FOR UPDATE OF p
    """, (project_id,))
    row = cur.fetchone()
    if not row:
        return None, "Проект не найден"
    pid, pstatus, slot_id, deal_id, slot_status = row

    # Проверяем статус: планирование или active (для случая, когда проект
    # был создан старым кодом с status='active', но сделка ещё в planning)
    if pstatus not in ('planning', 'active'):
        return None, "Проект уже завершён или в архиве"

    # Проверяем: если сделка уже закрыта — повторное нажатие кнопки, ничего не делаем
    if deal_id:
        cur.execute(f"SELECT stage FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
        drow = cur.fetchone()
        if drow and drow[0] == 'closed':
            # Сделка уже закрыта — просто убеждаемся что проект active
            if pstatus != 'active':
                cur.execute(f"UPDATE {SCHEMA}.projects SET status='active', updated_at=now() WHERE id=%s", (project_id,))
            return {"project_id": project_id, "status": "active", "deal_closed": None}, None

    # Переводим проект в active
    cur.execute(f"""
        UPDATE {SCHEMA}.projects
        SET status='active', updated_at=now()
        WHERE id=%s
    """, (project_id,))

    # Переводим слот в busy (защита: только если booked)
    if slot_id:
        cur.execute(f"""
            UPDATE {SCHEMA}.slots
            SET status='busy'
            WHERE id=%s AND status='booked'
        """, (slot_id,))

    # Закрываем связанную сделку с начислением комиссии риэлтору
    deal_closed = None
    if deal_id:
        deal_closed, err = close_deal_with_commission(cur, deal_id)
        if err:
            deal_closed = {"error": err}

    return {
        "project_id": project_id,
        "status": "active",
        "deal_closed": deal_closed,
    }, None


def cancel_project(cur, project_id: int):
    """
    Расторжение договора / отмена проекта:
    - проект → archived
    - слот → free (освобождается для других сделок)
    - сделка → lost
    """
    cur.execute(f"""
        SELECT p.id, p.status, p.slot_id, p.deal_id, s.status as slot_status
        FROM {SCHEMA}.projects p
        LEFT JOIN {SCHEMA}.slots s ON s.id = p.slot_id
        WHERE p.id = %s FOR UPDATE
    """, (project_id,))
    row = cur.fetchone()
    if not row:
        return None, "Проект не найден"
    pid, pstatus, slot_id, deal_id, slot_status = row

    if pstatus in ('archived', 'completed'):
        return None, "Проект уже архивирован или завершён"

    # Архивируем проект
    cur.execute(f"""
        UPDATE {SCHEMA}.projects SET status='archived', updated_at=now() WHERE id=%s
    """, (project_id,))

    # Освобождаем слот
    if slot_id and slot_status in ('booked', 'busy'):
        cur.execute(f"""
            UPDATE {SCHEMA}.slots SET status='free', deal_id=NULL WHERE id=%s
        """, (slot_id,))

    # Сделку переводим в lost
    if deal_id:
        cur.execute(f"""
            UPDATE {SCHEMA}.deals SET stage='lost', updated_at=now() WHERE id=%s
        """, (deal_id,))

    return {"project_id": project_id, "status": "archived", "slot_freed": bool(slot_id)}, None


def complete_project(cur, project_id: int):
    """
    Завершение проекта (сдан):
    - проект → completed
    - слот → archived (не влияет на загрузку, не удаляется)
    """
    cur.execute(f"""
        SELECT p.id, p.status, p.slot_id, s.status as slot_status
        FROM {SCHEMA}.projects p
        LEFT JOIN {SCHEMA}.slots s ON s.id = p.slot_id
        WHERE p.id = %s FOR UPDATE
    """, (project_id,))
    row = cur.fetchone()
    if not row:
        return None, "Проект не найден"
    pid, pstatus, slot_id, slot_status = row

    if pstatus not in ('active', 'planning'):
        return None, "Можно завершить только активный проект"

    cur.execute(f"""
        UPDATE {SCHEMA}.projects SET status='completed', updated_at=now() WHERE id=%s
    """, (project_id,))

    if slot_id:
        cur.execute(f"""
            UPDATE {SCHEMA}.slots SET status='archived' WHERE id=%s
        """, (slot_id,))

    return {"project_id": project_id, "status": "completed"}, None


def get_projects(cur, archived=False):
    where = "WHERE p.status = 'archived'" if archived else "WHERE p.status != 'archived'"
    cur.execute(f"""
        SELECT p.id, p.code, p.start_date, p.deadline, p.status, p.brigade, p.total_cost,
               c.name as client_name, c.phone as client_phone,
               p.address,
               (SELECT COUNT(*) FROM {SCHEMA}.project_stages ps WHERE ps.project_id=p.id) as total_stages,
               (SELECT COUNT(*) FROM {SCHEMA}.project_stages ps WHERE ps.project_id=p.id AND ps.status='done') as done_stages,
               d.id as deal_id, d.code as deal_code, d.budget as deal_budget, d.signed_date, d.contract_status,
               d.stage as deal_stage, d.client_token,
               sm.name as manager_name,
               sp.name as serial_project_name,
               cfg.name as configuration_name,
               p.slot_id,
               s.status as slot_status,
               s.start_date as slot_start_date
        FROM {SCHEMA}.projects p
        LEFT JOIN {SCHEMA}.clients c ON c.id = p.client_id
        LEFT JOIN {SCHEMA}.deals d ON d.id = p.deal_id
        LEFT JOIN {SCHEMA}.staff sm ON sm.id = d.manager_id
        LEFT JOIN {SCHEMA}.serial_projects sp ON sp.id = d.serial_project_id
        LEFT JOIN {SCHEMA}.configurations cfg ON cfg.id = d.configuration_id
        LEFT JOIN {SCHEMA}.slots s ON s.id = p.slot_id
        {where}
        ORDER BY p.created_at DESC
        LIMIT 50
    """)
    cols = [desc[0] for desc in cur.description]
    projects = [dict(zip(cols, row)) for row in cur.fetchall()]

    for proj in projects:
        cur.execute(f"""
            SELECT id, name, order_num, duration_days, planned_start, planned_end,
                   actual_start, actual_end, status
            FROM {SCHEMA}.project_stages WHERE project_id=%s ORDER BY order_num
        """, (proj["id"],))
        scols = [desc[0] for desc in cur.description]
        proj["stages"] = [dict(zip(scols, r)) for r in cur.fetchall()]

        total = proj["total_stages"] or 0
        done = proj["done_stages"] or 0
        proj["progress"] = int((done / total * 100) if total > 0 else 0)

        if proj["deadline"] and proj["start_date"]:
            today = date.today()
            dl = proj["deadline"] if isinstance(proj["deadline"], date) else date.fromisoformat(str(proj["deadline"]))
            proj["days_left"] = (dl - today).days

    return projects

# ─── PROCUREMENT ─────────────────────────────────────────────────────────────

def get_procurement(cur):
    cur.execute(f"""
        SELECT mr.id, mr.code, mr.material, mr.quantity, mr.unit, mr.required_date,
               mr.priority, mr.status, mr.notes, mr.created_at,
               p.code as project_code, f.name as foreman_name
        FROM {SCHEMA}.material_requests mr
        LEFT JOIN {SCHEMA}.projects p ON p.id = mr.project_id
        LEFT JOIN {SCHEMA}.staff f ON f.id = mr.foreman_id
        ORDER BY mr.created_at DESC LIMIT 50
    """)
    cols = [desc[0] for desc in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]

def create_material_request(cur, body):
    project_id = int(body["project_id"])
    material = body["material"]
    quantity = float(body.get("quantity", 1))
    unit = body.get("unit", "шт")
    required_date = body.get("required_date", (date.today() + timedelta(days=3)).isoformat())
    priority = body.get("priority", "normal")
    foreman_id = body.get("foreman_id")
    notes = body.get("notes", "")

    code = next_code(cur, "material_requests", "ЗМ")
    foreman_val = int(foreman_id) if foreman_id else None

    cur.execute(f"""
        INSERT INTO {SCHEMA}.material_requests
            (code, project_id, material, quantity, unit, required_date, priority, foreman_id, notes, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'new')
        RETURNING id, code
    """, (code, project_id, material, quantity, unit, required_date, priority, foreman_val, notes))
    rid, rcode = cur.fetchone()
    return {"id": rid, "code": rcode}

def update_request_status(cur, req_id, new_status):
    cur.execute(f"""
        UPDATE {SCHEMA}.material_requests SET status=%s, updated_at=now() WHERE id=%s RETURNING id
    """, (new_status, req_id))
    return bool(cur.fetchone())

# ─── CLIENT PORTAL ────────────────────────────────────────────────────────────

def get_client_portal(cur, token: str):
    """Возвращает данные ЛК клиента по токену."""
    cur.execute(f"""
        SELECT d.id, d.code, d.stage, d.budget,
               COALESCE(p.address, d.address) as address,
               d.signed_date,
               d.client_token, d.project_id,
               c.name as client_name, c.phone as client_phone,
               p.status as project_status, p.start_date, p.deadline, p.code as project_code
        FROM {SCHEMA}.deals d
        LEFT JOIN {SCHEMA}.clients c ON c.id = d.client_id
        LEFT JOIN {SCHEMA}.projects p ON p.id = d.project_id
        WHERE d.client_token = %s
    """, (token,))
    row = cur.fetchone()
    if not row:
        return None
    cols = ["deal_id","deal_code","stage","budget","address","signed_date",
            "client_token","project_id","client_name","client_phone",
            "project_status","start_date","deadline","project_code"]
    deal = dict(zip(cols, row))

    # Этапы проекта
    stages = []
    today = date.today().isoformat()
    if deal["project_id"]:
        cur.execute(f"""
            SELECT id, name, order_num, planned_start, planned_end, actual_start, actual_end, status
            FROM {SCHEMA}.project_stages
            WHERE project_id = %s ORDER BY order_num, id
        """, (deal["project_id"],))
        scols = ["id","name","order_num","planned_start","planned_end","actual_start","actual_end","status"]
        rows = cur.fetchall()
        for r in rows:
            s = dict(zip(scols, r))
            # Вычисляем effective_status для фронтенда
            pe = s["planned_end"].isoformat() if s["planned_end"] else None
            if s["status"] == "done" or s["actual_end"]:
                s["effective_status"] = "done"
            elif s["status"] == "in_progress" or s["actual_start"]:
                s["effective_status"] = "in_progress"
                if pe and pe < today:
                    s["effective_status"] = "overdue"
            else:
                if pe and pe < today:
                    s["effective_status"] = "overdue"
                else:
                    s["effective_status"] = "pending"
            stages.append(s)

    # Акты
    cur.execute(f"""
        SELECT id, code, title, amount, status, signed_at, created_at
        FROM {SCHEMA}.client_acts WHERE deal_id = %s ORDER BY created_at
    """, (deal["deal_id"],))
    acols = ["id","code","title","amount","status","signed_at","created_at"]
    acts = [dict(zip(acols, r)) for r in cur.fetchall()]

    # Акты появляются только после явного создания через раздел «Строительство»

    # Платежи по сделке (история)
    budget = float(deal["budget"] or 0)
    cur.execute(f"""
        SELECT id, code, amount, category, payment_date, description
        FROM {SCHEMA}.payments
        WHERE deal_id=%s AND type='income'
        ORDER BY payment_date DESC, created_at DESC
    """, (deal["deal_id"],))
    pcols = ["id","code","amount","category","payment_date","description"]
    payments_history = [dict(zip(pcols, r)) for r in cur.fetchall()]

    # Полоса оплаты: только «Основной договор»
    paid_main = sum(float(p["amount"] or 0) for p in payments_history if p["category"] == "Основной договор")
    paid_extra = sum(float(p["amount"] or 0) for p in payments_history if p["category"] != "Основной договор")
    balance = round(budget - paid_main, 2)
    paid_pct = round(paid_main / budget * 100, 1) if budget > 0 else 0

    for p in payments_history:
        for k, v in p.items():
            if hasattr(v, 'isoformat'):
                p[k] = v.isoformat()

    for d in [deal] + stages + acts:
        for k, v in d.items():
            if hasattr(v, 'isoformat'):
                d[k] = v.isoformat()

    return {
        "deal": deal,
        "stages": stages,
        "acts": acts,
        "payments_history": payments_history,
        "paid_main": paid_main,
        "paid_extra": paid_extra,
        "balance": balance,
        "paid_pct": paid_pct,
        "budget": budget,
    }

def create_client_act(cur, project_id: int, stage_id: int, amount: float, title: str = ""):
    """Директор/прораб создаёт акт по этапу для подписания клиентом."""
    cur.execute(f"""
        SELECT d.id, ps.name, p.code
        FROM {SCHEMA}.projects p
        JOIN {SCHEMA}.deals d ON d.id = p.deal_id
        JOIN {SCHEMA}.project_stages ps ON ps.id = %s AND ps.project_id = p.id
        WHERE p.id = %s
    """, (stage_id, project_id))
    row = cur.fetchone()
    if not row:
        return None, "Проект или этап не найден"
    deal_id, stage_name, project_code = row
    act_title = title.strip() or f"Акт по этапу «{stage_name}»"
    act_code = "АКТ-" + "".join(random.choices(string.digits, k=4))
    cur.execute(f"""
        INSERT INTO {SCHEMA}.client_acts
            (code, deal_id, project_id, stage_id, title, amount, status)
        VALUES (%s, %s, %s, %s, %s, %s, 'pending_signature')
        RETURNING id, code, title, amount, status, signed_at, created_at
    """, (act_code, deal_id, project_id, stage_id, act_title, amount))
    row = cur.fetchone()
    cols = ["id","code","title","amount","status","signed_at","created_at"]
    act = dict(zip(cols, row))
    for k, v in act.items():
        if hasattr(v, 'isoformat'):
            act[k] = v.isoformat()
    return act, None

def sign_client_act(cur, act_id: int):
    """Клиент подписывает акт → статус signed. Если привязан stage_id — переводим платёж в paid."""
    cur.execute(f"""
        UPDATE {SCHEMA}.client_acts
        SET status='signed', signed_at=now()
        WHERE id=%s AND status='pending_signature'
        RETURNING id, stage_id
    """, (act_id,))
    row = cur.fetchone()
    if not row:
        return None, "Акт не найден или уже подписан"
    act_id_val, stage_id = row
    # Если у акта есть stage_id — переводим соответствующий платёж в paid
    if stage_id:
        cur.execute(f"""
            UPDATE {SCHEMA}.payment_schedule
            SET status='paid', updated_at=now()
            WHERE stage_id=%s AND status='pending'
        """, (stage_id,))
    return {"ok": True, "act_id": act_id_val}, None

# ─── PAYMENT SCHEDULE ─────────────────────────────────────────────────────────

def get_payment_schedule(cur, deal_id: int):
    """Возвращает график оплат по сделке."""
    cur.execute(f"""
        SELECT id, deal_id, order_index, stage_name, amount, status, stage_id, created_at
        FROM {SCHEMA}.payment_schedule
        WHERE deal_id=%s ORDER BY order_index, id
    """, (deal_id,))
    cols = ["id","deal_id","order_index","stage_name","amount","status","stage_id","created_at"]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for r in rows:
        for k, v in r.items():
            if hasattr(v, 'isoformat'):
                r[k] = v.isoformat()
    return rows

def upsert_payment_schedule(cur, deal_id: int, items: list):
    """Полная замена графика оплат для сделки (идемпотентно по deal_id)."""
    # Удаляем старые строки через UPDATE статуса невозможно — используем INSERT/UPDATE по id
    # Получаем текущие id
    cur.execute(f"SELECT id FROM {SCHEMA}.payment_schedule WHERE deal_id=%s", (deal_id,))
    existing_ids = {r[0] for r in cur.fetchall()}
    incoming_ids = {item["id"] for item in items if item.get("id")}

    # Удаляем строки которых нет в новом списке — через UPDATE status=deleted нельзя, просто пропускаем
    # На самом деле даём менеджеру явно управлять через add/update/delete actions
    result = []
    for item in items:
        iid = item.get("id")
        order_index = int(item.get("order_index", 1))
        stage_name  = item.get("stage_name", "")
        amount      = float(item.get("amount", 0))
        status      = item.get("status", "pending")
        stage_id    = item.get("stage_id") or None
        if iid and iid in existing_ids:
            cur.execute(f"""
                UPDATE {SCHEMA}.payment_schedule
                SET order_index=%s, stage_name=%s, amount=%s, status=%s, stage_id=%s, updated_at=now()
                WHERE id=%s AND deal_id=%s RETURNING id
            """, (order_index, stage_name, amount, status, stage_id, iid, deal_id))
        else:
            cur.execute(f"""
                INSERT INTO {SCHEMA}.payment_schedule (deal_id, order_index, stage_name, amount, status, stage_id)
                VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
            """, (deal_id, order_index, stage_name, amount, status, stage_id))
        row = cur.fetchone()
        if row:
            result.append(row[0])

    # Удаляем строки которых не было в списке (менеджер убрал)
    ids_to_keep = set(result) | incoming_ids
    for old_id in existing_ids - ids_to_keep:
        cur.execute(f"UPDATE {SCHEMA}.payment_schedule SET status='cancelled', updated_at=now() WHERE id=%s", (old_id,))

    return result

def ensure_client_token(cur, deal_id: int) -> str:
    """Генерирует и сохраняет client_token если ещё нет."""
    cur.execute(f"SELECT client_token FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
    row = cur.fetchone()
    if row and row[0]:
        return row[0]
    token = "CL-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=4)) + "-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=4))
    cur.execute(f"UPDATE {SCHEMA}.deals SET client_token=%s WHERE id=%s", (token, deal_id))
    return token

# ─── PAYMENTS ────────────────────────────────────────────────────────────────

def get_payments(cur):
    cur.execute(f"""
        SELECT py.id, py.code, py.type, py.category, py.amount, py.payment_date, py.description,
               p.code as project_code, d.code as deal_code,
               s.name as created_by_name
        FROM {SCHEMA}.payments py
        LEFT JOIN {SCHEMA}.projects p ON p.id = py.project_id
        LEFT JOIN {SCHEMA}.deals d ON d.id = py.deal_id
        LEFT JOIN {SCHEMA}.staff s ON s.id = py.created_by
        ORDER BY py.payment_date DESC, py.created_at DESC LIMIT 50
    """)
    cols = [desc[0] for desc in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]

def create_payment(cur, body):
    import math
    project_id = body.get("project_id")
    deal_id = body.get("deal_id")
    pay_type = body.get("type")  # income / expense
    if pay_type not in ("income", "expense"):
        raise ValueError("Тип платежа должен быть income или expense")
    category = (body.get("category") or "").strip()[:100]
    try:
        amount = float(body.get("amount") or 0)
    except (TypeError, ValueError):
        raise ValueError("Сумма платежа должна быть числом")
    if math.isnan(amount) or math.isinf(amount):
        raise ValueError("Сумма платежа некорректна")
    if amount <= 0:
        raise ValueError("Сумма платежа должна быть больше нуля")
    if amount > 1_000_000_000:
        raise ValueError("Сумма платежа превышает допустимый лимит")
    payment_date = body.get("payment_date") or date.today().isoformat()
    description = (body.get("description") or "")[:1000]
    created_by = body.get("created_by")

    # Хотя бы одна привязка должна быть
    if not project_id and not deal_id:
        raise ValueError("Платёж должен быть привязан к сделке или проекту")

    code = next_code(cur, "payments", "ПЛТ")
    proj_val = int(project_id) if project_id else None
    deal_val = int(deal_id) if deal_id else None
    cb_val = int(created_by) if created_by else None

    cur.execute(f"""
        INSERT INTO {SCHEMA}.payments
            (code, project_id, deal_id, type, category, amount, payment_date, description, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id, code
    """, (code, proj_val, deal_val, pay_type, category, amount, payment_date, description, cb_val))
    pid, pcode = cur.fetchone()
    return {"id": pid, "code": pcode, "amount": amount, "type": pay_type}

def get_pl_summary(cur):
    cur.execute(f"""
        SELECT
            type,
            category,
            SUM(amount) as total
        FROM {SCHEMA}.payments
        WHERE EXTRACT(YEAR FROM payment_date) = EXTRACT(YEAR FROM CURRENT_DATE)
          AND EXTRACT(MONTH FROM payment_date) = EXTRACT(MONTH FROM CURRENT_DATE)
        GROUP BY type, category
        ORDER BY type, category
    """)
    cols = [desc[0] for desc in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    income = sum(r["total"] for r in rows if r["type"] == "income")
    expense = sum(r["total"] for r in rows if r["type"] == "expense")
    return {
        "income": float(income),
        "expense": float(expense),
        "profit": float(income - expense),
        "margin": round((income - expense) / income * 100, 1) if income > 0 else 0,
        "rows": rows
    }

# ─── K_COMPANY ───────────────────────────────────────────────────────────────

def calculate_kcompany(cur):
    # Нормативы
    cur.execute(f"SELECT key, value FROM {SCHEMA}.system_norms WHERE key IN ('sales_plan_month','houses_plan_month','build_days','k_company_alert')")
    norms = {r[0]: float(r[1]) for r in cur.fetchall()}
    sales_plan = norms.get("sales_plan_month", 20000000)
    houses_plan = norms.get("houses_plan_month", 4)
    norm_days = norms.get("build_days", 62)
    alert_threshold = norms.get("k_company_alert", 0.80)

    today = date.today()

    # Продажи факт (договоры этого месяца)
    cur.execute(f"""
        SELECT COALESCE(SUM(budget),0) FROM {SCHEMA}.deals
        WHERE stage='contract'
          AND EXTRACT(YEAR FROM start_date)=EXTRACT(YEAR FROM CURRENT_DATE)
          AND EXTRACT(MONTH FROM start_date)=EXTRACT(MONTH FROM CURRENT_DATE)
    """)
    sales_fact = float(cur.fetchone()[0])

    # Сдано домов этого месяца
    cur.execute(f"""
        SELECT COUNT(*) FROM {SCHEMA}.projects
        WHERE status='done'
          AND EXTRACT(YEAR FROM updated_at)=EXTRACT(YEAR FROM CURRENT_DATE)
          AND EXTRACT(MONTH FROM updated_at)=EXTRACT(MONTH FROM CURRENT_DATE)
    """)
    houses_fact = int(cur.fetchone()[0])

    # Средняя длительность завершённых домов
    cur.execute(f"""
        SELECT COALESCE(AVG(EXTRACT(DAY FROM (updated_at - created_at))),{norm_days})
        FROM {SCHEMA}.projects WHERE status='done'
    """)
    avg_duration = float(cur.fetchone()[0])
    if avg_duration <= 0:
        avg_duration = norm_days

    # Оборачиваемость (упрощённо: отношение платежей к нормативу)
    cur.execute(f"""
        SELECT COALESCE(SUM(amount),0) FROM {SCHEMA}.payments
        WHERE type='income'
          AND EXTRACT(YEAR FROM payment_date)=EXTRACT(YEAR FROM CURRENT_DATE)
          AND EXTRACT(MONTH FROM payment_date)=EXTRACT(MONTH FROM CURRENT_DATE)
    """)
    turnover_fact = float(cur.fetchone()[0])
    turnover_norm = sales_plan

    k_sales = min((sales_fact / sales_plan), 1.0) if sales_plan > 0 else 0
    k_production = min((houses_fact / houses_plan), 1.0) if houses_plan > 0 else 0
    k_speed = min((norm_days / avg_duration), 1.0) if avg_duration > 0 else 0
    k_turnover = min((turnover_fact / turnover_norm), 1.0) if turnover_norm > 0 else 0

    k_total = round(k_sales * 0.4 + k_production * 0.3 + k_speed * 0.2 + k_turnover * 0.1, 4)

    # Сохранить в лог
    cur.execute(f"""
        INSERT INTO {SCHEMA}.k_company_log
            (calc_date, k_total, k_sales, k_production, k_speed, k_turnover,
             sales_fact, sales_plan, houses_fact, houses_plan, avg_duration_days, alert_sent)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT DO NOTHING
    """, (today, k_total, round(k_sales,4), round(k_production,4), round(k_speed,4), round(k_turnover,4),
          sales_fact, sales_plan, houses_fact, int(houses_plan), round(avg_duration,2),
          k_total < alert_threshold))

    return {
        "date": today.isoformat(),
        "k_total": k_total,
        "k_sales": round(k_sales, 4),
        "k_production": round(k_production, 4),
        "k_speed": round(k_speed, 4),
        "k_turnover": round(k_turnover, 4),
        "sales_fact": sales_fact,
        "sales_plan": sales_plan,
        "houses_fact": houses_fact,
        "houses_plan": int(houses_plan),
        "avg_duration_days": round(avg_duration, 1),
        "alert": k_total < alert_threshold,
    }

def get_kcompany_last(cur):
    cur.execute(f"""
        SELECT calc_date, k_total, k_sales, k_production, k_speed, k_turnover,
               sales_fact, sales_plan, houses_fact, houses_plan, avg_duration_days, alert_sent
        FROM {SCHEMA}.k_company_log ORDER BY calc_date DESC LIMIT 1
    """)
    row = cur.fetchone()
    if not row:
        return None
    cols = ["calc_date","k_total","k_sales","k_production","k_speed","k_turnover",
            "sales_fact","sales_plan","houses_fact","houses_plan","avg_duration_days","alert_sent"]
    return dict(zip(cols, row))

# ─── DASHBOARD ───────────────────────────────────────────────────────────────

def get_dashboard(cur):
    cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.deals WHERE stage NOT IN ('lost','done')")
    active_deals = int(cur.fetchone()[0])

    cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.projects WHERE status='active'")
    active_projects = int(cur.fetchone()[0])

    cur.execute(f"""
        SELECT COALESCE(SUM(amount),0) FROM {SCHEMA}.payments
        WHERE type='income'
          AND EXTRACT(YEAR FROM payment_date)=EXTRACT(YEAR FROM CURRENT_DATE)
          AND EXTRACT(MONTH FROM payment_date)=EXTRACT(MONTH FROM CURRENT_DATE)
    """)
    revenue_month = float(cur.fetchone()[0])

    cur.execute(f"""
        SELECT COUNT(*) FROM {SCHEMA}.material_requests WHERE status='new'
    """)
    pending_requests = int(cur.fetchone()[0])

    # K_company last
    k = get_kcompany_last(cur)

    return {
        "active_deals": active_deals,
        "active_projects": active_projects,
        "revenue_month": revenue_month,
        "pending_requests": pending_requests,
        "k_company": k,
    }

# ─── SERIAL PROJECTS & CONFIGURATIONS ────────────────────────────────────────

def get_serial_projects(cur):
    cur.execute(f"""
        SELECT sp.id, sp.name, sp.area_sqm, sp.base_price, sp.base_duration_days,
               sp.description, sp.is_active, sp.created_at,
               COUNT(cfg.id) AS config_count
        FROM {SCHEMA}.serial_projects sp
        LEFT JOIN {SCHEMA}.configurations cfg ON cfg.serial_project_id = sp.id AND cfg.is_active = TRUE
        WHERE sp.is_active = TRUE
        GROUP BY sp.id, sp.name, sp.area_sqm, sp.base_price, sp.base_duration_days,
                 sp.description, sp.is_active, sp.created_at
        ORDER BY sp.name
    """)
    cols = [desc[0] for desc in cur.description]
    projects = [dict(zip(cols, r)) for r in cur.fetchall()]
    # Нормативы для пересчёта duration_days комплектаций
    stage_norms = get_stage_durations(cur)
    for p in projects:
        cur.execute(f"""
            SELECT id, name, description, price_coefficient, duration_days, included_stages,
                   COALESCE(discount_pct, 0) as discount_pct,
                   discount_until,
                   COALESCE(is_popular, false) as is_popular
            FROM {SCHEMA}.configurations
            WHERE serial_project_id=%s AND is_active=TRUE
            ORDER BY price_coefficient
        """, (p["id"],))
        ccols = [desc[0] for desc in cur.description]
        cfgs = [dict(zip(ccols, r)) for r in cur.fetchall()]
        # Пересчитываем duration_days из актуальных нормативов
        for cfg in cfgs:
            included = list(cfg["included_stages"]) if cfg["included_stages"] else []
            if included:
                filtered = [s for s in stage_norms if s["stage_num"] in included]
                if filtered:
                    plan = _build_plan_from_norms(date(2000, 1, 1), filtered)
                    if plan:
                        max_end = max(pp[4] for pp in plan)
                        cfg["duration_days"] = (max_end - date(2000, 1, 1)).days + 1
        p["configurations"] = cfgs
    return projects

def get_configurations(cur, serial_project_id):
    """Возвращает комплектации. duration_days пересчитывается из актуальных нормативов stage_durations."""
    cur.execute(f"""
        SELECT id, name, description, price_coefficient, duration_days, included_stages,
               COALESCE(discount_pct, 0) as discount_pct,
               discount_until,
               COALESCE(is_popular, false) as is_popular
        FROM {SCHEMA}.configurations
        WHERE serial_project_id=%s AND is_active=TRUE
        ORDER BY price_coefficient
    """, (serial_project_id,))
    cols = [desc[0] for desc in cur.description]
    cfgs = [dict(zip(cols, r)) for r in cur.fetchall()]

    today = date.today()
    # Пересчитываем duration_days из актуальных нормативов (чтобы отражать изменения директора)
    stage_norms = get_stage_durations(cur)
    for cfg in cfgs:
        included = list(cfg["included_stages"]) if cfg["included_stages"] else []
        if included:
            filtered = [s for s in stage_norms if s["stage_num"] in included]
            if filtered:
                plan = _build_plan_from_norms(date(2000, 1, 1), filtered)
                if plan:
                    max_end = max(p[4] for p in plan)
                    cfg["duration_days"] = (max_end - date(2000, 1, 1)).days + 1
        # Если скидка истекла — не показываем
        du = cfg.get("discount_until")
        if du and du < today:
            cfg["discount_pct"] = 0
            cfg["discount_until"] = None
    return cfgs

def create_serial_project(cur, body):
    name  = body["name"]
    area  = float(body.get("area_sqm", 0))
    price = float(body["base_price"])
    dur   = int(body.get("base_duration_days", 62))
    desc  = body.get("description", "")
    cur.execute(f"""
        INSERT INTO {SCHEMA}.serial_projects (name, area_sqm, base_price, base_duration_days, description)
        VALUES (%s, %s, %s, %s, %s) RETURNING id
    """, (name, area, price, dur, desc))
    return {"id": cur.fetchone()[0]}

def create_configuration(cur, body):
    sp_id   = int(body["serial_project_id"])
    name    = body["name"]
    desc    = body.get("description", "")
    coeff   = float(body.get("price_coefficient", 1.0))
    dur     = int(body.get("duration_days", 62))
    stages  = body.get("included_stages", list(range(1, 12)))
    cur.execute(f"""
        INSERT INTO {SCHEMA}.configurations
            (serial_project_id, name, description, price_coefficient, duration_days, included_stages)
        VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
    """, (sp_id, name, desc, coeff, dur, stages))
    return {"id": cur.fetchone()[0]}

def update_configuration(cur, cfg_id: int, body: dict):
    """Директор устанавливает скидку и/или популярную метку."""
    sets, vals = [], []
    if "discount_pct" in body:
        sets.append("discount_pct=%s"); vals.append(float(body["discount_pct"]))
    if "discount_until" in body:
        sets.append("discount_until=%s"); vals.append(body["discount_until"] or None)
    if "is_popular" in body:
        sets.append("is_popular=%s"); vals.append(bool(body["is_popular"]))
    if not sets:
        return {"ok": False}
    vals.append(cfg_id)
    cur.execute(f"UPDATE {SCHEMA}.configurations SET {', '.join(sets)} WHERE id=%s", vals)
    return {"id": cfg_id, "ok": True}

# ─── INDIVIDUAL PROJECT REQUESTS ─────────────────────────────────────────────

def get_individual_requests(cur):
    cur.execute(f"""
        SELECT ipr.id, ipr.deal_id, ipr.desired_area, ipr.special_requests,
               ipr.status, ipr.design_deadline, ipr.estimate_file_url,
               ipr.created_at, ipr.updated_at,
               c.name as client_name,
               d.code as deal_code
        FROM {SCHEMA}.individual_project_requests ipr
        LEFT JOIN {SCHEMA}.clients c ON c.id = ipr.client_id
        LEFT JOIN {SCHEMA}.deals d ON d.id = ipr.deal_id
        ORDER BY ipr.created_at DESC
    """)
    cols = [desc[0] for desc in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]

def update_individual_request(cur, req_id, body):
    status   = body.get("status")
    designer = body.get("assigned_designer_id")
    deadline = body.get("design_deadline")
    url      = body.get("estimate_file_url")
    sets, vals = [], []
    if status:   sets.append("status=%s");                  vals.append(status)
    if designer: sets.append("assigned_designer_id=%s");    vals.append(int(designer))
    if deadline: sets.append("design_deadline=%s");         vals.append(deadline)
    if url:      sets.append("estimate_file_url=%s");       vals.append(url)
    if not sets:
        return False
    sets.append("updated_at=now()")
    vals.append(req_id)
    cur.execute(f"UPDATE {SCHEMA}.individual_project_requests SET {', '.join(sets)} WHERE id=%s", vals)
    return True

# ─── SLOT PLAN (управление лимитами) ─────────────────────────────────────────

def update_slot_limit(cur, year, month, new_limit):
    """Обновляет monthly_limit для всех слотов указанного месяца."""
    cur.execute(f"""
        UPDATE {SCHEMA}.slots SET monthly_limit=%s
        WHERE year=%s AND month=%s
    """, (int(new_limit), int(year), int(month)))
    return {"year": year, "month": month, "monthly_limit": new_limit}

# ─── ESTIMATE: WORKS & MATERIALS ─────────────────────────────────────────────

def get_estimate(cur, serial_project_id: int):
    """Возвращает полную смету проекта: работы + материалы по этапам."""
    # Нормативы этапов
    cur.execute(f"""
        SELECT stage_num, stage_name FROM {SCHEMA}.stage_durations
        WHERE stage_name NOT LIKE '_DISABLED_%%'
        ORDER BY sort_order, stage_num
    """)
    stages = {r[0]: r[1] for r in cur.fetchall()}

    # Работы
    cur.execute(f"""
        SELECT id, stage_num, work_name, unit, quantity, unit_price, notes, sort_order
        FROM {SCHEMA}.stage_works
        WHERE serial_project_id = %s
        ORDER BY stage_num, sort_order, id
    """, (serial_project_id,))
    wcols = [d[0] for d in cur.description]
    works = [dict(zip(wcols, r)) for r in cur.fetchall()]

    # Материалы
    cur.execute(f"""
        SELECT id, stage_num, material_name, unit, quantity, unit_price, supplier_hint, notes, sort_order
        FROM {SCHEMA}.stage_materials
        WHERE serial_project_id = %s
        ORDER BY stage_num, sort_order, id
    """, (serial_project_id,))
    mcols = [d[0] for d in cur.description]
    materials = [dict(zip(mcols, r)) for r in cur.fetchall()]

    # Группируем по этапам
    result = []
    for snum, sname in sorted(stages.items()):
        stage_works = [w for w in works if w["stage_num"] == snum]
        stage_mats  = [m for m in materials if m["stage_num"] == snum]
        works_total = sum(float(w["quantity"]) * float(w["unit_price"]) for w in stage_works)
        mats_total  = sum(float(m["quantity"]) * float(m["unit_price"]) for m in stage_mats)
        result.append({
            "stage_num":    snum,
            "stage_name":   sname,
            "works":        stage_works,
            "materials":    stage_mats,
            "works_total":  works_total,
            "mats_total":   mats_total,
            "stage_total":  works_total + mats_total,
        })

    grand_works = sum(float(w["unit_price"]) * float(w["quantity"]) for w in works)
    grand_mats  = sum(float(m["unit_price"]) * float(m["quantity"]) for m in materials)
    return {
        "stages": result,
        "total_works": grand_works,
        "total_materials": grand_mats,
        "grand_total": grand_works + grand_mats,
    }

def upsert_estimate_row(cur, table: str, body: dict):
    """Создаёт или обновляет строку сметы (работа или материал)."""
    row_id = body.get("id")
    sp_id  = int(body["serial_project_id"])
    snum   = int(body["stage_num"])
    sort   = int(body.get("sort_order", 0))

    if table == "stage_works":
        name  = body["work_name"]
        unit  = body.get("unit", "шт")
        qty   = float(body.get("quantity", 0))
        price = float(body.get("unit_price", 0))
        notes = body.get("notes", "")
        if row_id:
            cur.execute(f"""
                UPDATE {SCHEMA}.stage_works
                SET work_name=%s, unit=%s, quantity=%s, unit_price=%s, notes=%s, sort_order=%s, updated_at=now()
                WHERE id=%s AND serial_project_id=%s
                RETURNING id
            """, (name, unit, qty, price, notes, sort, int(row_id), sp_id))
        else:
            cur.execute(f"""
                INSERT INTO {SCHEMA}.stage_works
                    (serial_project_id, stage_num, work_name, unit, quantity, unit_price, notes, sort_order)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
            """, (sp_id, snum, name, unit, qty, price, notes, sort))
    else:  # stage_materials
        name     = body["material_name"]
        unit     = body.get("unit", "шт")
        qty      = float(body.get("quantity", 0))
        price    = float(body.get("unit_price", 0))
        supplier = body.get("supplier_hint", "")
        notes    = body.get("notes", "")
        if row_id:
            cur.execute(f"""
                UPDATE {SCHEMA}.stage_materials
                SET material_name=%s, unit=%s, quantity=%s, unit_price=%s, supplier_hint=%s, notes=%s, sort_order=%s, updated_at=now()
                WHERE id=%s AND serial_project_id=%s
                RETURNING id
            """, (name, unit, qty, price, supplier, notes, sort, int(row_id), sp_id))
        else:
            cur.execute(f"""
                INSERT INTO {SCHEMA}.stage_materials
                    (serial_project_id, stage_num, material_name, unit, quantity, unit_price, supplier_hint, notes, sort_order)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
            """, (sp_id, snum, name, unit, qty, price, supplier, notes, sort))
    row = cur.fetchone()
    return {"id": row[0]} if row else {"id": None}

# ─── NOTIFICATIONS ────────────────────────────────────────────────────────────

def create_notification(cur, type_: str, title: str, body_text: str = "",
                        role: str = None, staff_id: int = None, deal_id: int = None):
    cur.execute(f"""
        INSERT INTO {SCHEMA}.notifications (type, title, body, role, staff_id, deal_id)
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (type_, title, body_text, role, staff_id, deal_id))
    return cur.fetchone()[0]

def get_notifications(cur, role: str = None, staff_id: int = None, unread_only: bool = False):
    wheres, vals = ["1=1"], []
    if role:
        wheres.append("(n.role=%s OR n.role IS NULL)"); vals.append(role)
    if staff_id:
        wheres.append("(n.staff_id=%s OR n.staff_id IS NULL)"); vals.append(staff_id)
    if unread_only:
        wheres.append("n.is_read=FALSE")
    cur.execute(f"""
        SELECT n.id, n.type, n.title, n.body, n.role, n.staff_id,
               n.deal_id, n.is_read, n.created_at,
               d.code as deal_code
        FROM {SCHEMA}.notifications n
        LEFT JOIN {SCHEMA}.deals d ON d.id = n.deal_id
        WHERE {' AND '.join(wheres)}
        ORDER BY n.created_at DESC
        LIMIT 50
    """, vals)
    cols = [x[0] for x in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]

def mark_notifications_read(cur, notif_ids: list):
    if not notif_ids:
        return
    placeholders = ",".join(["%s"] * len(notif_ids))
    cur.execute(f"UPDATE {SCHEMA}.notifications SET is_read=TRUE WHERE id IN ({placeholders})", notif_ids)

def get_unread_count(cur, role: str = None, staff_id: int = None):
    wheres, vals = ["is_read=FALSE"], []
    if role:
        wheres.append("(role=%s OR role IS NULL)"); vals.append(role)
    if staff_id:
        wheres.append("(staff_id=%s OR staff_id IS NULL)"); vals.append(staff_id)
    cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.notifications WHERE {' AND '.join(wheres)}", vals)
    return int(cur.fetchone()[0])

# ─── S3 HELPER ───────────────────────────────────────────────────────────────

def get_s3():
    return boto3.client(
        "s3",
        endpoint_url="https://bucket.poehali.dev",
        aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"],
    )

def upload_file_to_s3(file_b64: str, file_name: str, folder: str = "documents") -> tuple:
    """Декодирует base64, заливает в S3, возвращает (cdn_url, size_kb)."""
    s3 = get_s3()
    data = base64.b64decode(file_b64)
    # Определяем Content-Type по расширению
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
    content_types = {
        "pdf": "application/pdf",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "doc": "application/msword",
        "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "png": "image/png",
    }
    ct = content_types.get(ext, "application/octet-stream")
    key = f"{folder}/{file_name}"
    s3.put_object(Bucket="files", Key=key, Body=data, ContentType=ct)
    cdn_url = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{key}"
    size_kb = len(data) // 1024
    return cdn_url, size_kb

# ─── DOC TEMPLATES (пакет документов по договору) ────────────────────────────

def get_doc_templates(cur, active_only: bool = True):
    """Список шаблонов документов (управляет директор)."""
    where = "WHERE is_active = TRUE" if active_only else ""
    cur.execute(f"""
        SELECT id, name, description, is_required, sort_order,
               file_url, file_name, file_size_kb, file_updated_at,
               is_active, created_at, version, prev_file_url, prev_file_name
        FROM {SCHEMA}.doc_templates
        {where}
        ORDER BY sort_order, id
    """)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]

def update_doc_template(cur, tpl_id: int, body: dict):
    """Директор обновляет метаданные шаблона (или загружает новый файл)."""
    sets, vals = [], []
    for field in ["name", "description", "is_required", "sort_order", "is_active"]:
        if field in body:
            sets.append(f"{field}=%s"); vals.append(body[field])
    # Если передан файл — заливаем в S3, старый файл сохраняем как prev
    file_b64 = body.get("file_b64")
    file_name = body.get("file_name")
    if file_b64 and file_name:
        # Сохраняем старый URL как prev_file
        cur.execute(f"SELECT file_url, file_name, version FROM {SCHEMA}.doc_templates WHERE id=%s", (tpl_id,))
        old = cur.fetchone()
        cdn_url, size_kb = upload_file_to_s3(file_b64, f"template_{tpl_id}_{file_name}", "doc_templates")
        sets += ["file_url=%s", "file_name=%s", "file_size_kb=%s", "file_updated_at=now()",
                 "version=version+1",
                 "prev_file_url=%s", "prev_file_name=%s"]
        vals += [cdn_url, file_name, size_kb, old[0] if old else None, old[1] if old else None]
    if not sets:
        return False
    sets.append("updated_at=now()")
    vals.append(tpl_id)
    cur.execute(f"UPDATE {SCHEMA}.doc_templates SET {', '.join(sets)} WHERE id=%s", vals)
    return True

def create_doc_template(cur, body: dict):
    """Директор создаёт новый шаблон."""
    name      = body["name"]
    desc      = body.get("description", "")
    required  = bool(body.get("is_required", True))
    sort      = int(body.get("sort_order", 99))
    cur.execute(f"""
        INSERT INTO {SCHEMA}.doc_templates (name, description, is_required, sort_order)
        VALUES (%s,%s,%s,%s) RETURNING id
    """, (name, desc, required, sort))
    tpl_id = cur.fetchone()[0]
    # Если сразу передан файл
    file_b64 = body.get("file_b64")
    file_name = body.get("file_name")
    if file_b64 and file_name:
        cdn_url, size_kb = upload_file_to_s3(file_b64, f"template_{tpl_id}_{file_name}", "doc_templates")
        cur.execute(f"""
            UPDATE {SCHEMA}.doc_templates
            SET file_url=%s, file_name=%s, file_size_kb=%s, file_updated_at=now()
            WHERE id=%s
        """, (cdn_url, file_name, size_kb, tpl_id))
    return {"id": tpl_id, "name": name}

# ─── CONTRACT DOCUMENTS (пакет по конкретной сделке) ─────────────────────────

def get_contract_docs(cur, deal_id: int):
    """
    Возвращает статус пакета документов по сделке:
    все шаблоны + загружены ли файлы менеджером + подписанные директором.
    """
    # Все активные шаблоны
    templates = get_doc_templates(cur, active_only=True)
    # Уже загруженные для этой сделки
    cur.execute(f"""
        SELECT id, template_id, file_url, file_name, file_size_kb,
               uploaded_at, status, notes,
               signed_file_url, signed_file_name, signed_at,
               payment_confirmed, manager_seen_signed, reject_reason
        FROM {SCHEMA}.contract_documents
        WHERE deal_id = %s
    """, (deal_id,))
    cols = [d[0] for d in cur.description]
    uploaded = {r[1]: dict(zip(cols, r)) for r in cur.fetchall()}

    result = []
    for tpl in templates:
        doc = uploaded.get(tpl["id"])
        result.append({
            "template_id":   tpl["id"],
            "template_name": tpl["name"],
            "description":   tpl["description"],
            "is_required":   tpl["is_required"],
            "sort_order":    tpl["sort_order"],
            "template_version": tpl.get("version", 1),
            # Шаблон для скачивания
            "template_file_url":  tpl["file_url"],
            "template_file_name": tpl["file_name"],
            # Загруженный менеджером файл
            "doc_id":     doc["id"]        if doc else None,
            "file_url":   doc["file_url"]  if doc else None,
            "file_name":  doc["file_name"] if doc else None,
            "status":     doc["status"]    if doc else "pending",
            "uploaded_at":doc["uploaded_at"] if doc else None,
            # Подписанный директором вариант
            "signed_file_url":  doc["signed_file_url"]  if doc else None,
            "signed_file_name": doc["signed_file_name"] if doc else None,
            "signed_at":        doc["signed_at"]         if doc else None,
            "manager_seen_signed": doc["manager_seen_signed"] if doc else False,
            "reject_reason":        doc["reject_reason"]       if doc else None,
        })

    all_required_uploaded = all(
        r["status"] in ("uploaded", "approved", "review")
        for r in result if r["is_required"]
    )
    # Последняя причина отклонения — из любого отклонённого документа
    last_reject_reason = next(
        (r["reject_reason"] for r in result if r.get("reject_reason")),
        None
    )
    return {
        "items": result,
        "all_required_done": all_required_uploaded,
        "total": len(result),
        "uploaded_count": sum(1 for r in result if r["status"] != "pending"),
        "last_reject_reason": last_reject_reason,
    }

def upload_contract_doc(cur, deal_id: int, template_id: int, body: dict):
    """Менеджер загружает подписанный скан документа."""
    file_b64  = body["file_b64"]
    file_name = body["file_name"]
    notes     = body.get("notes", "")

    cdn_url, size_kb = upload_file_to_s3(
        file_b64,
        f"deal_{deal_id}_tpl_{template_id}_{file_name}",
        "contract_docs"
    )

    # Upsert — если уже было, обновляем
    cur.execute(f"""
        INSERT INTO {SCHEMA}.contract_documents
            (deal_id, template_id, file_url, file_name, file_size_kb, uploaded_at, status, notes)
        VALUES (%s, %s, %s, %s, %s, now(), 'uploaded', %s)
        ON CONFLICT (deal_id, template_id) DO UPDATE SET
            file_url = EXCLUDED.file_url,
            file_name = EXCLUDED.file_name,
            file_size_kb = EXCLUDED.file_size_kb,
            uploaded_at = now(),
            status = 'uploaded',
            notes = EXCLUDED.notes
        RETURNING id
    """, (deal_id, template_id, cdn_url, file_name, size_kb, notes))
    doc_id = cur.fetchone()[0]

    # Автоматически создаём/обновляем запись в таблице documents
    cur.execute(f"""
        SELECT name FROM {SCHEMA}.doc_templates WHERE id=%s
    """, (template_id,))
    tpl_row = cur.fetchone()
    tpl_name = tpl_row[0] if tpl_row else "Документ"

    cur.execute(f"""
        SELECT d.code, c.name FROM {SCHEMA}.deals d
        LEFT JOIN {SCHEMA}.clients c ON c.id = d.client_id
        WHERE d.id = %s
    """, (deal_id,))
    deal_row = cur.fetchone()
    deal_code   = deal_row[0] if deal_row else ""
    client_name = deal_row[1] if deal_row else ""

    # Обновляем или создаём запись в общем архиве документов
    cur.execute(f"""
        INSERT INTO {SCHEMA}.documents
            (doc_type, category, title, status, deal_id, file_url, file_name, file_size_kb)
        VALUES ('deal_contract', 'deal', %s, 'signed', %s, %s, %s, %s)
        ON CONFLICT DO NOTHING
    """, (f"{tpl_name} — {client_name} ({deal_code})",
          deal_id, cdn_url, file_name, size_kb))

    return {"doc_id": doc_id, "file_url": cdn_url, "file_name": file_name}

def submit_docs_for_review(cur, deal_id: int):
    """
    Менеджер отправляет пакет документов на проверку директору (шаг 2→3).
    Устанавливает статус пакета 'docs_review', уведомляет директора.
    """
    # Проверяем что все обязательные загружены
    pkg = get_contract_docs(cur, deal_id)
    if not pkg["all_required_done"]:
        return None, "Загрузите все обязательные документы перед отправкой"

    # Переводим все uploaded → review
    cur.execute(f"""
        UPDATE {SCHEMA}.contract_documents
        SET status='review'
        WHERE deal_id=%s AND status='uploaded'
    """, (deal_id,))

    # Обновляем статус сделки
    cur.execute(f"""
        UPDATE {SCHEMA}.deals SET contract_status='docs_review', updated_at=now() WHERE id=%s
        RETURNING code, client_id
    """, (deal_id,))
    row = cur.fetchone()
    deal_code = row[0] if row else ""

    # Уведомление директору
    create_notification(cur,
        type_="docs_for_review",
        title=f"Документы на проверку: {deal_code}",
        body_text="Менеджер загрузил пакет документов и ожидает проверки.",
        role="director",
        deal_id=deal_id,
    )
    return {"deal_id": deal_id, "contract_status": "docs_review"}, None

def approve_docs(cur, deal_id: int, approved: bool, reject_reason: str = ""):
    """
    Директор подтверждает или отклоняет документы (шаг 3).
    approved=True → статус 'docs_approved' + уведомление менеджеру.
    approved=False → статус 'docs_rejected' + причина отклонения.
    """
    new_doc_status = "approved" if approved else "rejected"
    new_deal_status = "docs_approved" if approved else "docs_uploaded"

    cur.execute(f"""
        UPDATE {SCHEMA}.contract_documents
        SET status=%s, reviewed_at=now(), reject_reason=%s
        WHERE deal_id=%s AND status='review'
    """, (new_doc_status, reject_reason or None, deal_id))

    cur.execute(f"""
        UPDATE {SCHEMA}.deals
        SET contract_status=%s, updated_at=now(),
            last_reject_reason=%s
        WHERE id=%s
        RETURNING code
    """, (new_deal_status, (reject_reason or None) if not approved else None, deal_id))
    deal_code = (cur.fetchone() or [""])[0]

    # Находим менеджера сделки
    cur.execute(f"SELECT manager_id FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
    manager_row = cur.fetchone()

    if approved:
        create_notification(cur,
            type_="docs_approved",
            title=f"Документы подтверждены: {deal_code}",
            body_text="Директор подтвердил документы. Переходите к ожиданию оплаты.",
            role="crm_manager",
            staff_id=manager_row[0] if manager_row else None,
            deal_id=deal_id,
        )
    else:
        create_notification(cur,
            type_="docs_rejected",
            title=f"Документы отклонены: {deal_code}",
            body_text=f"Причина: {reject_reason}. Исправьте и загрузите повторно.",
            role="crm_manager",
            staff_id=manager_row[0] if manager_row else None,
            deal_id=deal_id,
        )
    return {"deal_id": deal_id, "contract_status": new_deal_status, "approved": approved}, None

def set_payment_pending(cur, deal_id: int):
    """
    Переводим сделку в ожидание оплаты (шаг 3→4).
    Вызывается директором после approve_docs или отдельно.
    """
    cur.execute(f"""
        UPDATE {SCHEMA}.deals SET contract_status='payment_pending', updated_at=now() WHERE id=%s
        RETURNING code, manager_id
    """, (deal_id,))
    row = cur.fetchone()
    if not row:
        return None, "Сделка не найдена"
    deal_code, manager_id = row

    create_notification(cur,
        type_="payment_pending",
        title=f"Ожидание оплаты: {deal_code}",
        body_text="Документы подтверждены. Ожидайте поступления оплаты от заказчика.",
        role="crm_manager",
        staff_id=manager_id,
        deal_id=deal_id,
    )
    return {"deal_id": deal_id, "contract_status": "payment_pending"}, None

def confirm_payment(cur, deal_id: int):
    """
    Директор подтверждает оплату:
    - слот (kp_slot_id) → booked (SELECT FOR UPDATE)
    - создаётся проект со статусом planning
    - сделка → stage=planning, contract_status=payment_confirmed
    Автоматически, без кнопки «Перевести в планирование» у менеджера.
    """
    cur.execute(f"""
        SELECT d.id, d.client_id, d.kp_slot_id, d.code, d.manager_id,
               d.configuration_id, d.selected_stages, d.buffer_days,
               d.address, d.budget, d.project_id
        FROM {SCHEMA}.deals d
        WHERE d.id=%s
    """, (deal_id,))
    row = cur.fetchone()
    if not row:
        return None, "Сделка не найдена"
    (did, client_id, kp_slot_id, deal_code, manager_id,
     cfg_id, sel_stages, buf_days, address, budget, existing_project_id) = row

    if not kp_slot_id:
        return None, "Слот не выбран — невозможно подтвердить оплату"

    # Блокируем и проверяем слот
    cur.execute(f"""
        SELECT id, year, month, start_date, status, monthly_limit
        FROM {SCHEMA}.slots WHERE id=%s FOR UPDATE
    """, (kp_slot_id,))
    slot_row = cur.fetchone()
    if not slot_row:
        return None, "Слот не найден"
    s_id, s_year, s_month, s_start_date, s_status, s_limit = slot_row

    if s_status not in ('free',):
        return None, "Слот уже занят — выберите другой в настройках КП"

    cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.slots WHERE year=%s AND month=%s AND status IN ('booked','busy')", (s_year, s_month))
    if int(cur.fetchone()[0]) + 1 > s_limit:
        return None, f"Месяц перегружен (лимит {s_limit})"

    # Бронируем слот
    cur.execute(f"UPDATE {SCHEMA}.slots SET status='booked', deal_id=%s WHERE id=%s", (deal_id, kp_slot_id))

    slot_start = s_start_date if isinstance(s_start_date, date) else date.fromisoformat(str(s_start_date))
    start_for_project = slot_start.isoformat()

    # Обновляем сделку: слот, дата начала, статус контракта
    cur.execute(f"""
        UPDATE {SCHEMA}.deals
        SET contract_status='payment_confirmed', payment_confirmed=true,
            slot_id=%s, planned_start_date=%s, updated_at=now()
        WHERE id=%s
    """, (kp_slot_id, start_for_project, deal_id))

    # Создаём проект если ещё не создан
    project_id = existing_project_id
    if not project_id:
        project_id, perr = create_project_from_deal(cur, deal_id, client_id, start_for_project, kp_slot_id)
        if perr:
            return None, perr

    # Переводим сделку в planning
    cur.execute(f"""
        UPDATE {SCHEMA}.deals SET stage='planning', updated_at=now() WHERE id=%s
    """, (deal_id,))

    create_notification(cur,
        type_="payment_confirmed",
        title=f"Оплата подтверждена: {deal_code}",
        body_text="Оплата получена! Проект автоматически создан в разделе «Строительство».",
        role="crm_manager",
        staff_id=manager_id,
        deal_id=deal_id,
    )
    create_notification(cur,
        type_="payment_confirmed",
        title=f"Новый проект готов к производству: {deal_code}",
        body_text="Оплата подтверждена, проект создан. Нажмите «Взять в производство».",
        role="construction_director",
        deal_id=deal_id,
    )
    return {"deal_id": deal_id, "contract_status": "payment_confirmed", "stage": "planning", "project_id": project_id, "slot_id": kp_slot_id}, None

def upload_signed_doc(cur, deal_id: int, template_id: int, body: dict):
    """Директор загружает подписанный вариант документа. UPSERT — создаёт запись если нет."""
    file_b64  = body["file_b64"]
    file_name = body["file_name"]

    cdn_url, size_kb = upload_file_to_s3(
        file_b64,
        f"deal_{deal_id}_signed_{template_id}_{file_name}",
        "contract_docs"
    )

    # UPSERT: если запись менеджера есть — обновляем signed_*, если нет — создаём новую
    cur.execute(f"""
        INSERT INTO {SCHEMA}.contract_documents
            (deal_id, template_id, signed_file_url, signed_file_name, signed_at, status, manager_seen_signed)
        VALUES (%s, %s, %s, %s, now(), 'review', false)
        ON CONFLICT (deal_id, template_id) DO UPDATE SET
            signed_file_url    = EXCLUDED.signed_file_url,
            signed_file_name   = EXCLUDED.signed_file_name,
            signed_at          = now(),
            manager_seen_signed = false
        RETURNING id
    """, (deal_id, template_id, cdn_url, file_name))
    row = cur.fetchone()
    if not row:
        return None, "Ошибка сохранения документа"

    # Уведомляем менеджера
    cur.execute(f"SELECT code, manager_id FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
    deal_row = cur.fetchone()
    if deal_row:
        deal_code, manager_id = deal_row
        create_notification(cur,
            type_="docs_signed_returned",
            title=f"Подписанный документ готов: {deal_code}",
            body_text="Директор подписал документы. Скачайте подписанные варианты.",
            role="crm_manager",
            staff_id=manager_id,
            deal_id=deal_id,
        )
    return {"doc_id": row[0], "signed_file_url": cdn_url}, None

def confirm_doc_payment(cur, deal_id: int, template_id: int):
    """Директор нажимает 'Оплата прошла' у конкретного документа."""
    cur.execute(f"""
        UPDATE {SCHEMA}.contract_documents
        SET payment_confirmed=true
        WHERE deal_id=%s AND template_id=%s
        RETURNING id
    """, (deal_id, template_id))
    if not cur.fetchone():
        return None, "Документ не найден"
    return {"ok": True}, None

def get_payout_deals(cur, manager_id: int = None):
    """Сделки с подтверждённой оплатой для заявок на вознаграждение."""
    where_mgr = f"AND d.manager_id = {int(manager_id)}" if manager_id else ""
    cur.execute(f"""
        SELECT d.id, d.code, d.budget, d.contract_status, d.signed_date,
               c.name AS client_name, c.phone AS client_phone,
               s.name AS manager_name,
               sp.name AS serial_project_name,
               d.project_id,
               pr.id AS payout_id, pr.status AS payout_status,
               pr.amount AS payout_amount, pr.requested_at, pr.notes,
               pr.invoice_file_url, pr.invoice_file_name, pr.reject_comment
        FROM {SCHEMA}.deals d
        LEFT JOIN {SCHEMA}.clients c ON c.id = d.client_id
        LEFT JOIN {SCHEMA}.staff s ON s.id = d.manager_id
        LEFT JOIN {SCHEMA}.serial_projects sp ON sp.id = d.serial_project_id
        LEFT JOIN {SCHEMA}.payout_requests pr ON pr.deal_id = d.id
            AND (pr.manager_id = d.manager_id)
        WHERE d.contract_status = 'payment_confirmed'
        {where_mgr}
        ORDER BY d.updated_at DESC
    """)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]

def create_payout_request(cur, deal_id: int, manager_id: int, body: dict):
    """Менеджер подаёт заявку на выплату вознаграждения (со счётом)."""
    amount    = body.get("amount")
    notes     = body.get("notes", "")
    file_b64  = body.get("invoice_file_b64")
    file_name = body.get("invoice_file_name")

    # Проверяем что нет активной заявки
    cur.execute(f"""
        SELECT id FROM {SCHEMA}.payout_requests
        WHERE deal_id=%s AND manager_id=%s AND status NOT IN ('rejected')
    """, (deal_id, manager_id))
    if cur.fetchone():
        return None, "Заявка по этой сделке уже подана"

    invoice_url = None
    if file_b64 and file_name:
        invoice_url, _ = upload_file_to_s3(file_b64, f"invoice_{deal_id}_{file_name}", "payouts")

    cur.execute(f"""
        INSERT INTO {SCHEMA}.payout_requests (deal_id, manager_id, amount, notes, status, invoice_file_url, invoice_file_name)
        VALUES (%s, %s, %s, %s, 'pending', %s, %s)
        RETURNING id
    """, (deal_id, manager_id, amount, notes, invoice_url, file_name if file_b64 else None))
    payout_id = cur.fetchone()[0]

    # Уведомляем директора
    cur.execute(f"SELECT code FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
    deal_code = (cur.fetchone() or [""])[0]
    create_notification(cur,
        type_="payout_requested",
        title=f"Счёт на оплату: {deal_code}",
        body_text=f"Менеджер загрузил счёт на вознаграждение{f' ₽{int(amount):,}' if amount else ''}. Требует согласования.",
        role="director",
        deal_id=deal_id,
    )
    return {"payout_id": payout_id, "invoice_url": invoice_url, "ok": True}, None

def update_payout_request(cur, payout_id: int, body: dict):
    """Директор одобряет/отклоняет заявку на выплату (с комментарием)."""
    status  = body["status"]  # approved | rejected
    comment = body.get("reject_comment", "")

    cur.execute(f"""
        UPDATE {SCHEMA}.payout_requests
        SET status=%s, reviewed_at=now(),
            reject_comment=CASE WHEN %s != '' THEN %s ELSE reject_comment END
        WHERE id=%s
        RETURNING deal_id, manager_id
    """, (status, comment, comment, payout_id))
    row = cur.fetchone()
    if not row:
        return None, "Заявка не найдена"
    deal_id, manager_id = row

    cur.execute(f"SELECT code FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
    deal_code = (cur.fetchone() or [""])[0]
    notif_type = "payout_approved" if status == "approved" else "payout_rejected"
    if status == "approved":
        notif_text = "Ваш счёт согласован! Ожидайте поступления оплаты."
    else:
        notif_text = f"Счёт отклонён.{f' Причина: {comment}' if comment else ' Уточните детали у директора.'}"
    create_notification(cur,
        type_=notif_type,
        title=f"Счёт {'согласован' if status == 'approved' else 'отклонён'}: {deal_code}",
        body_text=notif_text,
        role="crm_manager",
        staff_id=manager_id,
        deal_id=deal_id,
    )
    return {"ok": True, "status": status}, None

# ─── CONTRACTORS ─────────────────────────────────────────────────────────────

CONTRACTOR_TYPES = {
    "client":         "Заказчик",
    "supplier":       "Поставщик",
    "contractor":     "Подрядчик",
    "subcontractor":  "Субподрядчик",
    "internal":       "Внутренний",
    "general":        "Общий",
}

def get_contractors(cur, ctype: str = None):
    """Список контрагентов, опционально фильтр по типу."""
    if ctype:
        cur.execute(f"""
            SELECT id, contractor_type, name, inn, phone, email, contact_person,
                   legal_address, bank_name, bank_account, notes, is_active, created_at
            FROM {SCHEMA}.contractors
            WHERE contractor_type = %s AND is_active = TRUE
            ORDER BY name
        """, (ctype,))
    else:
        cur.execute(f"""
            SELECT id, contractor_type, name, inn, phone, email, contact_person,
                   legal_address, bank_name, bank_account, notes, is_active, created_at
            FROM {SCHEMA}.contractors
            WHERE is_active = TRUE
            ORDER BY contractor_type, name
        """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for r in rows:
        r["type_label"] = CONTRACTOR_TYPES.get(r["contractor_type"], r["contractor_type"])
    return rows

def create_contractor(cur, body: dict):
    """Создание нового контрагента."""
    ctype         = body.get("contractor_type", "client")
    name          = body["name"]
    inn           = body.get("inn", "")
    kpp           = body.get("kpp", "")
    legal_address = body.get("legal_address", "")
    actual_address= body.get("actual_address", "")
    phone         = body.get("phone", "")
    email         = body.get("email", "")
    contact       = body.get("contact_person", "")
    bank_name     = body.get("bank_name", "")
    bank_account  = body.get("bank_account", "")
    bik           = body.get("bik", "")
    corr_account  = body.get("corr_account", "")
    notes         = body.get("notes", "")
    cur.execute(f"""
        INSERT INTO {SCHEMA}.contractors
            (contractor_type, name, inn, kpp, legal_address, actual_address,
             phone, email, contact_person, bank_name, bank_account, bik, corr_account, notes)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id, name, contractor_type
    """, (ctype, name, inn, kpp, legal_address, actual_address,
          phone, email, contact, bank_name, bank_account, bik, corr_account, notes))
    row = cur.fetchone()
    return {"id": row[0], "name": row[1], "contractor_type": row[2]}

def update_contractor(cur, cid: int, body: dict):
    """Обновление контрагента."""
    sets, vals = [], []
    for field in ["contractor_type","name","inn","kpp","legal_address","actual_address",
                  "phone","email","contact_person","bank_name","bank_account","bik","corr_account","notes"]:
        if field in body:
            sets.append(f"{field}=%s"); vals.append(body[field])
    if not sets:
        return False
    sets.append("updated_at=now()")
    vals.append(cid)
    cur.execute(f"UPDATE {SCHEMA}.contractors SET {', '.join(sets)} WHERE id=%s", vals)
    return True

# ─── DOCUMENTS ────────────────────────────────────────────────────────────────

# Метки типов документов по категориям
DOC_TYPE_LABELS = {
    # Сделки
    "deal_kp":             "КП",
    "deal_contract":       "Договор подряда",
    "deal_act":            "Акт приёмки",
    "deal_supplement":     "Доп. соглашение",
    # Поставщики
    "supply_contract":     "Договор поставки",
    "supply_invoice":      "Счёт",
    "supply_upd":          "УПД",
    "supply_waybill":      "Товарная накладная",
    "supply_certificate":  "Сертификат",
    # Подрядчики / субподрядчики
    "contractor_contract": "Договор подряда",
    "ks2":                 "Акт КС-2",
    "ks3":                 "Справка КС-3",
    "contractor_invoice":  "Счёт подрядчика",
    "contractor_estimate": "Смета подрядчика",
    # Внутренние
    "internal_regulation": "Регламент",
    "internal_order":      "Приказ",
    "internal_hr":         "Должностная инструкция",
    # Общие/компания
    "company_license":     "Лицензия",
    "company_certificate": "Сертификат компании",
    "company_permit":      "Разрешение",
}

def get_documents(cur, category: str = None, deal_id: int = None,
                  contractor_id: int = None, project_id: int = None):
    """Получить документы с фильтрами."""
    wheres, vals = ["1=1"], []
    if category:
        wheres.append("d.category=%s"); vals.append(category)
    if deal_id:
        wheres.append("d.deal_id=%s"); vals.append(deal_id)
    if contractor_id:
        wheres.append("d.contractor_id=%s"); vals.append(contractor_id)
    if project_id:
        wheres.append("d.project_id=%s"); vals.append(project_id)

    cur.execute(f"""
        SELECT d.id, d.doc_type, d.category, d.title, d.status, d.amount,
               d.doc_date, d.deal_id, d.project_id, d.contractor_id,
               d.file_url, d.file_name, d.file_size_kb, d.notes, d.created_at,
               deal.code  AS deal_code,
               proj.code  AS project_code,
               cont.name  AS contractor_name,
               cont.contractor_type
        FROM {SCHEMA}.documents d
        LEFT JOIN {SCHEMA}.deals     deal ON deal.id = d.deal_id
        LEFT JOIN {SCHEMA}.projects  proj ON proj.id = d.project_id
        LEFT JOIN {SCHEMA}.contractors cont ON cont.id = d.contractor_id
        WHERE {' AND '.join(wheres)}
        ORDER BY d.created_at DESC
        LIMIT 200
    """, vals)
    cols = [x[0] for x in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for r in rows:
        r["doc_type_label"] = DOC_TYPE_LABELS.get(r["doc_type"], r["doc_type"])
    return rows

def create_document(cur, body: dict):
    """Создание документа вручную."""
    doc_type    = body["doc_type"]
    title       = body["title"]
    category    = body.get("category", _infer_category(doc_type))
    status      = body.get("status", "draft")
    amount      = body.get("amount")
    doc_date    = body.get("doc_date")
    deal_id     = body.get("deal_id")
    project_id  = body.get("project_id")
    contractor_id = body.get("contractor_id")
    file_url    = body.get("file_url")
    file_name   = body.get("file_name")
    file_size   = body.get("file_size_kb")
    notes       = body.get("notes", "")
    created_by  = body.get("created_by")

    cur.execute(f"""
        INSERT INTO {SCHEMA}.documents
            (doc_type, category, title, status, amount, doc_date,
             deal_id, project_id, contractor_id,
             file_url, file_name, file_size_kb, notes, created_by)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id, title, doc_type, category
    """, (doc_type, category, title, status,
          float(amount) if amount else None,
          doc_date, deal_id, project_id, contractor_id,
          file_url, file_name, file_size, notes, created_by))
    row = cur.fetchone()
    return {"id": row[0], "title": row[1], "doc_type": row[2], "category": row[3]}

def update_document_status(cur, doc_id: int, status: str, file_url: str = None, file_name: str = None):
    cur.execute(f"""
        UPDATE {SCHEMA}.documents
        SET status=%s,
            file_url=COALESCE(%s, file_url),
            file_name=COALESCE(%s, file_name),
            updated_at=now()
        WHERE id=%s
        RETURNING id, status
    """, (status, file_url, file_name, doc_id))
    row = cur.fetchone()
    return {"id": row[0], "status": row[1]} if row else None

def auto_create_deal_documents(cur, deal_id: int, stage: str, deal_data: dict):
    """
    Автоматически создаёт документы при переходе сделки на стадию.
    stage='kp'       → создаёт КП
    stage='contract' → создаёт Договор подряда
    """
    # Не создаём дубли
    cur.execute(f"""
        SELECT COUNT(*) FROM {SCHEMA}.documents
        WHERE deal_id=%s AND doc_type=%s
    """, (deal_id, "deal_kp" if stage == "kp" else "deal_contract"))
    if cur.fetchone()[0] > 0:
        return

    client_name = deal_data.get("client_name", "")
    deal_code   = deal_data.get("code", "")
    budget      = deal_data.get("budget")
    contractor_id = deal_data.get("contractor_id")

    if stage == "kp":
        cur.execute(f"""
            INSERT INTO {SCHEMA}.documents
                (doc_type, category, title, status, amount, deal_id, contractor_id)
            VALUES ('deal_kp', 'deal', %s, 'draft', %s, %s, %s)
        """, (f"КП — {client_name} ({deal_code})",
              float(budget) if budget else None,
              deal_id, contractor_id))

    elif stage == "contract":
        cur.execute(f"""
            INSERT INTO {SCHEMA}.documents
                (doc_type, category, title, status, amount, deal_id, contractor_id,
                 doc_date)
            VALUES ('deal_contract', 'deal', %s, 'signed', %s, %s, %s, CURRENT_DATE)
        """, (f"Договор подряда — {client_name} ({deal_code})",
              float(budget) if budget else None,
              deal_id, contractor_id))

def _infer_category(doc_type: str) -> str:
    if doc_type.startswith("deal_"):       return "deal"
    if doc_type.startswith("supply_"):     return "supply"
    if doc_type.startswith("contractor_") or doc_type in ("ks2","ks3"): return "contractor"
    if doc_type.startswith("internal_"):   return "internal"
    if doc_type.startswith("company_"):    return "general"
    return "general"

# ─── CLIENTS / STAFF ─────────────────────────────────────────────────────────

def get_clients(cur):
    """Возвращает клиентов из таблицы clients + заказчиков из contractors (тип client), дедуплицируя по имени+телефону."""
    cur.execute(f"SELECT id, name, phone, email, source FROM {SCHEMA}.clients ORDER BY name")
    cols = [desc[0] for desc in cur.description]
    from_clients = [dict(zip(cols, r)) for r in cur.fetchall()]

    # Добавляем заказчиков из contractors которых нет в clients
    cur.execute(f"""
        SELECT id, name, phone, email FROM {SCHEMA}.contractors
        WHERE contractor_type = 'client' AND is_active = TRUE
        ORDER BY name
    """)
    seen_names = {c["name"].strip().lower() for c in from_clients}
    for row in cur.fetchall():
        cname = (row[1] or "").strip()
        if cname.lower() not in seen_names:
            from_clients.append({
                "id": row[0],
                "name": cname,
                "phone": row[2] or "",
                "email": row[3] or "",
                "source": "Контрагенты",
            })
            seen_names.add(cname.lower())

    from_clients.sort(key=lambda x: x["name"])
    return from_clients

def create_client(cur, name: str, phone: str, email: str = "", source: str = "CRM"):
    """Создаёт нового клиента-заказчика из формы лида — сохраняет в clients И в contractors."""
    name  = name.strip()
    phone = phone.strip()
    if not name:
        return None, "Укажите ФИО"
    cur.execute(f"""
        INSERT INTO {SCHEMA}.clients (name, phone, email, source)
        VALUES (%s, %s, %s, %s)
        RETURNING id, name, phone, email, source
    """, (name, phone, email.strip(), source))
    row = cur.fetchone()
    cols = ["id","name","phone","email","source"]
    result = dict(zip(cols, row))

    # Дублируем в справочник контрагентов (тип «Клиент») если ещё не существует
    cur.execute(f"""
        SELECT id FROM {SCHEMA}.contractors
        WHERE contractor_type = 'client' AND lower(name) = lower(%s)
        LIMIT 1
    """, (name,))
    if not cur.fetchone():
        cur.execute(f"""
            INSERT INTO {SCHEMA}.contractors
                (contractor_type, name, phone, email, contact_person, notes)
            VALUES ('client', %s, %s, %s, %s, %s)
        """, (name, phone, email.strip(), name, "Создан из лида/сделки"))

    return result, None

def get_staff(cur, role_filter=None):
    cols_sql = ("id, name, role, "
                "COALESCE(closed_deals_count, 0) AS closed_deals_count, "
                "COALESCE(qualification, 'novice') AS qualification")
    if role_filter:
        cur.execute(f"SELECT {cols_sql} FROM {SCHEMA}.staff WHERE role=%s ORDER BY name", (role_filter,))
    else:
        cur.execute(f"SELECT {cols_sql} FROM {SCHEMA}.staff ORDER BY name")
    cols = [desc[0] for desc in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def get_realtors_report(cur):
    """Отчёт для коммерческого директора: топ риэлторов.
    Для каждого активного риэлтора возвращаем агрегаты по сделкам:
    - количество закрытых сделок (точная цифра из deals, не из staff-кэша),
    - общая сумма закрытых сделок (revenue),
    - общая зафиксированная комиссия,
    - количество открытых сделок и их сумма,
    - текущая квалификация и до следующего уровня.
    """
    cur.execute(f"""
        SELECT
            s.id,
            s.name,
            COALESCE(s.qualification, 'novice')              AS qualification,
            COALESCE(s.closed_deals_count, 0)                AS closed_deals_count,
            COUNT(d.id) FILTER (WHERE d.stage = 'closed')    AS closed_count,
            COUNT(d.id) FILTER (WHERE d.stage <> 'closed' AND d.is_archived = FALSE) AS open_count,
            COALESCE(SUM(d.budget) FILTER (WHERE d.stage = 'closed'), 0)             AS closed_revenue,
            COALESCE(SUM(d.commission_amount) FILTER (WHERE d.stage = 'closed'), 0) AS commission_total,
            COALESCE(SUM(d.budget) FILTER (WHERE d.stage <> 'closed' AND d.is_archived = FALSE), 0) AS open_revenue
        FROM {SCHEMA}.staff s
        LEFT JOIN {SCHEMA}.deals d ON d.realtor_id = s.id
        WHERE s.role = 'realtor' AND COALESCE(s.is_active, TRUE) = TRUE
        GROUP BY s.id, s.name, s.qualification, s.closed_deals_count
        ORDER BY commission_total DESC, closed_count DESC, s.name
    """)
    cols = [desc[0] for desc in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    # Считаем "до следующего уровня"
    for row in rows:
        cnt = int(row["closed_deals_count"] or 0)
        if cnt >= 9:
            row["next_level"] = None
            row["to_next"]    = 0
            row["next_rate"]  = None
        elif cnt >= 5:
            row["next_level"] = "pro"
            row["to_next"]    = 9 - cnt
            row["next_rate"]  = 5.5
        else:
            row["next_level"] = "inTopic"
            row["to_next"]    = 5 - cnt
            row["next_rate"]  = 4.5

    # Суммарные итоги
    totals = {
        "realtors":         len(rows),
        "closed_total":     sum(int(r["closed_count"] or 0) for r in rows),
        "revenue_total":    float(sum(float(r["closed_revenue"] or 0) for r in rows)),
        "commission_total": float(sum(float(r["commission_total"] or 0) for r in rows)),
    }
    return {"realtors": rows, "totals": totals}

# ─── EMPLOYEES ───────────────────────────────────────────────────────────────

def get_employees(cur):
    cur.execute(f"SELECT id, name, role FROM {SCHEMA}.staff ORDER BY role, name")
    cols = [desc[0] for desc in cur.description]
    employees = [dict(zip(cols, r)) for r in cur.fetchall()]

    ROLE_DEPT = {
        "crm_manager": "Продажи", "realtor": "Продажи",
        "foreman": "Строительство", "quality": "Качество",
        "mechanic": "Техника", "supplier": "Снабжение",
        "accountant": "Финансы", "director": "Руководство",
        "commercial": "Руководство", "construction_director": "Строительство",
        "supply_director": "Снабжение", "finance_director": "Финансы",
        "project_manager": "Строительство",
    }

    for emp in employees:
        emp["dept"] = ROLE_DEPT.get(emp["role"], "Прочее")
        # Сделки сотрудника (для менеджеров/риэлторов)
        if emp["role"] in ("crm_manager", "realtor"):
            cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.deals WHERE manager_id=%s OR realtor_id=%s", (emp["id"], emp["id"]))
            emp["deals_count"] = int(cur.fetchone()[0])
            cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.deals WHERE (manager_id=%s OR realtor_id=%s) AND stage='contract'", (emp["id"], emp["id"]))
            contracts = int(cur.fetchone()[0])
            emp["contracts_count"] = contracts
            total = emp["deals_count"] or 1
            emp["kpi"] = min(round((contracts / total) * 100 + 50), 99)
        elif emp["role"] == "foreman":
            cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.projects WHERE brigade ILIKE %s AND status='active'", (f"%{emp['name'].split()[0]}%",))
            emp["active_projects"] = int(cur.fetchone()[0])
            emp["kpi"] = 85
            emp["deals_count"] = None
        else:
            emp["deals_count"] = None
            emp["kpi"] = 80

    return employees

def create_employee(cur, body):
    name = body["name"]
    role = body["role"]
    cur.execute(f"INSERT INTO {SCHEMA}.staff (name, role) VALUES (%s, %s) RETURNING id, name, role", (name, role))
    row = cur.fetchone()
    return {"id": row[0], "name": row[1], "role": row[2]}

# ─── REPORTS ─────────────────────────────────────────────────────────────────

def get_reports(cur):
    # --- Менеджеры CRM ---
    cur.execute(f"""
        SELECT s.id, s.name, s.role,
               COUNT(d.id) AS leads,
               COUNT(CASE WHEN d.stage='contract' THEN 1 END) AS contracts,
               COALESCE(SUM(CASE WHEN d.stage='contract' THEN d.budget ELSE 0 END),0) AS revenue
        FROM {SCHEMA}.staff s
        LEFT JOIN {SCHEMA}.deals d ON d.manager_id = s.id OR d.realtor_id = s.id
        WHERE s.role IN ('crm_manager','realtor')
        GROUP BY s.id, s.name, s.role
        ORDER BY contracts DESC, revenue DESC
    """)
    cols = [desc[0] for desc in cur.description]
    managers = []
    for r in cur.fetchall():
        row = dict(zip(cols, r))
        total = row["leads"] or 1
        row["conversion"] = round(row["contracts"] / total * 100, 1)
        row["kpi"] = min(round(row["conversion"] + 50), 99)
        managers.append(row)

    # --- Бригады (прорабы) ---
    cur.execute(f"""
        SELECT s.id, s.name,
               COUNT(p.id) AS total_projects,
               COUNT(CASE WHEN p.status='done' THEN 1 END) AS done_projects,
               COALESCE(AVG(CASE WHEN p.status='done'
                   THEN EXTRACT(DAY FROM (p.updated_at - p.created_at)) END), 0) AS avg_days
        FROM {SCHEMA}.staff s
        LEFT JOIN {SCHEMA}.projects p ON p.brigade ILIKE '%' || split_part(s.name,' ',1) || '%'
        WHERE s.role = 'foreman'
        GROUP BY s.id, s.name
        ORDER BY done_projects DESC
    """)
    cols = [desc[0] for desc in cur.description]
    brigades = [dict(zip(cols, r)) for r in cur.fetchall()]
    for b in brigades:
        b["avg_days"] = round(float(b["avg_days"]), 1) if b["avg_days"] else 0
        on_time = b["done_projects"]
        total = b["total_projects"] or 1
        b["rating"] = round(3.5 + (on_time / total) * 1.4, 1)

    # --- KPI метрики ---
    # active_deals: только активные сделки в работе (не архив, не отказ, не закрытые)
    # contracts: те, у которых дошли до планирования или закрылись (реально подписан договор)
    cur.execute(f"""
        SELECT
            COUNT(*) FILTER (WHERE is_archived = FALSE AND stage NOT IN ('lost','closed')) AS active_deals,
            COUNT(*) FILTER (WHERE stage IN ('planning','closed') OR contract_status = 'signed') AS contracts,
            COUNT(*) FILTER (WHERE is_archived = FALSE) AS total_deals
        FROM {SCHEMA}.deals
    """)
    d = cur.fetchone()
    active_deals, contracts, total_deals = d
    conversion = round(contracts / total_deals * 100, 1) if total_deals > 0 else 0

    cur.execute(f"""
        SELECT COUNT(*) FILTER (WHERE status='active') AS active,
               COUNT(*) FILTER (WHERE status='done')   AS done,
               COALESCE(AVG(CASE WHEN status='done'
                   THEN EXTRACT(DAY FROM (updated_at - created_at)) END),0) AS avg_dur
        FROM {SCHEMA}.projects
    """)
    p = cur.fetchone()
    active_proj, done_proj, avg_dur = p
    avg_dur = round(float(avg_dur), 1) if avg_dur else 62.0

    cur.execute(f"""
        SELECT
            COALESCE(SUM(CASE WHEN type='income'  THEN amount END),0) AS income,
            COALESCE(SUM(CASE WHEN type='expense' THEN amount END),0) AS expense
        FROM {SCHEMA}.payments
        WHERE EXTRACT(YEAR FROM payment_date)  = EXTRACT(YEAR  FROM CURRENT_DATE)
          AND EXTRACT(MONTH FROM payment_date) = EXTRACT(MONTH FROM CURRENT_DATE)
    """)
    fin = cur.fetchone()
    income, expense = float(fin[0]), float(fin[1])
    margin = round((income - expense) / income * 100, 1) if income > 0 else 0

    kpis = [
        {"name": "Конверсия лид → договор",  "value": f"{conversion}%",  "target": "15%",  "status": "success" if conversion >= 15 else "warning" if conversion >= 8 else "error",  "trend": f"{total_deals} сделок всего"},
        {"name": "Средний срок сдачи дома",  "value": f"{avg_dur} дн.",  "target": "62 дн.", "status": "success" if avg_dur <= 62 else "warning" if avg_dur <= 68 else "error",  "trend": f"{done_proj} домов сдано"},
        {"name": "Маржинальность (месяц)",   "value": f"{margin}%",      "target": "35%",  "status": "success" if margin >= 35 else "warning" if margin >= 20 else "error",  "trend": f"₽{income:,.0f} доходов"},
        {"name": "Загрузка (активных домов)","value": str(active_proj),  "target": "4+",   "status": "success" if active_proj >= 4 else "warning", "trend": f"план 4 дома/мес"},
    ]

    # --- Поставщики (заявки) ---
    cur.execute(f"""
        SELECT status, COUNT(*) AS cnt,
               SUM(quantity) AS total_qty
        FROM {SCHEMA}.material_requests
        GROUP BY status
    """)
    req_stats = {r[0]: {"count": int(r[1]), "qty": float(r[2] or 0)} for r in cur.fetchall()}

    return {
        "managers": managers,
        "brigades": brigades,
        "kpis": kpis,
        "req_stats": req_stats,
        "summary": {
            "active_deals": int(active_deals),
            "contracts": int(contracts),
            "conversion": conversion,
            "active_projects": int(active_proj),
            "avg_duration": avg_dur,
            "income": income,
            "expense": expense,
            "margin": margin,
        }
    }

# ─── GANTT ────────────────────────────────────────────────────────────────────

def get_gantt_stages(cur, project_id: int):
    """Возвращает иерархию Гант-этапов проекта: группы + подэтапы."""
    cur.execute(f"""
        SELECT id, project_id, parent_id, name, order_num, stage_num,
               planned_start, planned_end, actual_start, actual_end,
               status, progress_percent, group_name, duration_days
        FROM {SCHEMA}.project_stages
        WHERE project_id = %s
        ORDER BY COALESCE(parent_id, id), order_num, id
    """, (project_id,))
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    today = date.today().isoformat()
    for r in rows:
        for k, v in r.items():
            if hasattr(v, 'isoformat'):
                r[k] = v.isoformat()
        # Вычисляем отклонение
        pe = r.get("planned_end")
        ae = r.get("actual_end")
        if ae and pe:
            delta = (date.fromisoformat(ae) - date.fromisoformat(pe)).days
            r["deviation_days"] = delta
            r["deviation_label"] = "Опережение" if delta < 0 else ("Отставание" if delta > 0 else "По плану")
        elif not ae and pe and pe < today and r.get("status") not in ("done",):
            delta = (date.today() - date.fromisoformat(pe)).days
            r["deviation_days"] = delta
            r["deviation_label"] = "Отставание"
        else:
            r["deviation_days"] = 0
            r["deviation_label"] = None

    # Группируем: сначала родительские (parent_id IS NULL), потом вложенные
    groups = {r["id"]: {**r, "children": []} for r in rows if r["parent_id"] is None}
    orphans = []
    for r in rows:
        if r["parent_id"] is not None:
            if r["parent_id"] in groups:
                groups[r["parent_id"]]["children"].append(r)
            else:
                orphans.append(r)

    # Усреднённый прогресс для группы
    for g in groups.values():
        children = g["children"]
        if children:
            g["progress_percent"] = int(sum(c["progress_percent"] for c in children) / len(children))

    result = list(groups.values()) + orphans
    result.sort(key=lambda x: x["order_num"])
    return result


def update_stage_progress(cur, stage_id: int, progress: int):
    """Обновляет прогресс этапа (0-100). При 100% ставит actual_end=сегодня и status=done."""
    if progress < 0 or progress > 100:
        return None, "Прогресс должен быть от 0 до 100"
    progress = (progress // 25) * 25  # Округляем до ближайших 25%

    today = date.today()
    new_status = "done" if progress == 100 else ("in_progress" if progress > 0 else "pending")
    actual_end = today if progress == 100 else None
    actual_start_sql = ""

    cur.execute(f"""
        SELECT id, status, actual_start, parent_id, project_id
        FROM {SCHEMA}.project_stages WHERE id=%s
    """, (stage_id,))
    row = cur.fetchone()
    if not row:
        return None, "Этап не найден"
    _, cur_status, cur_actual_start, parent_id, project_id = row

    # Фиксируем actual_start при первом изменении прогресса с 0
    if progress > 0 and not cur_actual_start:
        actual_start_sql = ", actual_start=%s"

    if actual_start_sql:
        cur.execute(f"""
            UPDATE {SCHEMA}.project_stages
            SET progress_percent=%s, status=%s, actual_end=%s {actual_start_sql}, updated_at=now()
            WHERE id=%s
            RETURNING id, progress_percent, status, actual_start, actual_end
        """, (progress, new_status, actual_end, today, stage_id))
    else:
        cur.execute(f"""
            UPDATE {SCHEMA}.project_stages
            SET progress_percent=%s, status=%s, actual_end=%s, updated_at=now()
            WHERE id=%s
            RETURNING id, progress_percent, status, actual_start, actual_end
        """, (progress, new_status, actual_end, stage_id))

    r = cur.fetchone()
    if not r:
        return None, "Не удалось обновить этап"
    result = {"id": r[0], "progress_percent": r[1], "status": r[2],
              "actual_start": r[3].isoformat() if r[3] else None,
              "actual_end": r[4].isoformat() if r[4] else None}

    # Если это подэтап — пересчитываем прогресс родительской группы
    if parent_id:
        cur.execute(f"""
            SELECT AVG(progress_percent) FROM {SCHEMA}.project_stages
            WHERE parent_id=%s
        """, (parent_id,))
        avg_row = cur.fetchone()
        if avg_row and avg_row[0] is not None:
            parent_progress = int(avg_row[0])
            parent_status = "done" if parent_progress == 100 else ("in_progress" if parent_progress > 0 else "pending")
            cur.execute(f"""
                UPDATE {SCHEMA}.project_stages
                SET progress_percent=%s, status=%s, updated_at=now()
                WHERE id=%s
            """, (parent_progress, parent_status, parent_id))

    return result, None


def add_gantt_group(cur, project_id: int, body: dict):
    """Добавляет группу этапов (parent_id IS NULL) в Гант-план проекта."""
    name = (body.get("name") or "").strip()
    if not name:
        return None, "Укажите название группы"

    planned_start = body.get("planned_start")
    planned_end   = body.get("planned_end")
    duration      = int(body.get("duration_days", 7))

    if planned_start and not planned_end:
        ps = date.fromisoformat(planned_start)
        planned_end = (ps + timedelta(days=duration)).isoformat()

    cur.execute(f"""
        SELECT COALESCE(MAX(order_num), 0) + 1 FROM {SCHEMA}.project_stages
        WHERE project_id=%s AND parent_id IS NULL
    """, (project_id,))
    order_num = cur.fetchone()[0]

    cur.execute(f"""
        INSERT INTO {SCHEMA}.project_stages
            (project_id, name, order_num, duration_days, planned_start, planned_end,
             status, progress_percent, group_name)
        VALUES (%s, %s, %s, %s, %s, %s, 'pending', 0, %s)
        RETURNING id
    """, (project_id, name, order_num, duration, planned_start, planned_end, name))
    new_id = cur.fetchone()[0]
    return {"id": new_id, "name": name, "order_num": order_num}, None


def delete_gantt_stage(cur, stage_id: int):
    """Удаляет этап или группу из Гант-плана.
    Если это группа (parent_id IS NULL) — удаляет также все её подэтапы.
    Не позволяет удалять этапы в статусе done.
    """
    cur.execute(f"""
        SELECT id, parent_id, status, name FROM {SCHEMA}.project_stages WHERE id=%s
    """, (stage_id,))
    row = cur.fetchone()
    if not row:
        return None, "Этап не найден"
    sid, parent_id, status, name = row
    if status == "done":
        return None, f"Нельзя удалить завершённый этап «{name}»"

    # Если это группа — удаляем подэтапы
    if parent_id is None:
        cur.execute(f"""
            SELECT COUNT(*) FROM {SCHEMA}.project_stages
            WHERE parent_id=%s AND status='done'
        """, (stage_id,))
        done_children = int(cur.fetchone()[0])
        if done_children > 0:
            return None, f"Нельзя удалить группу «{name}»: в ней есть завершённые подэтапы"
        cur.execute(f"UPDATE {SCHEMA}.project_stages SET parent_id=NULL WHERE parent_id=%s", (stage_id,))
        cur.execute(f"DELETE FROM {SCHEMA}.project_stages WHERE parent_id=%s", (stage_id,))

    cur.execute(f"DELETE FROM {SCHEMA}.project_stages WHERE id=%s", (stage_id,))
    return {"deleted": stage_id, "name": name}, None


def add_gantt_substage(cur, project_id: int, body: dict):
    """Добавляет подэтап к группе (parent_id указан)."""
    parent_id = body.get("parent_id")
    if not parent_id:
        return None, "parent_id обязателен"
    name = (body.get("name") or "").strip()
    if not name:
        return None, "Укажите название"

    planned_start = body.get("planned_start")
    planned_end   = body.get("planned_end")
    duration      = int(body.get("duration_days", 5))

    if planned_start and not planned_end:
        ps = date.fromisoformat(planned_start)
        planned_end = (ps + timedelta(days=duration)).isoformat()

    cur.execute(f"""
        SELECT COALESCE(MAX(order_num), 0) + 1 FROM {SCHEMA}.project_stages
        WHERE project_id=%s AND parent_id=%s
    """, (project_id, int(parent_id)))
    order_num = cur.fetchone()[0]

    cur.execute(f"""
        INSERT INTO {SCHEMA}.project_stages
            (project_id, parent_id, name, order_num, duration_days, planned_start, planned_end,
             status, progress_percent)
        VALUES (%s, %s, %s, %s, %s, %s, %s, 'pending', 0)
        RETURNING id
    """, (project_id, int(parent_id), name, order_num, duration, planned_start, planned_end))
    new_id = cur.fetchone()[0]
    return {"id": new_id, "name": name, "parent_id": int(parent_id), "order_num": order_num}, None


# ─── INVOICE AI RECOGNITION ──────────────────────────────────────────────────

CHATGPT_URL = "https://functions.poehali.dev/778ceb38-0039-4da4-9a48-0cb34a7527cf"

ALLOWED_EXTS = {'jpg','jpeg','png','pdf','xls','xlsx'}

# ── Промпты ──────────────────────────────────────────────────────────────────

_PROMPT_HEADER = """Извлеки из текста шапки счёта следующие поля и верни ТОЛЬКО JSON-объект:
{"supplier_name":"Название поставщика или null","invoice_date":"YYYY-MM-DD или null","invoice_number":"Номер счёта или null"}
Текст шапки:
{header}"""

_PROMPT_TABLE_CHUNK = """Ты получаешь фрагмент таблицы товаров из счёта. Извлеки ВСЕ позиции.
Для каждой строки товара верни объект. Игнорируй строки итогов, НДС, заголовков, пустые строки.
Не придумывай данные — если поле неизвестно, ставь null.

Ответ — ТОЛЬКО JSON-массив:
[{{"material":"Полное наименование товара","unit":"шт|м3|т|пог.м|м2|компл","unit_price":число_или_null,"quantity":число_или_null}}]

unit_price и quantity — только числа, без символов валюты и единиц.

Фрагмент таблицы:
{chunk}"""

_PROMPT_OCR_IMAGE = """Извлеки ВЕСЬ текст с этого изображения счёта или накладной.
Не пропускай ни одной строки. Сохраняй структуру таблицы: строки разделяй переводом строки, столбцы — двумя пробелами или табуляцией.
Если значение нечёткое — всё равно попытайся его прочитать. Верни только текст без пояснений."""

_PROMPT_IMAGE_DIRECT = """Ты получаешь изображение или файл счёта. Твоя задача — извлечь данные абсолютно точно.

ВАЖНО ПРО ОФОРМЛЕНИЕ:
- Если видишь QR-код, штрих-код или логотип компании — полностью игнорируй их. Они не являются частью данных счёта.
- В начале документа (шапке) обычно указаны: поставщик, дата, номер счёта — ищи их там.
- Таблица с позициями содержит: №, наименование, количество, единицу, цену, сумму.

ИЗВЛЕКИ:
- supplier_name: название поставщика (из шапки, над таблицей или в первых строках)
- invoice_date: дата в формате YYYY-MM-DD
- invoice_number: номер счёта
- footer_total: итоговая сумма счёта (общая сумма внизу документа)
- items: КАЖДАЯ строка таблицы — ровно одна позиция, ни одной не пропускай

ДЛЯ КАЖДОЙ ПОЗИЦИИ:
- material: ПОЛНОЕ название включая размеры (например "Доска 50*200*6000 мм"). Копируй цифры точно, цифра за цифрой.
- quantity: число (может быть дробным, например 3.37)
- unit: единица (м3, шт, м, пог.м, м2, компл и т.д.)
- unit_price: цена за единицу
- amount: сумма строки (unit_price × quantity)

ПРАВИЛА:
1. Каждая строка таблицы = РОВНО ОДНА позиция. Никогда не объединяй строки.
2. Копируй названия и размеры дословно, не перефразируй.
3. Числа — только цифры (без символов валюты, единиц, пробелов-разделителей).
4. Если значение неразборчиво — ставь null, не придумывай.
5. Для каждой позиции проверь: unit_price * quantity ≈ amount (допуск 1%).
   Если не совпадает — всё равно включи строку, но добавь поле "sum_check": false.
6. QR-коды, печати и логотипы — игнорируй полностью.

Верни строго JSON без комментариев:
{"supplier_name":"...","invoice_date":"YYYY-MM-DD","invoice_number":"...","footer_total":число,"items":[{"material":"...","quantity":число,"unit":"...","unit_price":число,"amount":число,"sum_check":true}]}"""

_PROMPT_IMAGE_CORRECTION = """Итоговая сумма позиций не совпала с общей суммой счёта.

Общая сумма счёта (из документа): {footer_total}
Сумма всех позиций в твоём ответе: {items_total}

Перепроверь каждую строку счёта на изображении:
- Убедись, что quantity и unit_price считаны правильно (особенно дробные числа).
- Убедись, что ты не пропустил ни одной строки и не продублировал строки.
- Уточни названия материалов, если они отличаются от оригинала.

Верни полностью исправленный JSON в том же формате:
{"supplier_name":"...","invoice_date":"...","invoice_number":"...","footer_total":число,"items":[...]}"""

_ITEM_KEY_ALIASES = {
    "supplier_name":  ["supplier", "vendor", "company", "поставщик", "организация", "from", "seller"],
    "material":       ["product", "item", "goods", "description", "наименование", "товар", "услуга", "name"],
    "unit":           ["uom", "measure", "единица", "ед_изм", "ед.изм"],
    "unit_price":     ["price", "cost", "цена", "стоимость", "price_per_unit", "rate"],
    "quantity":       ["qty", "кол_во", "количество", "count", "volume"],
    "invoice_date":   ["date", "дата", "issue_date", "doc_date"],
    "invoice_number": ["number", "num", "номер", "invoice_no", "doc_number", "#"],
}
_VALID_UNITS = ['шт','м3','т','пог.м','м2','компл','кг','л','уп','рул','пач','компл','г','мл']


def _normalize_obj(obj: dict) -> dict:
    low = {k.lower().strip(): v for k, v in obj.items()}
    result = {}
    for canonical, aliases in _ITEM_KEY_ALIASES.items():
        if canonical in low:
            result[canonical] = low[canonical]
        else:
            for alias in aliases:
                if alias in low:
                    result[canonical] = low[alias]
                    break
    for k, v in low.items():
        if k not in result:
            result[k] = v
    return result


def _is_null(v) -> bool:
    return v is None or str(v).strip().lower() in ("", "null", "none")


def _match_or_create_supplier(cur, name: str):
    if not name or _is_null(name):
        return None, False
    cur.execute(
        f"SELECT id FROM {SCHEMA}.suppliers WHERE lower(name)=lower(%s) AND name!='(не указан)' LIMIT 1",
        (name,)
    )
    row = cur.fetchone()
    if not row:
        cur.execute(
            f"SELECT id FROM {SCHEMA}.suppliers WHERE name ILIKE %s AND name!='(не указан)' LIMIT 1",
            (f"%{name}%",)
        )
        row = cur.fetchone()
    if row:
        return row[0], False
    cur.execute(
        f"INSERT INTO {SCHEMA}.suppliers (name, category) VALUES (%s,'прочее') RETURNING id",
        (name,)
    )
    return cur.fetchone()[0], True


def _match_or_create_material(cur, name: str, raw_unit: str):
    if not name or _is_null(name):
        return None, False, 'шт'
    # Используем unit из AI как есть, дефолт "шт" только если пусто
    unit = raw_unit.strip() if raw_unit and raw_unit.strip() and raw_unit.strip().lower() not in ('null','none','') else 'шт'
    cur.execute(
        f"SELECT id FROM {SCHEMA}.materials WHERE lower(name)=lower(%s) AND name!='(не указан)' LIMIT 1",
        (name,)
    )
    row = cur.fetchone()
    if not row:
        cur.execute(
            f"SELECT id FROM {SCHEMA}.materials WHERE name ILIKE %s AND name!='(не указан)' LIMIT 1",
            (f"%{name}%",)
        )
        row = cur.fetchone()
    if row:
        return row[0], False, unit
    cur.execute(
        f"INSERT INTO {SCHEMA}.materials (name, unit) VALUES (%s,%s) RETURNING id",
        (name, unit)
    )
    return cur.fetchone()[0], True, unit


def _parse_ai_response(raw_content: str):
    import re
    json_str = raw_content.strip()
    json_str = re.sub(r'```(?:json)?\s*', '', json_str)
    json_str = re.sub(r'\s*```', '', json_str).strip()
    try:
        return json.loads(json_str), None
    except json.JSONDecodeError as e1:
        m = re.search(r'\{[\s\S]+\}', json_str)
        if m:
            try:
                return json.loads(m.group(0)), None
            except Exception as e2:
                return {}, f"Err1:{e1} Err2:{e2} Raw:{json_str[:300]}"
        return {}, f"JSON не найден. Raw:{json_str[:300]}"


def _pdf_to_jpg_b64(file_bytes: bytes) -> str:
    """
    Конвертирует PDF в JPEG.
    Пытается через PyMuPDF (fitz) если доступен, иначе возвращает None — 
    тогда PDF будет передан в Gemini напрямую как application/pdf.
    """
    try:
        import fitz, base64, io as _io
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        page = doc[0]
        mat = fitz.Matrix(150 / 72, 150 / 72)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        from PIL import Image as _PIL
        img = _PIL.frombytes("RGB", [pix.width, pix.height], pix.samples)
        MAX = 1400
        w, h = img.size
        if w > MAX or h > MAX:
            scale = MAX / max(w, h)
            img = img.resize((int(w * scale), int(h * scale)), _PIL.LANCZOS)
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=88)
        doc.close()
        return base64.b64encode(buf.getvalue()).decode("utf-8")
    except ImportError:
        # PyMuPDF не установлен — вернём None, PDF пойдёт напрямую в Gemini
        return None


def _excel_to_jpg_b64(file_bytes: bytes, ext: str) -> str:
    """Читает Excel и отрисовывает таблицу в JPEG через Pillow. Возвращает base64."""
    import base64, io as _io
    from PIL import Image as _PIL, ImageDraw as _Draw, ImageFont as _Font

    # ── Читаем строки Excel ───────────────────────────────────────────────────
    if ext == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active or wb.worksheets[0]
        rows_raw = []
        for row in ws.iter_rows(max_row=100, values_only=True):
            rows_raw.append([str(c) if c is not None else "" for c in row])
        wb.close()
    else:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        rows_raw = []
        for ri in range(min(ws.nrows, 100)):
            rows_raw.append([str(ws.cell_value(ri, ci)) for ci in range(min(ws.ncols, 15))])

    # Убираем полностью пустые строки, ограничиваем 12 колонок
    rows = [r[:12] for r in rows_raw if any(str(c).strip() for c in r)]
    if not rows:
        rows = [["(пустой файл)"]]

    ncols = max(len(r) for r in rows)
    rows  = [r + [""] * (ncols - len(r)) for r in rows]

    # ── Шрифт ─────────────────────────────────────────────────────────────────
    FONT_SZ = 14
    PAD     = 6
    font    = None
    for font_path in (
        "/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/DejaVuSans.ttf",
    ):
        try:
            font = _Font.truetype(font_path, FONT_SZ)
            break
        except Exception:
            continue
    if font is None:
        font = _Font.load_default()

    # ── Автоподбор ширины колонок по содержимому ──────────────────────────────
    CHAR_W = 8.5     # пикс. на символ для 14px моноширинного
    MIN_W  = ([50, 300] + [90]  * max(0, ncols - 2))[:ncols]
    MAX_W  = ([80, 520] + [200] * max(0, ncols - 2))[:ncols]

    col_widths = []
    for ci in range(ncols):
        max_chars = max((len(str(row[ci])) for row in rows), default=0)
        cw = int(max_chars * CHAR_W) + PAD * 2
        col_widths.append(max(MIN_W[ci], min(MAX_W[ci], cw)))

    ROW_H = FONT_SZ + PAD * 2 + 2
    IMG_W = sum(col_widths) + 2
    IMG_H = len(rows) * ROW_H + 2

    img  = _PIL.new("RGB", (IMG_W, IMG_H), "#ffffff")
    draw = _Draw.Draw(img)

    for ri, row in enumerate(rows):
        y  = ri * ROW_H
        bg = "#e8eef5" if ri % 2 == 0 else "#ffffff"
        draw.rectangle([0, y, IMG_W, y + ROW_H], fill=bg)
        x = 1
        for ci, cell in enumerate(row):
            cw       = col_widths[ci]
            text     = str(cell)
            max_chars = max(1, int((cw - PAD * 2) / CHAR_W))
            if len(text) > max_chars:
                text = text[:max_chars - 1] + "…"
            draw.text((x + PAD, y + PAD), text, fill="#111827", font=font)
            draw.line([(x + cw - 1, y), (x + cw - 1, y + ROW_H)], fill="#c7d2da")
            x += cw
        draw.line([(0, y + ROW_H - 1), (IMG_W, y + ROW_H - 1)], fill="#c7d2da")

    # Масштабируем только если шире 2400px (Gemini хорошо читает крупные изображения)
    MAX_PX = 2400
    w, h   = img.size
    if w > MAX_PX or h > MAX_PX:
        scale = MAX_PX / max(w, h)
        img   = img.resize((int(w * scale), int(h * scale)), _PIL.LANCZOS)

    buf = _io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def upload_invoice_file(cur, invoice_id: int, file_b64: str, file_name: str):
    """
    Загружает файл счёта в S3.
    PDF и Excel автоматически конвертируются в JPG на бэкенде перед сохранением.
    """
    import base64 as _b64
    ext = file_name.rsplit('.', 1)[-1].lower() if '.' in file_name else ''
    if ext not in ALLOWED_EXTS:
        return None, f"Неподдерживаемый формат. Разрешены: {', '.join(sorted(ALLOWED_EXTS))}"

    file_bytes = _b64.b64decode(file_b64)

    # ── Конвертируем PDF → JPG (если fitz доступен), иначе сохраняем as-is ────
    if ext == "pdf":
        try:
            jpg_b64 = _pdf_to_jpg_b64(file_bytes)
            if jpg_b64 is not None:
                file_b64  = jpg_b64
                file_name = file_name[:-4] + ".jpg"
            # else: fitz недоступен — сохраняем PDF, Gemini прочитает напрямую
        except Exception as e:
            return None, f"Не удалось обработать PDF: {e}"

    # Excel — сохраняем as-is, recognize_invoice обработает через DeepSeek V4 Pro
    # (конвертация в JPG убрана — она мешала работе TSV-пайплайна)

    # Уникальное имя: inv_{id}_{name}, папка invoices/
    safe_name = f"inv_{invoice_id}_{file_name}"
    cdn_url, _ = upload_file_to_s3(file_b64, safe_name, "invoices")
    cur.execute(f"""
        UPDATE {SCHEMA}.invoices
        SET pdf_file_url=%s, pdf_file_name=%s, updated_at=now()
        WHERE id=%s
        RETURNING id
    """, (cdn_url, file_name, invoice_id))
    if not cur.fetchone():
        return None, "Счёт не найден"
    return {"cdn_url": cdn_url, "file_name": file_name}, None


# ── Утилиты ──────────────────────────────────────────────────────────────────

def _call_polza(req_lib, messages: list, max_tokens: int = 4096) -> str:
    """Вызов Polza.ai. Возвращает строку-ответ модели."""
    model = "google/gemini-3.1-flash-lite"
    payload = {"messages": messages, "model": model,
               "temperature": 0.0, "max_tokens": max_tokens}
    logger.info(f"_call_polza: model={model} messages_count={len(messages)} max_tokens={max_tokens}")
    resp = req_lib.post(
        f"{CHATGPT_URL}?action=generate",
        json=payload,
        timeout=120
    )
    if not resp.ok:
        err_text = ""
        try: err_text = resp.json()
        except Exception: err_text = resp.text[:500]
        logger.error(f"_call_polza error {resp.status_code}: {err_text}")
    resp.raise_for_status()
    return resp.json().get("content", "")


def _parse_json_list(raw: str):
    """
    Надёжный парсинг ответа модели в список позиций.
    Если ответ — объект с полем "items" — возвращает items[].
    Возвращает (list | None, error_str | None).
    """
    import re
    s = raw.strip()
    s = re.sub(r'```(?:json)?\s*', '', s)
    s = re.sub(r'```', '', s).strip()

    # 1. Прямой json.loads
    try:
        r = json.loads(s)
        # Если объект с items — извлекаем items (новый формат AI)
        if isinstance(r, dict):
            if "items" in r and isinstance(r["items"], list):
                return r["items"], None
            # Объект без items — единичная позиция
            return [r], None
        if isinstance(r, list):
            return r, None
    except json.JSONDecodeError:
        pass

    # 2. Жадный поиск [...] — самый длинный блок
    for m in sorted(re.finditer(r'\[[\s\S]+\]', s), key=lambda x: -len(x.group(0))):
        try:
            r = json.loads(m.group(0))
            if isinstance(r, list): return r, None
        except json.JSONDecodeError:
            continue

    # 3. Жадный поиск {...} с items
    for m in sorted(re.finditer(r'\{[\s\S]+\}', s), key=lambda x: -len(x.group(0))):
        try:
            r = json.loads(m.group(0))
            if isinstance(r, dict):
                if "items" in r and isinstance(r["items"], list):
                    return r["items"], None
                return [r], None
        except json.JSONDecodeError:
            continue

    return None, f"JSON не найден. Фрагмент: {s[:400]}"


def _parse_json_obj(raw: str) -> dict:
    """Парсит ответ как JSON-объект. При ошибке возвращает {}."""
    import re
    s = raw.strip()
    s = re.sub(r'```(?:json)?\s*', '', s)
    s = re.sub(r'```', '', s).strip()
    try:
        r = json.loads(s)
        return r if isinstance(r, dict) else {}
    except json.JSONDecodeError:
        for m in sorted(re.finditer(r'\{[\s\S]+\}', s), key=lambda x: -len(x.group(0))):
            try:
                r = json.loads(m.group(0))
                if isinstance(r, dict): return r
            except json.JSONDecodeError:
                continue
    return {}


def _safe_float(v) -> float | None:
    """Конвертирует значение в float, убирает разделители тысяч."""
    if v is None: return None
    sv = str(v).lower().strip().replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if sv in ("null", "none", "", "-", "—"): return None
    try: return float(sv)
    except (ValueError, TypeError): return None


def _clean(v) -> str | None:
    s = str(v or "").strip()
    return None if s.lower() in ("null", "none", "") else s


# ── Шаг 1: извлечение текста (openpyxl для Excel, polza.ai для остального) ────

def _extract_text_excel(file_bytes: bytes) -> str:
    """Читает все ячейки Excel через openpyxl и возвращает текст строка за строкой.
    openpyxl есть в requirements.txt — единственная нативная библиотека для документов.
    """
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets[:3]:
        for r in list(ws.rows)[:200]:
            cells = []
            for c in r:
                v = c.value
                if v is None: continue
                if isinstance(v, float):
                    cells.append(str(int(v)) if v == int(v) else str(round(v, 4)))
                else:
                    cells.append(str(v).strip())
            if cells:
                lines.append("\t".join(cells))
    return "\n".join(lines)


def _ocr_image_polza(req_lib, file_b64: str, mime: str) -> str:
    """OCR изображения через Polza.ai/GPT-4o."""
    sys_msg  = {"role": "system", "content": "Ты — OCR. Верни полный текст документа, строка за строкой."}
    user_msg = {"role": "user", "content": [
        {"type": "text",      "text": _PROMPT_OCR_IMAGE},
        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{file_b64}"}},
    ]}
    return _call_polza(req_lib, [sys_msg, user_msg], max_tokens=4096)


# ── Шаг 2: разбивка текста на header / table_lines / footer ──────────────────

_TABLE_START_RE = r'(?i)(№\s*п?/?п?|наименование|товар|материал|услуга|позиция)'
_TABLE_END_RE   = r'(?i)(итого|всего|в том числе|ндс|подпись|м\.?п\.?|директор|главный\s+бухгалтер|отпустил|принял)'

def _split_document(text: str) -> dict:
    """
    Разбивает текст документа на зоны: header, table_lines, footer.
    Возвращает dict с ключами header, table_lines (list[str]), footer, expected_count.
    """
    import re
    lines = text.splitlines()

    table_start = None
    table_end   = None

    for i, line in enumerate(lines):
        if table_start is None and re.search(_TABLE_START_RE, line):
            table_start = i
        if table_start is not None and table_end is None:
            if i > table_start + 2 and re.search(_TABLE_END_RE, line):
                table_end = i
                break

    if table_start is None:
        # Не нашли явную таблицу — пробуем эвристику: строки с числами
        num_lines = []
        for i, line in enumerate(lines):
            if re.search(r'\d+[\.,]\d+', line) and len(line.strip()) > 10:
                num_lines.append(i)
        if num_lines:
            table_start = max(0, num_lines[0] - 1)
            table_end   = num_lines[-1] + 1

    if table_start is None:
        table_start = 0
    if table_end is None or table_end <= table_start:
        table_end = len(lines)

    header_lines = lines[:table_start]
    table_lines  = [l for l in lines[table_start:table_end] if l.strip()]
    footer_lines = lines[table_end:]

    # Ищем «Всего наименований N» в footer
    import re as _re
    expected_count = 0
    footer_text = "\n".join(footer_lines)
    m = _re.search(r'(?i)всего\s+наименований[:\s]+(\d+)', footer_text)
    if not m:
        m = _re.search(r'(?i)итого\s+позиций[:\s]+(\d+)', footer_text)
    if m:
        expected_count = int(m.group(1))

    return {
        "header":         "\n".join(header_lines[:40]),
        "table_lines":    table_lines,
        "footer":         footer_text[:500],
        "expected_count": expected_count,
    }


def _make_chunks(table_lines: list, chunk_size: int = 15) -> list:
    """Режет список строк на чанки по chunk_size строк."""
    return [table_lines[i:i+chunk_size] for i in range(0, len(table_lines), chunk_size)]


# ── Шаг 3: AI-запросы к чанкам ───────────────────────────────────────────────

def _extract_header_meta(req_lib, header_text: str) -> dict:
    """Один запрос: извлечь supplier_name, invoice_date, invoice_number из шапки."""
    if not header_text.strip():
        return {}
    sys_msg  = {"role": "system", "content": "Ты — парсер финансовых документов. Отвечай ТОЛЬКО JSON-объектом."}
    user_msg = {"role": "user", "content": _PROMPT_HEADER.format(header=header_text[:1500])}
    try:
        raw = _call_polza(req_lib, [sys_msg, user_msg], max_tokens=256)
        return _parse_json_obj(raw)
    except Exception:
        return {}


def _extract_chunk_items(req_lib, chunk_lines: list, supplier_name: str = "") -> list:
    """Один запрос: извлечь позиции из чанка строк таблицы."""
    chunk_text = "\n".join(chunk_lines)
    sys_msg  = {"role": "system", "content": "Ты — парсер таблиц счетов. Отвечай ТОЛЬКО JSON-массивом без комментариев."}
    user_msg = {"role": "user", "content": _PROMPT_TABLE_CHUNK.format(chunk=chunk_text[:2000])}
    try:
        raw = _call_polza(req_lib, [sys_msg, user_msg], max_tokens=4096)
        items, _ = _parse_json_list(raw)
        return items or []
    except Exception:
        return []


# ── Шаг 4: постобработка и валидация ─────────────────────────────────────────

# Ключевые слова для детектирования единицы "тонна"
_TONNE_KEYWORDS = ['т', 'тон', 'тонн', 'тонна', 'тонны', '/т', 'mt', 'ton']
# Ключевые слова для детектирования материалов с ценой за тонну
_HEAVY_MATERIALS = ['арматур', 'металл', 'прокат', 'швеллер', 'балк', 'уголок',
                    'труб', 'лист', 'профиль', 'сталь', 'жби', 'бетон', 'цемент']

def _postprocess_items(raw_items: list, supplier_name: str, invoice_date: str | None,
                       invoice_number: str | None) -> list:
    """
    Постобработка и валидация позиций:
    - Заполняет invoice_date / invoice_number из мета если нет
    - Исправляет unit_price × 1000 для тяжёлых материалов с ценой < 1000 (путаница руб/коп или т/кг)
    - Проставляет quality: 'ok' | 'suspicious' | 'bad'
    """
    import re
    result = []
    for item in raw_items:
        if not isinstance(item, dict): continue

        m_name = _clean(item.get("material")) or ""
        unit   = _clean(item.get("unit")) or "шт"
        up     = _safe_float(item.get("unit_price"))
        qty    = _safe_float(item.get("quantity"))
        # supplier_name берём из мета для отображения (не для записи в БД)
        s_name = _clean(item.get("supplier_name")) or supplier_name or None
        idate  = _clean(item.get("invoice_date"))  or invoice_date  or None
        inum   = _clean(item.get("invoice_number")) or invoice_number or None

        # unit нормализация: приводим к нашим допустимым значениям
        unit_low = unit.lower().strip()
        if unit_low in _TONNE_KEYWORDS:
            unit = "т"
        elif unit_low in ('м2', 'кв.м', 'кв м', 'квм', 'm2', 'sq.m', 'sqm'):
            unit = "м2"
        elif unit_low in ('м3', 'куб.м', 'куб м', 'кубм', 'm3', 'cu.m', 'cum'):
            unit = "м3"
        elif unit_low in ('пог.м', 'пог м', 'погм', 'п.м', 'п/м', 'lm', 'lin.m'):
            unit = "пог.м"
        elif unit_low in ('м', 'метр', 'метры', 'метров', 'ml', 'lin', 'rm'):
            unit = "пог.м"
        elif unit_low in ('компл', 'комплект', 'компл.', 'set', 'комп'):
            unit = "компл"
        elif unit_low in ('шт', 'шт.', 'piece', 'pc', 'pcs', 'ед', 'ед.', 'штук', 'штуки', 'штука'):
            unit = "шт"
        elif unit_low in ('кг', 'kg', 'килограмм', 'килограммов', 'кило'):
            unit = "кг"
        elif unit_low in ('г', 'gr', 'gram', 'грамм'):
            unit = "г"
        elif unit_low in ('л', 'liter', 'litre', 'литр', 'литров', 'lt'):
            unit = "л"
        elif unit_low in ('мл', 'ml', 'миллилитр'):
            unit = "мл"
        elif unit_low in ('уп', 'уп.', 'упак', 'упаковка', 'pack', 'pck'):
            unit = "уп"
        elif unit_low in ('рул', 'рул.', 'рулон', 'roll'):
            unit = "рул"
        elif unit_low in ('пач', 'пач.', 'пачка', 'пачек'):
            unit = "пач"
        # Если единица не распознана — оставляем как есть (не сбрасываем в "шт")
        elif unit_low and unit_low not in ('null', 'none', '-', '—'):
            unit = unit  # сохраняем оригинал

        price_fixed = False

        # Исправление цены: если unit=т и цена подозрительно маленькая
        # (например, арматура по 55 руб/т вместо 55000 руб/т)
        if up is not None and unit == "т":
            m_low = m_name.lower()
            is_heavy = any(kw in m_low for kw in _HEAVY_MATERIALS)
            if is_heavy and 0 < up < 1000:
                up *= 1000
                price_fixed = True

        # Флаг quality
        if m_name and up is not None and qty is not None and qty > 0:
            if price_fixed:
                quality = "suspicious"
            else:
                quality = "ok"
        elif m_name:
            quality = "bad"
        else:
            continue  # совсем пустая строка — пропускаем

        result.append({
            "material":       m_name or None,
            "supplier_name":  s_name,
            "unit":           unit,
            "unit_price":     up,
            "quantity":       qty,
            "invoice_date":   idate,
            "invoice_number": inum,
            "quality":        quality,
            "price_fixed":    price_fixed,
        })
    return result


# ── Нормализация: матчинг справочников ───────────────────────────────────────

def _normalize_postprocessed(cur, items: list) -> list:
    """Матчит supplier/material в справочниках, возвращает финальные items."""
    result = []
    for item in items:
        s_name = item.get("supplier_name") or ""
        m_name = item.get("material")      or ""
        raw_u  = item.get("unit", "шт")
        up     = item.get("unit_price")
        qty    = item.get("quantity")
        idate  = item.get("invoice_date")
        inum   = item.get("invoice_number")

        s_id, s_created = _match_or_create_supplier(cur, s_name)
        m_id, m_created, unit = _match_or_create_material(cur, m_name, raw_u)

        complete = bool(m_name and up is not None and qty is not None)
        result.append({
            "supplier_name":    s_name or None,
            "supplier_id":      s_id,
            "supplier_created": s_created,
            "material":         m_name or None,
            "material_id":      m_id,
            "material_created": m_created,
            "unit":             unit,
            "unit_price":       up,
            "quantity":         qty,
            "invoice_date":     idate,
            "invoice_number":   inum,
            "complete":         complete,
            "quality":          item.get("quality", "ok"),
            "price_fixed":      item.get("price_fixed", False),
        })
    return result


# ── TABLE TEMPLATES: поиск, применение, сохранение ───────────────────────────

# Синонимы для сравнения заголовков
_HEADER_SYNONYMS = {
    "наименование": ["товар", "материал", "услуга", "номенклатура", "описание",
                     "name", "item", "goods", "description", "product"],
    "количество":   ["кол-во", "кол.", "кол", "qty", "кол-ть", "кол_во", "count"],
    "цена":         ["цена за ед", "цена за единицу", "ед.цена", "price",
                     "стоимость ед", "unit price", "цена/ед"],
    "единица":      ["ед", "ед.", "ед.изм", "ед. изм", "единица изм", "uom",
                     "unit", "measure", "ед.изм."],
    "сумма":        ["итого", "стоимость", "total", "amount", "итог", "сумма руб"],
    "номер":        ["№", "no", "n", "п/п", "п п", "поз", "позиция", "line"],
}

# Какие поля нас интересуют — только эти колонки можно маппить
_MAPPABLE_FIELDS = ["material", "quantity", "unit_price", "unit", "total", "row_num", "skip"]

# Метки полей для фронта
_FIELD_LABELS = {
    "material":   "Наименование",
    "quantity":   "Количество",
    "unit_price": "Цена",
    "unit":       "Ед. изм.",
    "total":      "Сумма",
    "row_num":    "№ строки",
    "skip":       "Пропустить",
}


def _normalize_header(h: str) -> str:
    """Приводит заголовок к нижнему регистру, убирает лишние символы."""
    import re
    return re.sub(r'[^\w\s]', '', h.lower()).strip()


def _header_to_canonical(h: str) -> str | None:
    """Пытается привести заголовок к одному из канонических значений через синонимы."""
    norm = _normalize_header(h)
    for canonical, synonyms in _HEADER_SYNONYMS.items():
        if norm == canonical or norm in synonyms:
            return canonical
        # Частичное вхождение (для составных заголовков типа «Кол-во шт»)
        for syn in [canonical] + synonyms:
            if syn in norm or norm in syn:
                return canonical
    return None


def _headers_similarity(headers_a: list, headers_b: list) -> float:
    """
    Возвращает долю совпавших заголовков (0.0 – 1.0).
    Сравнение через канонические синонимы.
    """
    if not headers_a or not headers_b:
        return 0.0
    canon_a = set(c for c in (_header_to_canonical(h) for h in headers_a) if c)
    canon_b = set(c for c in (_header_to_canonical(h) for h in headers_b) if c)
    if not canon_a or not canon_b:
        # Fallback: посимвольное сравнение нормализованных строк
        norm_a = set(_normalize_header(h) for h in headers_a)
        norm_b = set(_normalize_header(h) for h in headers_b)
        inter  = norm_a & norm_b
        return len(inter) / max(len(norm_a), len(norm_b))
    inter = canon_a & canon_b
    return len(inter) / max(len(canon_a), len(canon_b))


def _find_table_header_row(table_lines: list) -> tuple[list, int]:
    """
    Ищет строку-заголовок таблицы среди первых 5 строк.
    Возвращает (список заголовков, индекс строки) или ([], -1).
    """
    import re
    header_keywords = r'(?i)(наименование|товар|материал|кол|цена|сумма|ед|№|наим|услуг)'
    for i, line in enumerate(table_lines[:6]):
        parts = re.split(r'\s{2,}|\t', line.strip())
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) >= 3 and re.search(header_keywords, line):
            return parts, i
    return [], -1


def _find_matching_template(cur, headers: list) -> dict | None:
    """
    Ищет шаблон в БД с совпадением > 70%.
    Возвращает строку шаблона (dict) или None.
    """
    cur.execute(
        f"SELECT id, name, headers, column_map, use_count FROM {SCHEMA}.table_templates ORDER BY use_count DESC, id ASC"
    )
    rows = cur.fetchall()
    best_tmpl = None
    best_score = 0.0
    for row in rows:
        tid, tname, theaders_raw, tcol_map, tuse = row
        theaders = theaders_raw if isinstance(theaders_raw, list) else json.loads(theaders_raw or "[]")
        score = _headers_similarity(headers, theaders)
        if score > best_score:
            best_score = score
            best_tmpl = {"id": tid, "name": tname, "headers": theaders,
                         "column_map": tcol_map if isinstance(tcol_map, dict) else json.loads(tcol_map or "{}"),
                         "use_count": tuse, "score": round(score, 2)}
    if best_tmpl and best_score >= 0.7:
        return best_tmpl
    return None


def _apply_template(table_lines: list, col_map: dict, header_row_idx: int,
                    supplier_name: str, invoice_date: str | None,
                    invoice_number: str | None) -> list:
    """
    Парсит строки таблицы по col_map (без AI).
    col_map: {"material": 1, "quantity": 3, "unit_price": 4, "unit": 2}  — 0-based индексы.
    Возвращает список сырых позиций (dict) для _postprocess_items.
    """
    import re
    # Данные строки — всё что после заголовка
    data_lines = table_lines[header_row_idx + 1:]
    result = []
    for line in data_lines:
        parts = re.split(r'\s{2,}|\t', line.strip())
        parts = [p.strip() for p in parts if p.strip() != ""]
        if not parts or len(parts) < 2:
            continue
        def get_col(field):
            idx = col_map.get(field)
            if idx is None: return None
            try: return parts[int(idx)]
            except IndexError: return None

        material   = get_col("material")
        qty_raw    = get_col("quantity")
        price_raw  = get_col("unit_price")
        unit_raw   = get_col("unit")

        # Пропускаем строки без наименования или с нечисловой ценой
        if not material or len(material) < 2:
            continue
        # Пропускаем итоговые строки
        if re.match(r'(?i)^(итого|всего|ндс|в т\.ч\.|total)', material):
            continue

        result.append({
            "material":       material,
            "unit":           unit_raw or "шт",
            "unit_price":     _safe_float(price_raw),
            "quantity":       _safe_float(qty_raw),
            "supplier_name":  supplier_name or None,
            "invoice_date":   invoice_date,
            "invoice_number": invoice_number,
        })
    return result


def _ai_suggest_column_map(req_lib, headers: list) -> dict:
    """
    Просит AI предположить маппинг колонок по заголовкам.
    Возвращает dict {field: col_index, ...}.
    """
    prompt = (
        f"Дан список заголовков таблицы счёта (нумерация с 0): {json.dumps(headers, ensure_ascii=False)}.\n"
        f"Определи, какой индекс (число) соответствует каждому полю:\n"
        f"  material (наименование товара), quantity (количество), unit_price (цена за единицу), "
        f"unit (единица измерения), total (сумма строки), row_num (номер строки).\n"
        f"Верни ТОЛЬКО JSON-объект: {{\"material\":N,\"quantity\":N,\"unit_price\":N,\"unit\":N,\"total\":N,\"row_num\":N}}\n"
        f"Если поле не найдено — ставь null. Только JSON, без пояснений."
    )
    sys_msg  = {"role": "system", "content": "Отвечай ТОЛЬКО JSON-объектом."}
    user_msg = {"role": "user",   "content": prompt}
    try:
        raw = _call_polza(req_lib, [sys_msg, user_msg], max_tokens=128)
        return _parse_json_obj(raw)
    except Exception:
        return {}


def _save_template(cur, name: str, headers: list, column_map: dict, ai_suggested: bool = False) -> int:
    """Сохраняет новый шаблон в БД. Возвращает id."""
    cur.execute(
        f"INSERT INTO {SCHEMA}.table_templates (name, headers, column_map, ai_suggested) "
        f"VALUES (%s,%s,%s,%s) RETURNING id",
        (name, json.dumps(headers, ensure_ascii=False),
         json.dumps(column_map, ensure_ascii=False), ai_suggested)
    )
    return cur.fetchone()[0]


def _increment_template_use(cur, template_id: int):
    cur.execute(
        f"UPDATE {SCHEMA}.table_templates SET use_count=use_count+1, last_used_at=now(), updated_at=now() WHERE id=%s",
        (template_id,)
    )


def get_table_templates(cur) -> list:
    """CRUD: список всех шаблонов."""
    cur.execute(
        f"SELECT id, name, headers, column_map, ai_suggested, use_count, last_used_at, created_at "
        f"FROM {SCHEMA}.table_templates ORDER BY use_count DESC, id ASC"
    )
    rows = cur.fetchall()
    result = []
    for r in rows:
        result.append({
            "id":           r[0],
            "name":         r[1],
            "headers":      r[2] if isinstance(r[2], list) else json.loads(r[2] or "[]"),
            "column_map":   r[3] if isinstance(r[3], dict) else json.loads(r[3] or "{}"),
            "ai_suggested": r[4],
            "use_count":    r[5],
            "last_used_at": r[6].isoformat() if r[6] else None,
            "created_at":   r[7].isoformat() if r[7] else None,
        })
    return result


def delete_table_template(cur, template_id: int):
    cur.execute(f"DELETE FROM {SCHEMA}.table_templates WHERE id=%s RETURNING id", (template_id,))
    if not cur.fetchone():
        return False
    return True


def rename_table_template(cur, template_id: int, new_name: str):
    cur.execute(
        f"UPDATE {SCHEMA}.table_templates SET name=%s, updated_at=now() WHERE id=%s RETURNING id",
        (new_name, template_id)
    )
    return bool(cur.fetchone())


def save_user_template(cur, req_lib, name: str, headers: list, column_map: dict) -> dict:
    """
    Сохраняет шаблон подтверждённый снабженцем.
    Возвращает {id, name}.
    """
    tid = _save_template(cur, name, headers, column_map, ai_suggested=False)
    return {"id": tid, "name": name}


# ── Главная функция ───────────────────────────────────────────────────────────

def _convert_to_jpg_b64(file_bytes: bytes, ext: str, debug_log: list):
    """
    Подготавливает изображение для отправки в Gemini.
    Фронт уже конвертировал PDF/Excel в JPG — здесь принимаем только изображения.
    Возвращает (b64_str, mime) или (None, error_message).
    """
    import base64, io as _io

    # JPG/PNG — сжимаем через Pillow
    if ext in ("jpg", "jpeg", "png"):
        try:
            from PIL import Image as _PILImage
            img = _PILImage.open(_io.BytesIO(file_bytes)).convert("RGB")
            MAX = 1400
            w, h = img.size
            if w > MAX or h > MAX:
                scale = MAX / max(w, h)
                img = img.resize((int(w * scale), int(h * scale)), _PILImage.LANCZOS)
            buf = _io.BytesIO()
            img.save(buf, format="JPEG", quality=88)
            b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
            debug_log.append(f"image: {ext} → jpg size={len(buf.getvalue())}")
            return b64, "image/jpeg"
        except Exception:
            mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
            return base64.b64encode(file_bytes).decode("utf-8"), mime

    # PDF — пробуем fitz, иначе передаём напрямую в Gemini (поддерживает PDF)
    if ext == "pdf":
        jpg_b64 = _pdf_to_jpg_b64(file_bytes)
        if jpg_b64:
            debug_log.append(f"pdf → jpg via fitz, size={len(jpg_b64)}")
            return jpg_b64, "image/jpeg"
        # fitz недоступен: отправляем PDF напрямую
        debug_log.append(f"pdf → direct (fitz unavailable), size={len(file_bytes)}")
        return base64.b64encode(file_bytes).decode("utf-8"), "application/pdf"

    return None, (
        f"Формат «{ext.upper()}» не поддерживается. "
        "Загрузите JPG, PNG или PDF."
    )


def _recognize_excel_tsv(req_lib, file_bytes: bytes, ext: str, debug_log: list):
    """
    Excel → Markdown-таблица → один запрос в Gemini Flash Lite.
    Возвращает (ai_obj, all_items, error).
    """
    import io as _io, json as _json, re as _re2

    # ── 1. Читаем строки ───────────────────────────────────────────────────────
    if ext == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active or wb.worksheets[0]
        raw_rows = []
        for row in ws.iter_rows(max_row=200, values_only=True):
            raw_rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
    else:
        import xlrd
        wb_xls = xlrd.open_workbook(file_contents=file_bytes)
        ws_xls = wb_xls.sheet_by_index(0)
        raw_rows = []
        for ri in range(min(ws_xls.nrows, 200)):
            raw_rows.append([str(ws_xls.cell_value(ri, ci)) for ci in range(ws_xls.ncols)])

    rows = [r for r in raw_rows if any(str(c).strip() for c in r)]
    if not rows:
        return {}, [], "Excel пустой"

    debug_log.append(f"excel_markdown: {len(rows)} rows")

    # ── 2. Конвертируем в Markdown-таблицу ────────────────────────────────────
    ncols = max(len(r) for r in rows)
    rows  = [r + [""] * (ncols - len(r)) for r in rows]

    def _md_row(r):
        return "| " + " | ".join(str(c).replace("|", "\\|") for c in r) + " |"

    md_lines = [_md_row(rows[0])]
    md_lines.append("|" + "---|" * ncols)          # разделитель заголовка
    for r in rows[1:]:
        md_lines.append(_md_row(r))
    markdown_table = "\n".join(md_lines)

    # ── 3. Один запрос в Gemini Flash Lite ────────────────────────────────────
    _PROMPT = (
        "Ты получаешь счёт в формате Markdown-таблицы.\n"
        "Извлеки все позиции товаров/материалов.\n"
        "Поставщика, дату и номер счёта ищи в первых строках (шапке) над таблицей.\n\n"
        "Для каждой позиции верни:\n"
        "- material: полное название (не сокращай)\n"
        "- quantity: число (может быть дробным)\n"
        "- unit: единица измерения (м3, шт, м, пог.м и т.д.)\n"
        "- unit_price: цена за единицу\n"
        "- amount: сумма строки\n\n"
        "ПРАВИЛА:\n"
        "1. Строки-заголовки (№, Наименование, Кол-во...) — пропускай.\n"
        "2. Если значение отсутствует — ставь null.\n"
        "3. Числа — только цифры без пробелов и символов валюты.\n"
        "4. Не добавляй комментариев.\n\n"
        "Верни строго JSON без markdown-обёрток:\n"
        '{"supplier_name":"...","invoice_date":"YYYY-MM-DD","invoice_number":"...","footer_total":число,'
        '"items":[{"material":"...","quantity":число,"unit":"...","unit_price":число,"amount":число}]}\n\n'
        f"Счёт (Markdown):\n{markdown_table}"
    )

    messages = [
        {"role": "system", "content": "Ты — система извлечения данных из счетов. Отвечай ТОЛЬКО строгим JSON без пояснений и markdown-обёрток."},
        {"role": "user",   "content": _PROMPT},
    ]

    logger.info("Отправка Excel в Gemini (Markdown), строк=%d", len(rows))
    debug_log.append(f"excel_markdown: sending to gemini, rows={len(rows)}")

    try:
        raw = _call_polza_model(req_lib, messages=messages,
                                model="google/gemini-3.1-flash-lite", max_tokens=8192)
    except Exception as e:
        debug_log.append(f"excel_markdown gemini error: {e}")
        return {}, [], f"Gemini вернул ошибку: {e}"

    debug_log.append(f"excel_markdown: response len={len(raw)} first200={raw[:200]!r}")

    # ── 4. Парсим ответ ────────────────────────────────────────────────────────
    if not raw.strip():
        return {}, [], f"Gemini вернул пустой ответ. Первые 200 символов raw: {raw[:200]!r}"

    s = raw.strip()
    s = _re2.sub(r'```(?:json)?\s*', '', s)
    s = _re2.sub(r'```', '', s).strip()

    ai_obj, items = {}, []
    try:
        p = _json.loads(s)
        if isinstance(p, dict) and "items" in p:
            ai_obj, items = p, p.get("items") or []
    except Exception:
        for m in sorted(_re2.finditer(r'\{[\s\S]+\}', s), key=lambda x: -len(x.group(0))):
            try:
                p = _json.loads(m.group(0))
                if isinstance(p, dict) and "items" in p:
                    ai_obj, items = p, p.get("items") or []
                    break
            except Exception:
                continue

    logger.info("Excel Gemini: извлечено %d позиций", len(items))
    debug_log.append(f"excel_markdown: parsed {len(items)} items")

    if not items:
        return {}, [], f"Gemini не смог извлечь позиции. Ответ: {raw[:400]!r}"

    return ai_obj, items, None


def _call_polza_model(req_lib, messages: list, model: str, max_tokens: int = 4096) -> str:
    """Вызов Polza.ai с явным указанием модели. Возвращает строку-ответ."""
    import time as _time
    payload = {
        "messages": messages,
        "model":    model,
        "temperature": 0.0,
        "max_tokens": max_tokens,
    }
    t0 = _time.monotonic()
    logger.warning("_call_polza_model START: model=%s msg_count=%d", model, len(messages))
    resp = req_lib.post(
        f"{CHATGPT_URL}?action=generate",
        json=payload,
        timeout=180,   # 3 минуты — запас для медленных моделей
    )
    elapsed = _time.monotonic() - t0
    logger.warning("_call_polza_model END: model=%s elapsed=%.1fs status=%d", model, elapsed, resp.status_code)
    if not resp.ok:
        try:    err_text = str(resp.json())
        except Exception: err_text = resp.text[:500]
        raise RuntimeError(f"polza {resp.status_code}: {err_text}")
    content = resp.json().get("content", "")
    logger.warning("_call_polza_model CONTENT: model=%s len=%d first100=%r", model, len(content), content[:100])
    return content


def recognize_invoice(cur, invoice_id: int):
    """
    Конвейер распознавания счёта:
      Шаг 1. Конвертация файла в JPG (PDF→PyMuPDF, Excel→Pillow, Image→as is).
      Шаг 2. AI-запрос к Gemini с изображением.
      Шаг 3. Постобработка: нормализация единиц, матчинг справочников.
      Возвращает {status, meta, items, parse_error, debug}.
    """
    import requests as req_lib
    import base64, io as _io

    debug_log  = []
    parse_error = None

    # ── 0. Загружаем данные счёта из БД ──────────────────────────────────────
    cur.execute(
        f"SELECT id, pdf_file_url, pdf_file_name FROM {SCHEMA}.invoices WHERE id=%s",
        (invoice_id,)
    )
    row = cur.fetchone()
    if not row:
        return None, "Счёт не найден"
    _, file_url, file_name = row
    if not file_url:
        return None, "Файл не загружен. Сначала прикрепите файл счёта."

    ext = (file_name or "").rsplit(".", 1)[-1].lower() if file_name else ""

    # ── 1. Скачиваем файл ────────────────────────────────────────────────────
    try:
        resp = req_lib.get(file_url, timeout=30)
        resp.raise_for_status()
        file_bytes = resp.content
    except Exception as e:
        return None, f"Не удалось загрузить файл: {e}"

    debug_log.append(f"file={file_name} ext={ext} size={len(file_bytes)}")

    # ── 2a. Excel — отдельный пайплайн TSV + DeepSeek ─────────────────────────
    if ext in ("xls", "xlsx"):
        debug_log.append("routing to excel_tsv pipeline")
        ai_obj, all_raw_items, excel_err = _recognize_excel_tsv(req_lib, file_bytes, ext, debug_log)
        if excel_err and not all_raw_items:
            return None, f"Не удалось извлечь данные из Excel: {excel_err}"
        # Дальше используем общий постобработчик — перепрыгиваем через блок image
        supplier_name  = _clean(ai_obj.get("supplier_name")) or ""
        invoice_date   = _clean(ai_obj.get("invoice_date"))
        invoice_number = _clean(ai_obj.get("invoice_number"))
        footer_total   = _safe_float(ai_obj.get("footer_total"))
        debug_log.append(f"excel meta: supplier={supplier_name!r} date={invoice_date} num={invoice_number}")
        # Постобработка и матчинг — такой же как для JPG
        processed  = _postprocess_items(all_raw_items, supplier_name=supplier_name,
                                        invoice_date=invoice_date, invoice_number=invoice_number)
        norm_items = _normalize_postprocessed(cur, processed)
        all_ok     = bool(norm_items) and all(i.get("complete") for i in norm_items)
        status     = "обработан" if all_ok else "требуется_проверка"
        cur.execute(
            f"UPDATE {SCHEMA}.invoices SET recognition_status=%s, recognized_data=%s, updated_at=now() WHERE id=%s",
            (status, json.dumps(all_raw_items, ensure_ascii=False), invoice_id)
        )
        return {
            "status": status,
            "meta": {"supplier_name": supplier_name, "invoice_date": invoice_date,
                     "invoice_number": invoice_number, "footer_total": footer_total},
            "items": norm_items,
            "parse_error": excel_err,
            "debug": debug_log,
        }, None

    # ── 2. Конвертация в JPG если нужно (JPG/PNG/PDF) ────────────────────────
    file_b64, mime = _convert_to_jpg_b64(file_bytes, ext, debug_log)
    if file_b64 is None:
        return None, mime  # mime содержит сообщение об ошибке

    # ── Вспомогательная функция парсинга ответа polza.ai ─────────────────────
    import re as _re

    def _parse_ai_invoice_response(raw: str):
        """Парсит JSON-ответ polza.ai. Возвращает (ai_obj, items, error)."""
        s = raw.strip()
        s = _re.sub(r'```(?:json)?\s*', '', s)
        s = _re.sub(r'```', '', s).strip()
        # Прямой парсинг
        try:
            p = json.loads(s)
            if isinstance(p, dict) and "items" in p:
                return p, p.get("items") or [], None
            if isinstance(p, list):
                return {}, p, None
        except json.JSONDecodeError as je:
            pass
        # Fallback: ищем JSON-объект с items
        for m in sorted(_re.finditer(r'\{[\s\S]+\}', s), key=lambda x: -len(x.group(0))):
            try:
                p = json.loads(m.group(0))
                if isinstance(p, dict) and "items" in p:
                    return p, p.get("items") or [], None
            except Exception:
                continue
        # Fallback: ищем массив
        for m in sorted(_re.finditer(r'\[[\s\S]+\]', s), key=lambda x: -len(x.group(0))):
            try:
                arr = json.loads(m.group(0))
                if isinstance(arr, list):
                    return {}, arr, None
            except Exception:
                continue
        return {}, [], f"JSON не найден. Начало ответа: {s[:300]}"

    def _calc_items_total(items: list) -> float:
        """Считает сумму позиций из сырого списка AI."""
        total = 0.0
        for it in items:
            if not isinstance(it, dict): continue
            # Берём amount если есть, иначе unit_price * quantity
            amt = _safe_float(it.get("amount"))
            if amt:
                total += amt
            else:
                up  = _safe_float(it.get("unit_price"))
                qty = _safe_float(it.get("quantity"))
                if up and qty:
                    total += up * qty
        return total

    # ── 3. Формируем запрос к AI в зависимости от типа файла ─────────────────
    sys_msg = {
        "role": "system",
        "content": "Ты — система извлечения данных из счетов. Отвечай ТОЛЬКО JSON без пояснений и markdown."
    }

    # JPG / PNG / PDF (application/pdf) — всё передаём как data URL
    img_content = {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{file_b64}"}}
    usr_msg = {"role": "user", "content": [
        {"type": "text", "text": _PROMPT_IMAGE_DIRECT},
        img_content,
    ]}

    try:
        raw_response = _call_polza(req_lib, [sys_msg, usr_msg], max_tokens=4096)
        debug_log.append(f"attempt 1: {len(raw_response)} chars")
    except Exception as pe:
        return None, f"Ошибка обращения к сервису распознавания: {pe}"

    ai_obj, all_raw_items, parse_err = _parse_ai_invoice_response(raw_response)

    # ── PDF с пустым ответом: проверяем на QR-код ─────────────────────────────
    if not all_raw_items and ext == "pdf":
        debug_log.append(f"pdf empty response, checking for QR")
        # Если PDF содержит текстовый маркер QR или ответ намекает на графику — даём понятную подсказку
        raw_lower = raw_response.lower()
        has_qr_hint = (
            "qr" in raw_lower or "qr-код" in raw_lower or
            "штрих" in raw_lower or "barcode" in raw_lower or
            len(raw_response.strip()) < 30   # почти пустой ответ = Gemini не смог прочитать
        )
        if has_qr_hint:
            return None, (
                "qr_detected|"
                "Документ содержит QR‑код или графику, которую система не может корректно обработать. "
                "Пожалуйста, откройте счёт и сделайте скриншот первой страницы в JPG, затем загрузите его."
            )

    if not all_raw_items:
        debug_log.append(f"all attempts empty: {parse_err}")
        return None, (
            f"AI не смог распознать позиции из документа. "
            f"Попробуйте загрузить более чёткий скан или введите позиции вручную. "
            f"Детали: {parse_err or 'пустой ответ'}"
        )

    debug_log.append(f"attempt 1: parsed {len(all_raw_items)} items")

    # ── 4. Извлечь meta из первого ответа ────────────────────────────────────
    supplier_name  = _clean(ai_obj.get("supplier_name")) or ""
    invoice_date   = _clean(ai_obj.get("invoice_date"))
    invoice_number = _clean(ai_obj.get("invoice_number"))
    footer_total   = _safe_float(ai_obj.get("footer_total"))

    if not supplier_name and not invoice_date and not invoice_number and all_raw_items:
        first = all_raw_items[0]
        supplier_name  = _clean(first.get("supplier_name")) or ""
        invoice_date   = _clean(first.get("invoice_date"))
        invoice_number = _clean(first.get("invoice_number"))

    debug_log.append(
        f"meta: supplier={supplier_name!r} date={invoice_date} "
        f"num={invoice_number} footer_total={footer_total}"
    )

    # ── 5. Самокорректирующийся цикл (до 2 повторных запросов) ───────────────
    raw_response_2 = ""
    correction_attempts = 0
    MAX_CORRECTIONS = 2

    if footer_total and footer_total > 0:
        for attempt in range(MAX_CORRECTIONS):
            items_total = _calc_items_total(all_raw_items)
            if items_total <= 0:
                break
            diff_pct = abs(items_total - footer_total) / footer_total
            debug_log.append(
                f"sum check attempt {attempt + 1}: "
                f"items={items_total:.2f} footer={footer_total:.2f} diff={diff_pct:.1%}"
            )
            if diff_pct <= 0.01:
                break  # сумма сошлась — выходим

            # Запрашиваем исправление
            correction_attempts += 1
            corr_prompt = _PROMPT_IMAGE_CORRECTION.format(
                footer_total=footer_total,
                items_total=round(items_total, 2),
            )
            corr_usr = {"role": "user", "content": [
                {"type": "text", "text": corr_prompt},
                img_content,
            ]}
            try:
                raw_corr = _call_polza(req_lib, [sys_msg, corr_usr], max_tokens=4096)
                debug_log.append(f"correction {attempt + 1}: {len(raw_corr)} chars")
                raw_response_2 = raw_corr
                corr_obj, corr_items, corr_err = _parse_ai_invoice_response(raw_corr)
                if corr_items:
                    all_raw_items = corr_items
                    # Обновляем meta если пришла — используем .get() везде
                    supplier_name  = _clean(corr_obj.get("supplier_name"))  or supplier_name
                    invoice_date   = _clean(corr_obj.get("invoice_date"))   or invoice_date
                    invoice_number = _clean(corr_obj.get("invoice_number")) or invoice_number
                    debug_log.append(f"correction {attempt + 1}: {len(corr_items)} items")
                else:
                    debug_log.append(f"correction {attempt + 1} parse failed: {corr_err}")
                    break
            except Exception as ce:
                debug_log.append(f"correction {attempt + 1} error: {ce}")
                break

    # ── 6. Постобработка: нормализация единиц, исправление цен ──────────────
    processed = _postprocess_items(
        all_raw_items,
        supplier_name=supplier_name,
        invoice_date=invoice_date,
        invoice_number=invoice_number,
    )
    debug_log.append(f"after postprocess: {len(processed)} items")

    # ── 7. Матчинг справочников ───────────────────────────────────────────────
    norm_items = _normalize_postprocessed(cur, processed)

    # ── 8. Финальная проверка суммы ───────────────────────────────────────────
    total_warning = None
    if footer_total and footer_total > 0:
        final_total = sum(
            (it["unit_price"] or 0) * (it["quantity"] or 0)
            for it in norm_items
            if it.get("unit_price") and it.get("quantity")
        )
        if final_total > 0:
            final_diff = abs(final_total - footer_total) / footer_total
            if final_diff > 0.01:
                total_warning = (
                    f"Итоговая сумма не сходится: по позициям {final_total:,.2f} ₽, "
                    f"в счёте {footer_total:,.2f} ₽ "
                    f"(расхождение {final_diff:.1%}). Проверьте позиции."
                )
                debug_log.append(
                    f"FINAL MISMATCH after {correction_attempts} corrections: "
                    f"items={final_total:.2f} footer={footer_total:.2f} diff={final_diff:.1%}"
                )
            else:
                debug_log.append(f"sum OK after {correction_attempts} corrections: {final_total:.2f}")

    # ── 9. Сохраняем в БД ────────────────────────────────────────────────────
    all_ok = bool(norm_items) and all(i.get("complete") for i in norm_items)
    status = "обработан" if all_ok else "требуется_проверка"

    cur.execute(
        f"UPDATE {SCHEMA}.invoices SET recognition_status=%s, recognized_data=%s, updated_at=now() WHERE id=%s",
        (status, json.dumps(all_raw_items, ensure_ascii=False), invoice_id)
    )

    return {
        "status":       status,
        "meta":         {"invoice_date": invoice_date, "invoice_number": invoice_number},
        "items":        norm_items,
        "items_count":  len(norm_items),
        "parse_error":  total_warning,   # предупреждение о сумме (не блокирует)
        "footer_total": footer_total,
        "fallback_used": False,
        # шаблоны отключены
        "template_used":          False,
        "template":               {"id": None, "name": None, "score": None},
        "need_template_setup":    False,
        "table_headers":          [],
        "ai_col_suggestion":      {},
        "template_fallback_info": None,
        "debug": {
            "raw_response":      raw_response[:2000],
            "raw_response_2":    raw_response_2[:2000] if raw_response_2 else "",
            "parse_error":       parse_err,
            "fallback_used":     False,
            "text_source":       "image_direct",
            "correction_rounds": correction_attempts,
            "items_debug":       [
                f"[{it.get('material') or '?'}] up={it.get('unit_price')} "
                f"qty={it.get('quantity')} quality={it.get('quality')} "
                f"fix={it.get('price_fixed')} sid={it.get('supplier_id')} "
                f"sname={it.get('supplier_name')}"
                for it in norm_items
            ],
            "continuation_log":  debug_log,
        },
    }, None


def apply_invoice_items(cur, source_invoice_id: int, items: list, invoice_date, invoice_number, file_url: str, file_name: str):
    """Создаёт отдельные записи счетов для каждой выбранной позиции.
    Если supplier_id/material_id = 0 — матчит или создаёт по имени.
    Каждая позиция защищена try/except — ошибка одной не рушит всю транзакцию."""
    created_ids = []
    skipped     = []

    for idx, item in enumerate(items):
        if not isinstance(item, dict):
            logger.warning(f"apply_invoice_items: item[{idx}] не dict, пропуск: {type(item)}")
            continue

        try:
            up  = _safe_float(item.get("unit_price"))
            qty = _safe_float(item.get("quantity"))

            # ── Поставщик: матчим по имени если id не передан ─────────────────
            s_id = item.get("supplier_id") or 0
            if not s_id:
                s_name = _clean(item.get("supplier_name") or item.get("supplier") or "")
                if s_name:
                    try:
                        matched_id, _ = _match_or_create_supplier(cur, s_name)
                        s_id = matched_id or 0
                    except Exception as e_s:
                        logger.warning(f"apply_invoice_items[{idx}]: supplier match error '{s_name}': {e_s}")
                        s_id = 0

            # ── Материал: матчим по имени если id не передан ──────────────────
            m_id     = item.get("material_id") or 0
            raw_unit = _clean(item.get("unit") or "шт") or "шт"
            if not m_id:
                m_name = _clean(item.get("material") or item.get("material_name") or "")
                if m_name:
                    try:
                        matched_id, _, raw_unit = _match_or_create_material(cur, m_name, raw_unit)
                        m_id = matched_id or 0
                    except Exception as e_m:
                        logger.warning(f"apply_invoice_items[{idx}]: material match error '{m_name}': {e_m}")
                        m_id = 0

            material    = _clean(item.get("material") or item.get("material_name")) or None
            s_name_raw  = _clean(item.get("supplier_name") or item.get("supplier") or "") or None
            key_ok      = bool(material and up is not None and qty is not None)
            item_status = "обработан" if key_ok else "требуется_проверка"

            try:
                rec_data = json.dumps(item, ensure_ascii=False)
            except Exception:
                rec_data = None

            logger.info(
                f"apply_invoice_items[{idx}]: material={material!r} s_id={s_id} m_id={m_id} "
                f"up={up} qty={qty} status={item_status}"
            )

            cur.execute("SAVEPOINT _apply_item")
            cur.execute(f"""
                INSERT INTO {SCHEMA}.invoices
                    (supplier_id, material_id, material_name_raw, supplier_name_raw,
                     invoice_date, invoice_number,
                     unit_price, quantity, pdf_file_url, pdf_file_name,
                     recognition_status, recognized_data)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (
                int(s_id), int(m_id),
                material,       # всегда пишем сырое название — fallback если материал не нашёлся в справочнике
                s_name_raw,     # сырое название поставщика
                invoice_date or None,
                invoice_number or None,
                up,
                qty,
                file_url  or None,
                file_name or None,
                item_status,
                rec_data,
            ))
            created_ids.append(cur.fetchone()[0])

        except Exception as e_item:
            logger.error(
                f"apply_invoice_items[{idx}]: ОШИБКА, позиция пропущена. "
                f"item={json.dumps(item, ensure_ascii=False)[:300]} err={e_item}\n{traceback.format_exc()}"
            )
            skipped.append(idx)
            # Сбрасываем состояние курсора после ошибки, чтобы продолжить транзакцию
            try:
                cur.execute("ROLLBACK TO SAVEPOINT _apply_item")
            except Exception:
                pass

    if skipped:
        logger.warning(f"apply_invoice_items: пропущено {len(skipped)} позиций из {len(items)}: {skipped}")

    return created_ids


# ─── SUPPLIERS ────────────────────────────────────────────────────────────────

SUPPLIER_CATEGORIES = ['бетон','пиломатериалы','металл','кровля','инженерия','отделка','прочее']
MATERIAL_UNITS      = ['шт','м3','т','пог.м','м2','компл','кг','л','уп','рул','пач','г','мл']

def get_suppliers(cur):
    """Список поставщиков."""
    cur.execute(f"""
        SELECT id, name, inn, category, contact, rating, is_active, created_at
        FROM {SCHEMA}.suppliers WHERE is_active=TRUE ORDER BY name
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for r in rows:
        if r.get('created_at'): r['created_at'] = r['created_at'].isoformat()
    return rows

def create_supplier(cur, body):
    """Создать поставщика."""
    name     = (body.get('name') or '').strip()
    if not name: raise ValueError("Название обязательно")
    inn      = (body.get('inn') or '').strip()
    category = body.get('category', 'прочее')
    if category not in SUPPLIER_CATEGORIES: category = 'прочее'
    contact  = (body.get('contact') or '').strip()
    rating   = body.get('rating')
    if rating is not None:
        rating = int(rating)
        if not (1 <= rating <= 5): rating = None
    cur.execute(f"""
        INSERT INTO {SCHEMA}.suppliers (name, inn, category, contact, rating)
        VALUES (%s,%s,%s,%s,%s) RETURNING id, name
    """, (name, inn or None, category, contact or None, rating))
    row = cur.fetchone()
    return {"id": row[0], "name": row[1]}

def update_supplier(cur, sid, body):
    """Обновить поставщика."""
    sets, vals = [], []
    for field in ['name','inn','category','contact','rating','is_active']:
        if field in body:
            val = body[field]
            if field == 'name':
                val = (val or '').strip()
                if not val: raise ValueError("Название обязательно")
            if field == 'category' and val not in SUPPLIER_CATEGORIES:
                val = 'прочее'
            if field == 'rating' and val is not None:
                val = int(val)
                if not (1 <= val <= 5): val = None
            sets.append(f"{field}=%s"); vals.append(val)
    if not sets: return False
    sets.append("updated_at=now()"); vals.append(int(sid))
    cur.execute(f"UPDATE {SCHEMA}.suppliers SET {', '.join(sets)} WHERE id=%s", vals)
    return True

def import_suppliers_csv(cur, rows):
    """Импорт поставщиков из CSV (список словарей с ключами name,inn,category,contact,rating)."""
    created = 0
    for row in rows:
        name = (row.get('name') or '').strip()
        if not name: continue
        cat = row.get('category','прочее')
        if cat not in SUPPLIER_CATEGORIES: cat = 'прочее'
        rating = row.get('rating')
        try: rating = int(rating) if rating else None
        except: rating = None
        if rating and not (1 <= rating <= 5): rating = None
        cur.execute(f"""
            INSERT INTO {SCHEMA}.suppliers (name, inn, category, contact, rating)
            VALUES (%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING RETURNING id
        """, (name, (row.get('inn') or None), cat, (row.get('contact') or None), rating))
        if cur.fetchone(): created += 1
    return {"created": created}


# ─── MATERIALS ────────────────────────────────────────────────────────────────

def get_materials(cur):
    """Список материалов."""
    cur.execute(f"""
        SELECT id, name, unit, supplier_category, is_active, created_at
        FROM {SCHEMA}.materials WHERE is_active=TRUE ORDER BY name
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for r in rows:
        if r.get('created_at'): r['created_at'] = r['created_at'].isoformat()
    return rows

def create_material(cur, body):
    """Создать материал."""
    name = (body.get('name') or '').strip()
    if not name: raise ValueError("Наименование обязательно")
    unit = body.get('unit','шт')
    if unit not in MATERIAL_UNITS: unit = 'шт'
    cat  = body.get('supplier_category')
    if cat and cat not in SUPPLIER_CATEGORIES: cat = None
    cur.execute(f"""
        INSERT INTO {SCHEMA}.materials (name, unit, supplier_category)
        VALUES (%s,%s,%s) RETURNING id, name, unit
    """, (name, unit, cat))
    row = cur.fetchone()
    return {"id": row[0], "name": row[1], "unit": row[2]}

def update_material(cur, mid, body):
    """Обновить материал."""
    sets, vals = [], []
    for field in ['name','unit','supplier_category','is_active']:
        if field in body:
            val = body[field]
            if field == 'unit' and val not in MATERIAL_UNITS: val = 'шт'
            if field == 'supplier_category' and val and val not in SUPPLIER_CATEGORIES: val = None
            sets.append(f"{field}=%s"); vals.append(val)
    if not sets: return False
    sets.append("updated_at=now()"); vals.append(int(mid))
    cur.execute(f"UPDATE {SCHEMA}.materials SET {', '.join(sets)} WHERE id=%s", vals)
    return True


# ─── INVOICES ────────────────────────────────────────────────────────────────

def get_invoices(cur):
    """Список счетов с поставщиком и материалом.
    material_name: из справочника, иначе сырое из AI (material_name_raw).
    supplier_name: из справочника, иначе сырое из AI (supplier_name_raw)."""
    cur.execute(f"""
        SELECT i.id, i.supplier_id,
               COALESCE(
                   CASE WHEN s.name = '(не указан)' OR s.name IS NULL THEN NULL ELSE s.name END,
                   i.supplier_name_raw
               ) as supplier_name,
               i.material_id,
               COALESCE(
                   CASE WHEN m.name = '(не указан)' OR m.name IS NULL THEN NULL ELSE m.name END,
                   i.material_name_raw
               ) as material_name,
               COALESCE(m.unit, 'шт') as unit,
               i.invoice_date, i.invoice_number,
               i.unit_price, i.quantity, i.total_amount,
               i.pdf_file_url, i.pdf_file_name,
               i.recognition_status, i.recognized_data,
               i.created_at
        FROM {SCHEMA}.invoices i
        LEFT JOIN {SCHEMA}.suppliers s ON s.id = i.supplier_id
        LEFT JOIN {SCHEMA}.materials m ON m.id = i.material_id
        ORDER BY i.created_at DESC LIMIT 200
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for r in rows:
        for k in ('invoice_date','created_at'):
            if r.get(k): r[k] = r[k].isoformat()
        for k in ('unit_price','quantity','total_amount'):
            if r.get(k) is not None: r[k] = float(r[k])
    return rows

def create_invoice(cur, body):
    """Создать счёт. Поставщик и материал необязательны (могут быть заполнены после AI-распознавания).
    Если загружен PDF, но нет цены/даты — статус = новый."""
    logger.info(f"create_invoice body keys: {list(body.keys())}, supplier_id={body.get('supplier_id')!r}, recognition_status={body.get('recognition_status')!r}")
    supplier_id = body.get('supplier_id') or None
    material_id = body.get('material_id') or None
    inv_date    = body.get('invoice_date') or None
    inv_num     = (body.get('invoice_number') or '').strip() or None
    unit_price  = float(body['unit_price']) if body.get('unit_price') not in (None,'') else None
    quantity    = float(body['quantity'])   if body.get('quantity')   not in (None,'') else None
    pdf_url     = (body.get('pdf_file_url') or '').strip() or None
    pdf_name    = (body.get('pdf_file_name') or '').strip() or None
    rec_data    = body.get('recognized_data') or None
    # Автостатус
    status = body.get('recognition_status','новый')
    if pdf_url and (not unit_price or not inv_date):
        status = 'новый'
    if status not in ('новый','обработан','требуется_проверка'): status = 'новый'
    # supplier_id/material_id = 0 (заглушка «не указан») если не выбраны
    sid = int(supplier_id) if supplier_id else 0
    mid = int(material_id) if material_id else 0
    cur.execute(f"""
        INSERT INTO {SCHEMA}.invoices
            (supplier_id, material_id, invoice_date, invoice_number,
             unit_price, quantity, pdf_file_url, pdf_file_name,
             recognition_status, recognized_data)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (sid, mid, inv_date, inv_num, unit_price, quantity, pdf_url, pdf_name, status, rec_data))
    return {"id": cur.fetchone()[0]}

def update_invoice(cur, iid, body):
    """Обновить счёт (статус, распознанные данные, поля, включая supplier_id/material_id)."""
    sets, vals = [], []
    fields_map = {
        'invoice_date':'invoice_date', 'invoice_number':'invoice_number',
        'unit_price':'unit_price', 'quantity':'quantity',
        'recognition_status':'recognition_status', 'recognized_data':'recognized_data',
        'pdf_file_url':'pdf_file_url', 'pdf_file_name':'pdf_file_name',
    }
    for key, col in fields_map.items():
        if key in body:
            val = body[key]
            if col == 'recognition_status' and val not in ('новый','обработан','требуется_проверка'):
                val = 'требуется_проверка'
            sets.append(f"{col}=%s"); vals.append(val)
    # supplier_id и material_id — отдельная обработка (FK с заглушкой 0)
    if 'supplier_id' in body:
        sid = body['supplier_id']
        sets.append("supplier_id=%s"); vals.append(int(sid) if sid else 0)
    if 'material_id' in body:
        mid = body['material_id']
        sets.append("material_id=%s"); vals.append(int(mid) if mid else 0)
    if not sets: return False
    sets.append("updated_at=now()"); vals.append(int(iid))
    cur.execute(f"UPDATE {SCHEMA}.invoices SET {', '.join(sets)} WHERE id=%s", vals)
    return True


# ─── PURCHASE REQUESTS ────────────────────────────────────────────────────────

def get_purchase_requests(cur):
    """Список заявок на закупку с поставщиками."""
    cur.execute(f"""
        SELECT pr.id, pr.created_at, pr.staff_id, s.name as staff_name,
               pr.material_id, m.name as material_name, m.unit, m.supplier_category,
               pr.quantity, pr.needed_by, pr.status
        FROM {SCHEMA}.purchase_requests pr
        JOIN {SCHEMA}.staff s ON s.id = pr.staff_id
        JOIN {SCHEMA}.materials m ON m.id = pr.material_id
        ORDER BY pr.created_at DESC LIMIT 200
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    for r in rows:
        for k in ('created_at','needed_by'):
            if r.get(k): r[k] = r[k].isoformat()
        if r.get('quantity') is not None: r['quantity'] = float(r['quantity'])
        # Загружаем связанных поставщиков
        cur.execute(f"""
            SELECT sup.id, sup.name, sup.category, sup.rating
            FROM {SCHEMA}.purchase_request_suppliers prs
            JOIN {SCHEMA}.suppliers sup ON sup.id = prs.supplier_id
            WHERE prs.request_id = %s
        """, (r['id'],))
        r['suppliers'] = [{"id": rr[0], "name": rr[1], "category": rr[2], "rating": rr[3]}
                          for rr in cur.fetchall()]
    return rows

def create_purchase_request(cur, body):
    """Создать заявку на закупку. Автоподбор поставщиков по категории материала."""
    staff_id    = body.get('staff_id')
    material_id = body.get('material_id')
    if not staff_id or not material_id:
        raise ValueError("Отправитель и материал обязательны")
    quantity  = float(body.get('quantity', 1))
    needed_by = body.get('needed_by') or None
    status    = body.get('status', 'новая')
    if status not in ('новая','в_работе','закрыта'): status = 'новая'

    cur.execute(f"""
        INSERT INTO {SCHEMA}.purchase_requests (staff_id, material_id, quantity, needed_by, status)
        VALUES (%s,%s,%s,%s,%s) RETURNING id
    """, (int(staff_id), int(material_id), quantity, needed_by, status))
    req_id = cur.fetchone()[0]

    # Автоподбор поставщиков по категории материала
    cur.execute(f"SELECT supplier_category FROM {SCHEMA}.materials WHERE id=%s", (int(material_id),))
    mat_row = cur.fetchone()
    if mat_row and mat_row[0]:
        cur.execute(f"""
            SELECT id FROM {SCHEMA}.suppliers
            WHERE category=%s AND is_active=TRUE LIMIT 20
        """, (mat_row[0],))
        for (sup_id,) in cur.fetchall():
            cur.execute(f"""
                INSERT INTO {SCHEMA}.purchase_request_suppliers (request_id, supplier_id)
                VALUES (%s,%s) ON CONFLICT DO NOTHING
            """, (req_id, sup_id))

    # Добавить явно переданных поставщиков
    for sid in (body.get('supplier_ids') or []):
        cur.execute(f"""
            INSERT INTO {SCHEMA}.purchase_request_suppliers (request_id, supplier_id)
            VALUES (%s,%s) ON CONFLICT DO NOTHING
        """, (req_id, int(sid)))

    return {"id": req_id}

def update_purchase_request(cur, rid, body):
    """Обновить статус или данные заявки на закупку."""
    sets, vals = [], []
    if 'status' in body:
        st = body['status']
        if st not in ('новая','в_работе','закрыта'): st = 'в_работе'
        sets.append("status=%s"); vals.append(st)
    if 'quantity' in body:
        sets.append("quantity=%s"); vals.append(float(body['quantity']))
    if 'needed_by' in body:
        sets.append("needed_by=%s"); vals.append(body['needed_by'] or None)
    if sets:
        sets.append("updated_at=now()"); vals.append(int(rid))
        cur.execute(f"UPDATE {SCHEMA}.purchase_requests SET {', '.join(sets)} WHERE id=%s", vals)
    # Перезаписать связанных поставщиков если переданы
    if 'supplier_ids' in body:
        cur.execute(f"DELETE FROM {SCHEMA}.purchase_request_suppliers WHERE request_id=%s", (int(rid),))
        for sid in (body['supplier_ids'] or []):
            cur.execute(f"""
                INSERT INTO {SCHEMA}.purchase_request_suppliers (request_id, supplier_id)
                VALUES (%s,%s) ON CONFLICT DO NOTHING
            """, (int(rid), int(sid)))
    return True


# ─── PURCHASE PLAN ───────────────────────────────────────────────────────────

def get_purchase_plan(cur):
    """Плановые закупки."""
    cur.execute(f"""
        SELECT pp.id, pp.material_id, m.name as material_name, m.unit,
               pp.planned_volume, pp.period, pp.period_start, pp.created_at
        FROM {SCHEMA}.purchase_plan pp
        JOIN {SCHEMA}.materials m ON m.id = pp.material_id
        ORDER BY pp.period_start DESC, m.name
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    for r in rows:
        for k in ('period_start','created_at'):
            if r.get(k): r[k] = r[k].isoformat()
        if r.get('planned_volume') is not None: r['planned_volume'] = float(r['planned_volume'])
    return rows

def create_purchase_plan(cur, body):
    """Добавить позицию в план закупок."""
    material_id = body.get('material_id')
    if not material_id: raise ValueError("Материал обязателен")
    volume = float(body.get('planned_volume', 0))
    period = body.get('period','месяц')
    if period not in ('неделя','месяц'): period = 'месяц'
    period_start = body.get('period_start')
    if not period_start: raise ValueError("Дата начала периода обязательна")
    cur.execute(f"""
        INSERT INTO {SCHEMA}.purchase_plan (material_id, planned_volume, period, period_start)
        VALUES (%s,%s,%s,%s) RETURNING id
    """, (int(material_id), volume, period, period_start))
    return {"id": cur.fetchone()[0]}

def delete_purchase_plan(cur, pid):
    """Удалить позицию плана."""
    cur.execute(f"DELETE FROM {SCHEMA}.purchase_plan WHERE id=%s", (int(pid),))
    return {"deleted": pid}


# ─── HANDLER ─────────────────────────────────────────────────────────────────

def handler(event: dict, context) -> dict:
    """Единый ERP API: deals, projects, procurement, payments, kcompany, dashboard, clients, staff"""
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS_HEADERS, "body": ""}

    method = event.get("httpMethod", "GET")
    path = event.get("path", "/")
    qs = event.get("queryStringParameters") or {}
    body = {}
    if event.get("body"):
        try:
            body = json.loads(event["body"])
        except Exception:
            pass

    # Контекст пользователя из заголовков X-User-Id / X-User-Role
    user_id, user_role = get_user_context(event)

    # Определяем роут: сначала из querystring ?r=, затем из path
    ROUTES = {"deals", "projects", "procurement", "payments", "kcompany", "dashboard", "clients", "staff",
              "employees", "reports", "slots", "serial_projects", "configurations", "individual_requests",
              "stage_durations", "estimate_works", "estimate_materials", "estimate",
              "contractors", "documents", "doc_templates", "contract_docs",
              "notifications", "payout_requests", "realtors_report", "gantt_stages",
              "suppliers", "materials", "invoices", "purchase_requests", "purchase_plan",
              "extract_pdf_text"}
    resource = qs.get("r", "")
    if not resource:
        parts = [p for p in path.split("/") if p]
        resource = next((p for p in parts if p in ROUTES), "")

    conn = get_conn()
    cur = conn.cursor()

    try:
        # ── DEALS ──────────────────────────────────────────────────────────────
        if resource == "deals":
            if method == "GET":
                data = get_deals(
                    cur,
                    archived=qs.get("archived") == "1",
                    user_id=user_id,
                    user_role=user_role,
                )
                return ok(data)
            elif method == "POST":
                action = body.get("action", "create")
                # Owner-check: менеджер/риэлтор может менять только свою сделку.
                # Применяем ко всем действиям, изменяющим конкретную сделку.
                if action != "create" and body.get("deal_id"):
                    require_deal_owner(cur, int(body["deal_id"]), user_id, user_role)
                if action in ("update_stage", "kp", "contract", "lost", "planning", "closed"):
                    deal_id = int(body["deal_id"])
                    stage = body.get("stage") or action
                    result, error = update_deal_stage(cur, deal_id, stage, body)
                    if error:
                        return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "save_kp_slot":
                    deal_id = int(body["deal_id"])
                    result, error = save_kp_slot(cur, deal_id, body)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "confirm_kp_contract":
                    deal_id = int(body["deal_id"])
                    result, error = confirm_kp_contract(cur, deal_id)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "confirm_kp_payment":
                    deal_id = int(body["deal_id"])
                    result, error = confirm_kp_payment(cur, deal_id)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "to_planning":
                    deal_id = int(body["deal_id"])
                    result, error = move_to_planning(cur, deal_id, body)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "archive":
                    deal_id = int(body["deal_id"])
                    # Читаем слоты сделки
                    cur.execute(f"SELECT slot_id, kp_slot_id FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
                    drow = cur.fetchone()
                    if not drow:
                        return err("Сделка не найдена")
                    slot_id_val, kp_slot_id_val = drow
                    # Освобождаем слоты (booked или busy → free)
                    for sid in set(filter(None, [slot_id_val, kp_slot_id_val])):
                        cur.execute(f"""
                            UPDATE {SCHEMA}.slots SET status='free', deal_id=NULL
                            WHERE id=%s AND status IN ('booked','busy')
                        """, (sid,))
                    # Архивируем связанный проект если есть
                    cur.execute(f"""
                        UPDATE {SCHEMA}.projects SET status='archived', updated_at=now()
                        WHERE deal_id=%s AND status NOT IN ('archived','completed')
                    """, (deal_id,))
                    # Архивируем сделку
                    cur.execute(f"UPDATE {SCHEMA}.deals SET is_archived=TRUE, slot_id=NULL, updated_at=now() WHERE id=%s RETURNING code", (deal_id,))
                    row = cur.fetchone()
                    if not row:
                        return err("Сделка не найдена")
                    conn.commit()
                    return ok({"success": True, "code": row[0], "slots_freed": list(filter(None, [slot_id_val, kp_slot_id_val]))})
                elif action == "restore":
                    deal_id = int(body["deal_id"])
                    cur.execute(f"UPDATE {SCHEMA}.deals SET is_archived=FALSE, updated_at=now() WHERE id=%s RETURNING code, stage", (deal_id,))
                    row = cur.fetchone()
                    if not row:
                        return err("Сделка не найдена")
                    deal_code, deal_stage = row
                    # Если сделка не в финальной стадии — восстанавливаем и связанный проект из архива в planning
                    if deal_stage not in ('lost', 'closed'):
                        cur.execute(f"""
                            UPDATE {SCHEMA}.projects SET status='planning', updated_at=now()
                            WHERE deal_id=%s AND status='archived'
                        """, (deal_id,))
                    conn.commit()
                    return ok({"success": True, "code": deal_code})
                elif action == "delete":
                    if user_role not in ("director", "commercial"):
                        return err("Удалять сделки может только директор.", 403)
                    deal_id = int(body["deal_id"])
                    # Читаем слоты и текущую стадию
                    cur.execute(f"SELECT slot_id, kp_slot_id, stage FROM {SCHEMA}.deals WHERE id=%s", (deal_id,))
                    drow = cur.fetchone()
                    if not drow:
                        return err("Сделка не найдена")
                    slot_id_val, kp_slot_id_val, deal_stage = drow

                    # Слоты освобождаем если проект ещё НЕ взят в производство.
                    # closed = проект уже active (директор по строительству нажал «Взять в производство») → слот не трогаем.
                    # planning = проект создан, но ещё не одобрен → слот можно освободить.
                    # Но если у проекта status='active', оставляем слот занятым.
                    cur.execute(f"SELECT status FROM {SCHEMA}.projects WHERE deal_id=%s", (deal_id,))
                    prow = cur.fetchone()
                    project_in_production = bool(prow and prow[0] == 'active')
                    free_slots_allowed = (deal_stage != 'closed') and (not project_in_production)
                    freed = []
                    if free_slots_allowed:
                        for sid in set(filter(None, [slot_id_val, kp_slot_id_val])):
                            cur.execute(f"""
                                UPDATE {SCHEMA}.slots SET status='free', deal_id=NULL
                                WHERE id=%s AND status IN ('booked','busy','free')
                                RETURNING id
                            """, (sid,))
                            r = cur.fetchone()
                            if r: freed.append(r[0])

                    # Архивируем связанный проект если есть
                    cur.execute(f"""
                        UPDATE {SCHEMA}.projects SET status='archived', updated_at=now()
                        WHERE deal_id=%s AND status != 'archived'
                    """, (deal_id,))
                    # Удаляем сделку
                    cur.execute(f"DELETE FROM {SCHEMA}.deals WHERE id=%s RETURNING code", (deal_id,))
                    row = cur.fetchone()
                    if not row:
                        return err("Сделка не найдена")
                    conn.commit()
                    return ok({
                        "success": True,
                        "code": row[0],
                        "slots_freed": freed,
                        "slots_kept": (not free_slots_allowed),
                    })
                else:
                    result, error = create_deal(cur, body, user_id=user_id, user_role=user_role)
                    if error:
                        return err(error)
                    conn.commit()
                    return ok(result, 201)

        # ── PROJECTS ───────────────────────────────────────────────────────────
        elif resource == "projects":
            if method == "GET":
                data = get_projects(cur, archived=qs.get("archived") == "1")
                return ok(data)
            elif method == "PUT":
                action = body.get("action")
                if action == "update_project":
                    pid    = int(body["project_id"])
                    fields = {}
                    if "status"  in body: fields["status"]  = body["status"]
                    if "brigade" in body: fields["brigade"] = body["brigade"]
                    if "address" in body: fields["address"] = body["address"]
                    if fields:
                        set_clause = ", ".join(f"{k}=%s" for k in fields)
                        cur.execute(f"UPDATE {SCHEMA}.projects SET {set_clause}, updated_at=now() WHERE id=%s",
                                    list(fields.values()) + [pid])
                        conn.commit()
                    return ok({"success": True})
                elif action == "approve_project":
                    require_role(cur, user_id, user_role, {"director", "general_director", "construction_director"})
                    pid = int(body["project_id"])
                    result, error = approve_project(cur, pid)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "cancel_project":
                    require_role(cur, user_id, user_role, {"director", "general_director", "construction_director", "commercial"})
                    pid = int(body["project_id"])
                    result, error = cancel_project(cur, pid)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "complete_project":
                    require_role(cur, user_id, user_role, {"director", "general_director", "construction_director", "foreman"})
                    pid = int(body["project_id"])
                    result, error = complete_project(cur, pid)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)

        # ── GANTT STAGES ───────────────────────────────────────────────────────
        elif resource == "gantt_stages":
            if method == "GET":
                project_id = int(qs.get("project_id", 0))
                if not project_id:
                    return err("project_id обязателен")
                data = get_gantt_stages(cur, project_id)
                return ok(data)
            elif method == "POST":
                action = body.get("action", "update_progress")
                if action == "update_progress":
                    stage_id = int(body["stage_id"])
                    progress = int(body["progress_percent"])
                    result, error = update_stage_progress(cur, stage_id, progress)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "add_group":
                    project_id = int(body["project_id"])
                    result, error = add_gantt_group(cur, project_id, body)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "add_substage":
                    project_id = int(body["project_id"])
                    result, error = add_gantt_substage(cur, project_id, body)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "delete_stage":
                    stage_id = int(body["stage_id"])
                    result, error = delete_gantt_stage(cur, stage_id)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)

        # ── PROCUREMENT ────────────────────────────────────────────────────────
        elif resource == "procurement":
            if method == "GET":
                data = get_procurement(cur)
                return ok(data)
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update_status":
                    req_id     = int(body["id"])
                    new_status = body["status"]
                    ok_res = update_request_status(cur, req_id, new_status)
                    # Уведомление прорабу при статусе "закуплено"
                    if ok_res and new_status in ("purchased", "delivered"):
                        cur.execute(f"""
                            SELECT mr.code, mr.material, mr.foreman_id, p.code as project_code
                            FROM {SCHEMA}.material_requests mr
                            LEFT JOIN {SCHEMA}.projects p ON p.id = mr.project_id
                            WHERE mr.id = %s
                        """, (req_id,))
                        mr_row = cur.fetchone()
                        if mr_row:
                            mr_code, material, foreman_id, project_code = mr_row
                            create_notification(cur,
                                type_="material_purchased",
                                title=f"Материал закуплен: {material}",
                                body_text=f"Заявка {mr_code} по проекту {project_code or '—'} переведена в статус «Закуплено».",
                                role="foreman",
                                staff_id=foreman_id,
                            )
                    conn.commit()
                    return ok({"ok": ok_res})
                else:
                    result = create_material_request(cur, body)
                    # Уведомление снабженцу о новой заявке
                    cur.execute(f"""
                        SELECT mr.code, mr.material, p.code as project_code
                        FROM {SCHEMA}.material_requests mr
                        LEFT JOIN {SCHEMA}.projects p ON p.id = mr.project_id
                        WHERE mr.id = %s
                    """, (result["id"],))
                    mr_row = cur.fetchone()
                    if mr_row:
                        mr_code, material, project_code = mr_row
                        create_notification(cur,
                            type_="material_request_new",
                            title=f"Новая заявка на материалы по проекту {project_code or '—'}",
                            body_text=f"{material} · заявка {mr_code}. Перейдите в раздел Снабжение для обработки.",
                            role="supplier",
                        )
                    conn.commit()
                    return ok(result, 201)

        # ── SUPPLIERS ──────────────────────────────────────────────────────────
        elif resource == "suppliers":
            if method == "GET":
                return ok(get_suppliers(cur))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update":
                    sid = int(body["id"])
                    update_supplier(cur, sid, body)
                    conn.commit()
                    return ok({"ok": True})
                elif action == "import_csv":
                    result = import_suppliers_csv(cur, body.get("rows", []))
                    conn.commit()
                    return ok(result)
                else:
                    result = create_supplier(cur, body)
                    conn.commit()
                    return ok(result, 201)

        # ── MATERIALS ──────────────────────────────────────────────────────────
        elif resource == "materials":
            if method == "GET":
                return ok(get_materials(cur))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update":
                    mid = int(body["id"])
                    update_material(cur, mid, body)
                    conn.commit()
                    return ok({"ok": True})
                else:
                    try:
                        result = create_material(cur, body)
                        conn.commit()
                        return ok(result, 201)
                    except Exception as e:
                        if "unique" in str(e).lower():
                            return err("Материал с таким наименованием и единицей уже существует")
                        raise

        # ── INVOICES ───────────────────────────────────────────────────────────
        elif resource == "invoices":
            if method == "GET":
                return ok(get_invoices(cur))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update":
                    iid = int(body["id"])
                    update_invoice(cur, iid, body)
                    conn.commit()
                    return ok({"ok": True})
                elif action == "upload_file":
                    iid      = int(body["invoice_id"])
                    file_b64 = body.get("file_b64", "")
                    file_name= (body.get("file_name") or "").strip()
                    if not file_b64 or not file_name:
                        return err("file_b64 и file_name обязательны")
                    result, error = upload_invoice_file(cur, iid, file_b64, file_name)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "recognize":
                    if user_role not in ("director", "supply_director", "supplier"):
                        return err("Нет прав для AI-распознавания", 403)
                    iid = int(body["invoice_id"])
                    result, error = recognize_invoice(cur, iid)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "apply_items":
                    if user_role not in ("director", "supply_director", "supplier"):
                        return err("Нет прав", 403)
                    src_id       = int(body["invoice_id"])
                    items        = body.get("items") or []
                    inv_date     = body.get("invoice_date") or None
                    inv_number   = body.get("invoice_number") or None
                    file_url     = (body.get("file_url") or "").strip() or None
                    file_name    = (body.get("file_name") or "").strip() or None
                    if not items:
                        return err("Список позиций пуст")
                    created = apply_invoice_items(cur, src_id, items, inv_date, inv_number, file_url, file_name)
                    conn.commit()
                    return ok({"created_ids": created, "count": len(created)})
                else:
                    result = create_invoice(cur, body)
                    conn.commit()
                    return ok(result, 201)

        # ── TABLE TEMPLATES ────────────────────────────────────────────────────
        elif resource == "table_templates":
            if user_role not in ("director", "supply_director", "supplier"):
                return err("Нет прав", 403)
            if method == "GET":
                return ok(get_table_templates(cur))
            elif method == "POST":
                action = body.get("action", "")
                if action == "save":
                    import requests as _rl
                    name       = (body.get("name") or "").strip()
                    headers    = body.get("headers") or []
                    column_map = body.get("column_map") or {}
                    if not headers or not column_map:
                        return err("headers и column_map обязательны")
                    if not name:
                        name = "Шаблон: " + " / ".join(str(h) for h in headers[:4])
                    result = save_user_template(cur, _rl, name, headers, column_map)
                    conn.commit()
                    return ok(result, 201)
                elif action == "apply_locally":
                    # Парсинг таблицы счёта через col_map — без AI.
                    # Перечитываем файл из S3, извлекаем текст и применяем шаблон.
                    import requests as _rl, base64 as _b64, io as _io
                    invoice_id_al  = int(body["invoice_id"])
                    column_map     = body.get("column_map") or {}
                    supplier_name  = (body.get("supplier_name") or "").strip() or ""
                    invoice_date   = body.get("invoice_date") or None
                    invoice_number = body.get("invoice_number") or None

                    if not column_map:
                        return err("column_map обязателен")

                    # Получаем файл из БД
                    cur.execute(
                        f"SELECT pdf_file_url, pdf_file_name FROM {SCHEMA}.invoices WHERE id=%s",
                        (invoice_id_al,)
                    )
                    inv_row = cur.fetchone()
                    if not inv_row or not inv_row[0]:
                        return err("Файл счёта не найден")
                    file_url_al, file_name_al = inv_row
                    ext_al = (file_name_al or "").rsplit(".", 1)[-1].lower() if file_name_al else ""

                    # Скачиваем файл
                    try:
                        resp_al = _rl.get(file_url_al, timeout=30)
                        resp_al.raise_for_status()
                        file_bytes_al = resp_al.content
                    except Exception as fe:
                        return err(f"Не удалось загрузить файл: {fe}")

                    # Извлекаем текст — openpyxl для Excel, polza.ai для остального
                    import base64 as _b64al
                    file_b64_al = _b64al.b64encode(file_bytes_al).decode("utf-8")
                    raw_text_al = ""

                    if ext_al in ("xls", "xlsx"):
                        try:
                            raw_text_al = _extract_text_excel(file_bytes_al)
                        except Exception:
                            raw_text_al = ""
                        if not raw_text_al.strip():
                            # Fallback: polza.ai
                            try:
                                mime_al = ("application/vnd.ms-excel" if ext_al == "xls"
                                           else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                                _sys_al = {"role": "system", "content": "Ты — OCR. Верни текст документа."}
                                _usr_al = {"role": "user", "content": [
                                    {"type": "text", "text": "Извлеки ВЕСЬ текст из этого документа, сохраняя табличную структуру строка за строкой."},
                                    {"type": "image_url", "image_url": {"url": f"data:{mime_al};base64,{file_b64_al}"}},
                                ]}
                                raw_text_al = _call_polza(_rl, [_sys_al, _usr_al], max_tokens=4096)
                            except Exception:
                                raw_text_al = ""
                    elif ext_al == "pdf":
                        try:
                            _sys_al = {"role": "system", "content": "Ты — OCR. Верни текст документа."}
                            _usr_al = {"role": "user", "content": [
                                {"type": "text", "text": "Извлеки ВЕСЬ текст из этого PDF, сохраняя табличную структуру строка за строкой."},
                                {"type": "image_url", "image_url": {"url": f"data:application/pdf;base64,{file_b64_al}"}},
                            ]}
                            raw_text_al = _call_polza(_rl, [_sys_al, _usr_al], max_tokens=4096)
                        except Exception:
                            raw_text_al = ""
                    else:
                        return err(f"Формат {ext_al} не поддерживается для локального парсинга")

                    if not raw_text_al.strip():
                        return err("Не удалось извлечь текст из файла для локального парсинга")

                    zones_al = _split_document(raw_text_al)
                    table_lines_al = zones_al["table_lines"]
                    tbl_headers_al, hdr_idx_al = _find_table_header_row(table_lines_al)

                    raw_items = _apply_template(
                        table_lines_al, column_map, hdr_idx_al,
                        supplier_name, invoice_date, invoice_number
                    )
                    processed_al = _postprocess_items(
                        raw_items, supplier_name, invoice_date, invoice_number
                    )
                    norm_items = _normalize_postprocessed(cur, processed_al)

                    complete_ratio = (
                        sum(1 for it in norm_items if it.get("unit_price") and it.get("quantity"))
                        / max(len(norm_items), 1)
                    )

                    conn.commit()
                    return ok({
                        "items":          norm_items,
                        "items_count":    len(norm_items),
                        "complete_ratio": round(complete_ratio, 2),
                        "low_quality":    complete_ratio < 0.5,
                    })
                elif action == "delete":
                    tid = int(body["id"])
                    ok_res = delete_table_template(cur, tid)
                    if not ok_res: return err("Шаблон не найден", 404)
                    conn.commit()
                    return ok({"ok": True})
                elif action == "rename":
                    tid      = int(body["id"])
                    new_name = (body.get("name") or "").strip()
                    if not new_name: return err("name обязателен")
                    ok_res = rename_table_template(cur, tid, new_name)
                    if not ok_res: return err("Шаблон не найден", 404)
                    conn.commit()
                    return ok({"ok": True})

        # ── PURCHASE REQUESTS ──────────────────────────────────────────────────
        elif resource == "purchase_requests":
            if method == "GET":
                return ok(get_purchase_requests(cur))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update":
                    rid = int(body["id"])
                    update_purchase_request(cur, rid, body)
                    conn.commit()
                    return ok({"ok": True})
                else:
                    result = create_purchase_request(cur, body)
                    conn.commit()
                    return ok(result, 201)

        # ── PURCHASE PLAN ──────────────────────────────────────────────────────
        elif resource == "purchase_plan":
            if method == "GET":
                return ok(get_purchase_plan(cur))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "delete":
                    pid = int(body["id"])
                    result = delete_purchase_plan(cur, pid)
                    conn.commit()
                    return ok(result)
                else:
                    result = create_purchase_plan(cur, body)
                    conn.commit()
                    return ok(result, 201)

        # ── PAYMENTS ───────────────────────────────────────────────────────────
        elif resource == "payments":
            if method == "GET":
                action = qs.get("action", "list")
                if action == "pl":
                    data = get_pl_summary(cur)
                else:
                    data = get_payments(cur)
                return ok(data)
            elif method == "POST":
                # Создавать платежи могут только финансовые роли и директор
                require_role(cur, user_id, user_role,
                             {"director", "general_director", "finance_director", "accountant"})
                result = create_payment(cur, body)
                conn.commit()
                return ok(result, 201)

        # ── K_COMPANY ──────────────────────────────────────────────────────────
        elif resource == "kcompany":
            if method == "GET":
                action = qs.get("action", "last")
                if action == "calc":
                    data = calculate_kcompany(cur)
                    conn.commit()
                else:
                    data = get_kcompany_last(cur)
                    if not data:
                        data = calculate_kcompany(cur)
                        conn.commit()
                return ok(data)

        # ── DASHBOARD ──────────────────────────────────────────────────────────
        elif resource == "dashboard":
            data = get_dashboard(cur)
            return ok(data)

        # ── CLIENTS ────────────────────────────────────────────────────────────
        elif resource == "clients":
            if method == "GET":
                return ok(get_clients(cur))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "create":
                    client, err_msg = create_client(
                        cur,
                        name=body.get("name",""),
                        phone=body.get("phone",""),
                        email=body.get("email",""),
                        source=body.get("source","CRM"),
                    )
                    if err_msg:
                        return err(err_msg, 400)
                    conn.commit()
                    return ok(client, 201)

        # ── STAFF ──────────────────────────────────────────────────────────────
        elif resource == "staff":
            role_filter = qs.get("role")
            return ok(get_staff(cur, role_filter))

        # ── REALTORS REPORT (для коммерческого директора) ─────────────────────
        elif resource == "realtors_report":
            # Доступ только директору и коммерческому директору.
            allowed = {"director", "commercial", "general_director", "commercial_director"}
            role_lower = (user_role or "").strip().lower()
            if role_lower and role_lower not in allowed:
                return err("Нет доступа", 403)
            return ok(get_realtors_report(cur))

        # ── EMPLOYEES ──────────────────────────────────────────────────────────
        elif resource == "employees":
            if method == "GET":
                return ok(get_employees(cur))
            elif method == "POST":
                result = create_employee(cur, body)
                conn.commit()
                return ok(result, 201)

        # ── REPORTS ────────────────────────────────────────────────────────────
        elif resource == "reports":
            if method == "GET":
                return ok(get_reports(cur))

        # ── SLOTS ──────────────────────────────────────────────────────────────
        elif resource == "slots":
            if method == "GET":
                action = qs.get("action", "free")
                if action == "plan":
                    show_archived = qs.get("show_archived") == "1"
                    return ok(get_slot_plan(cur, show_archived))
                else:
                    # Передаём дату подписания для фильтрации: только слоты >= signed_date + 15д
                    signed_date = qs.get("signed_date")
                    return ok(get_free_slots(cur, signed_date))
            elif method == "POST":
                action = body.get("action", "update_limit")
                if action == "create_slots":
                    year  = int(body["year"])
                    month = int(body["month"])
                    count = int(body.get("count", 4))
                    limit = int(body.get("monthly_limit", 4))
                    result = create_slots(cur, year, month, count, limit)
                    conn.commit()
                    return ok(result)
                elif action == "delete_slot":
                    slot_id = int(body["slot_id"])
                    result, error = delete_slot(cur, slot_id)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                else:
                    # Обновить лимит месяца
                    result = update_slot_limit(cur, body["year"], body["month"], body["monthly_limit"])
                    conn.commit()
                    return ok(result)

        # ── SERIAL PROJECTS ────────────────────────────────────────────────────
        elif resource == "serial_projects":
            if method == "GET":
                return ok(get_serial_projects(cur))
            elif method == "POST":
                result = create_serial_project(cur, body)
                conn.commit()
                return ok(result, 201)

        # ── CONFIGURATIONS ─────────────────────────────────────────────────────
        elif resource == "configurations":
            if method == "GET":
                sp_id = qs.get("serial_project_id")
                if not sp_id:
                    return err("serial_project_id required")
                return ok(get_configurations(cur, int(sp_id)))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update":
                    cfg_id = int(body["id"])
                    # Если устанавливаем скидку — уведомляем менеджеров
                    is_new_discount = float(body.get("discount_pct", 0)) > 0
                    result = update_configuration(cur, cfg_id, body)
                    if is_new_discount:
                        cur.execute(f"SELECT name FROM {SCHEMA}.configurations WHERE id=%s", (cfg_id,))
                        cfg_name = (cur.fetchone() or [""])[0]
                        pct = float(body.get("discount_pct", 0))
                        until = body.get("discount_until", "")
                        create_notification(cur,
                            type_="discount_set",
                            title=f"Новая скидка на «{cfg_name}»",
                            body_text=f"Директор установил скидку {pct:.0f}%{f' до {until}' if until else ''}. Предложите клиентам!",
                            role="crm_manager",
                            deal_id=None,
                        )
                    conn.commit()
                    return ok(result)
                else:
                    result = create_configuration(cur, body)
                    conn.commit()
                    return ok(result, 201)

        # ── INDIVIDUAL REQUESTS ────────────────────────────────────────────────
        elif resource == "individual_requests":
            if method == "GET":
                return ok(get_individual_requests(cur))
            elif method == "POST":
                req_id = int(body.get("id", 0))
                if req_id:
                    ok_res = update_individual_request(cur, req_id, body)
                    conn.commit()
                    return ok({"ok": ok_res})
                return err("id required")

        # ── STAGE DURATIONS (нормативы директора) ──────────────────────────────
        elif resource == "stage_durations":
            if method == "GET":
                return ok(get_stage_durations(cur))
            elif method == "POST":
                stage_num    = int(body["stage_num"])
                dur          = int(body["duration_days"])
                cur.execute(f"""
                    UPDATE {SCHEMA}.stage_durations
                    SET duration_days=%s, updated_at=now()
                    WHERE stage_num=%s
                    RETURNING stage_num, stage_name, duration_days
                """, (dur, stage_num))
                row = cur.fetchone()
                if not row:
                    return err("Этап не найден")
                # Норматив обновлён — применяется только к новым проектам (lead/kp).
                # Активные проекты (status active/planning) не затрагиваем.
                conn.commit()
                return ok({
                    "stage_num": row[0],
                    "stage_name": row[1],
                    "duration_days": row[2],
                })

        # ── ESTIMATE (полная смета: работы + материалы) ────────────────────────
        elif resource == "estimate":
            sp_id = qs.get("serial_project_id") or body.get("serial_project_id")
            if not sp_id:
                return err("serial_project_id required")
            return ok(get_estimate(cur, int(sp_id)))

        elif resource == "estimate_works":
            if method == "GET":
                sp_id = qs.get("serial_project_id")
                if not sp_id:
                    return err("serial_project_id required")
                est = get_estimate(cur, int(sp_id))
                return ok([w for s in est["stages"] for w in s["works"]])
            elif method == "POST":
                result = upsert_estimate_row(cur, "stage_works", body)
                conn.commit()
                return ok(result, 201)

        elif resource == "estimate_materials":
            if method == "GET":
                sp_id = qs.get("serial_project_id")
                if not sp_id:
                    return err("serial_project_id required")
                est = get_estimate(cur, int(sp_id))
                return ok([m for s in est["stages"] for m in s["materials"]])
            elif method == "POST":
                result = upsert_estimate_row(cur, "stage_materials", body)
                conn.commit()
                return ok(result, 201)

        # ── CONTRACTORS ────────────────────────────────────────────────────────
        elif resource == "contractors":
            if method == "GET":
                ctype = qs.get("type")
                return ok(get_contractors(cur, ctype))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update":
                    cid = int(body["id"])
                    update_contractor(cur, cid, body)
                    conn.commit()
                    return ok({"id": cid, "ok": True})
                else:
                    result = create_contractor(cur, body)
                    conn.commit()
                    return ok(result, 201)

        # ── DOCUMENTS ──────────────────────────────────────────────────────────
        elif resource == "documents":
            if method == "GET":
                category    = qs.get("category")
                deal_id     = int(qs["deal_id"])    if qs.get("deal_id")     else None
                contractor_id = int(qs["contractor_id"]) if qs.get("contractor_id") else None
                project_id  = int(qs["project_id"]) if qs.get("project_id") else None
                return ok(get_documents(cur, category, deal_id, contractor_id, project_id))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update_status":
                    doc_id = int(body["id"])
                    result = update_document_status(
                        cur, doc_id, body["status"],
                        body.get("file_url"), body.get("file_name")
                    )
                    conn.commit()
                    return ok(result)
                else:
                    result = create_document(cur, body)
                    conn.commit()
                    return ok(result, 201)

        # ── DOC TEMPLATES (директор управляет пакетом) ────────────────────────
        elif resource == "doc_templates":
            if method == "GET":
                active_only = qs.get("all") != "1"
                return ok(get_doc_templates(cur, active_only))
            elif method == "POST":
                action = body.get("action", "create")
                if action == "update":
                    tpl_id = int(body["id"])
                    update_doc_template(cur, tpl_id, body)
                    conn.commit()
                    return ok({"id": tpl_id, "ok": True})
                else:
                    result = create_doc_template(cur, body)
                    conn.commit()
                    return ok(result, 201)

        # ── CONTRACT DOCUMENTS (пакет по сделке) ──────────────────────────────
        elif resource == "contract_docs":
            if method == "GET":
                deal_id = qs.get("deal_id")
                if not deal_id:
                    return err("deal_id required")
                # Возвращаем пакет + contract_status сделки
                pkg = get_contract_docs(cur, int(deal_id))
                cur.execute(f"SELECT contract_status FROM {SCHEMA}.deals WHERE id=%s", (int(deal_id),))
                cs_row = cur.fetchone()
                pkg["contract_status"] = cs_row[0] if cs_row else "none"
                return ok(pkg)
            elif method == "POST":
                action  = body.get("action", "upload")
                deal_id = int(body["deal_id"])

                if action == "upload":
                    template_id = int(body["template_id"])
                    result = upload_contract_doc(cur, deal_id, template_id, body)
                    conn.commit()
                    return ok(result)

                elif action == "submit_review":
                    # Менеджер отправляет на проверку директору
                    result, error = submit_docs_for_review(cur, deal_id)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)

                elif action == "approve":
                    # Директор подтверждает/отклоняет — только директорские роли
                    require_role(cur, user_id, user_role,
                                 {"director", "general_director", "commercial", "commercial_director"})
                    approved = bool(body.get("approved", True))
                    reason   = body.get("reject_reason", "")
                    result, error = approve_docs(cur, deal_id, approved, reason)
                    if error: return err(error)
                    # Если подтвердил — сразу переводим в ожидание оплаты
                    if approved:
                        set_payment_pending(cur, deal_id)
                    conn.commit()
                    return ok(result)

                elif action == "confirm_payment":
                    # Директор подтверждает оплату — только директорские/финансовые роли
                    require_role(cur, user_id, user_role,
                                 {"director", "general_director", "finance_director", "accountant"})
                    result, error = confirm_payment(cur, deal_id)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)

                elif action == "upload_signed":
                    # Директор загружает подписанный вариант документа
                    template_id = int(body["template_id"])
                    result, error = upload_signed_doc(cur, deal_id, template_id, body)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)

                elif action == "confirm_doc_payment":
                    # Директор нажимает «Оплата прошла» у документа
                    template_id = int(body["template_id"])
                    result, error = confirm_doc_payment(cur, deal_id, template_id)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)

        # ── PAYOUT REQUESTS ────────────────────────────────────────────────────
        elif resource == "payout_requests":
            if method == "GET":
                manager_id = int(qs["manager_id"]) if qs.get("manager_id") else None
                return ok({"deals": get_payout_deals(cur, manager_id)})
            elif method == "POST":
                action = body.get("action", "create")
                if action == "create":
                    deal_id_pr = int(body["deal_id"])
                    mgr_id     = int(body["manager_id"])
                    result, error = create_payout_request(cur, deal_id_pr, mgr_id, body)
                    if error: return err(error)
                    conn.commit()
                    return ok(result, 201)
                elif action == "update":
                    # Одобрять/отклонять выплаты может только директор
                    require_role(cur, user_id, user_role,
                                 {"director", "general_director", "finance_director"})
                    payout_id = int(body["payout_id"])
                    result, error = update_payout_request(cur, payout_id, body)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "resubmit":
                    # Менеджер повторно подаёт счёт после отклонения
                    payout_id = int(body["payout_id"])
                    file_b64  = body.get("invoice_file_b64")
                    file_name = body.get("invoice_file_name")
                    amount    = body.get("amount")
                    invoice_url = None
                    if file_b64 and file_name:
                        invoice_url, _ = upload_file_to_s3(file_b64, f"invoice_resubmit_{payout_id}_{file_name}", "payouts")
                    cur.execute(f"""
                        UPDATE {SCHEMA}.payout_requests
                        SET status='pending', invoice_file_url=COALESCE(%s, invoice_file_url),
                            invoice_file_name=COALESCE(%s, invoice_file_name),
                            amount=COALESCE(%s, amount), reject_comment=NULL
                        WHERE id=%s RETURNING deal_id
                    """, (invoice_url, file_name if file_b64 else None, amount, payout_id))
                    row = cur.fetchone()
                    if row:
                        cur.execute(f"SELECT code FROM {SCHEMA}.deals WHERE id=%s", (row[0],))
                        deal_code = (cur.fetchone() or [""])[0]
                        create_notification(cur, type_="payout_requested",
                            title=f"Новый счёт на оплату: {deal_code}",
                            body_text="Менеджер загрузил исправленный счёт.", role="director", deal_id=row[0])
                    conn.commit()
                    return ok({"ok": True})

        # ── NOTIFICATIONS ──────────────────────────────────────────────────────
        elif resource == "notifications":
            if method == "GET":
                role      = qs.get("role")
                staff_id  = int(qs["staff_id"]) if qs.get("staff_id") else None
                unread    = qs.get("unread") == "1"
                notifs    = get_notifications(cur, role, staff_id, unread)
                count     = get_unread_count(cur, role, staff_id)
                return ok({"notifications": notifs, "unread_count": count})
            elif method == "POST":
                action = body.get("action", "read")
                if action == "read":
                    ids = body.get("ids", [])
                    mark_notifications_read(cur, ids)
                    conn.commit()
                    return ok({"ok": True})

        # ── CLIENT PORTAL ──────────────────────────────────────────────────────
        elif resource == "client_portal":
            if method == "GET":
                token = qs.get("token", "")
                if not token:
                    return err("token обязателен", 400)
                data = get_client_portal(cur, token)
                if not data:
                    return err("Страница не найдена", 404)
                conn.commit()  # сохраняем авто-созданный акт если был
                return ok(data)
            elif method == "POST":
                action = body.get("action")
                if action == "sign_act":
                    act_id = int(body["act_id"])
                    result, error = sign_client_act(cur, act_id)
                    if error: return err(error)
                    conn.commit()
                    return ok(result)
                elif action == "get_token":
                    deal_id = int(body["deal_id"])
                    token = ensure_client_token(cur, deal_id)
                    conn.commit()
                    return ok({"client_token": token})
                elif action == "create_act":
                    project_id = int(body["project_id"])
                    stage_id   = int(body["stage_id"])
                    amount     = float(body.get("amount", 0))
                    title      = body.get("title", "")
                    act, error = create_client_act(cur, project_id, stage_id, amount, title)
                    if error: return err(error)
                    conn.commit()
                    return ok(act, 201)

        # ── EXTRACT PDF TEXT ───────────────────────────────────────────────────
        elif resource == "extract_pdf_text":
            if method != "POST":
                return err("Method not allowed", 405)
            pdf_b64 = body.get("pdf_b64", "")
            if not pdf_b64:
                return err("pdf_b64 обязателен", 400)
            import base64 as _b64
            import io as _io
            try:
                pdf_bytes = _b64.b64decode(pdf_b64)
            except Exception:
                return err("Некорректный base64", 400)
            try:
                from pypdf import PdfReader
                reader = PdfReader(_io.BytesIO(pdf_bytes))
                parts  = []
                for page in reader.pages:
                    text = page.extract_text() or ""
                    if text.strip():
                        parts.append(text.strip())
                full_text = "\n\n".join(parts)
                return ok({"text": full_text, "pages": len(reader.pages), "has_text": bool(full_text.strip())})
            except Exception as e:
                logger.error(f"PDF extract error: {type(e).__name__}: {e}")
                return err("Не удалось прочитать PDF. Попробуйте сохранить файл как JPG.", 500)

        return err("Маршрут не найден", 404)

    except PermissionError as pe:
        # Недостаточно прав — 403, понятное сообщение пользователю
        conn.rollback()
        logger.warning(f"PermissionDenied: user_id={user_id} role={user_role} resource={resource}: {pe}")
        return err(str(pe), 403)
    except ValueError as ve:
        # Бизнес-валидация — пользовательская ошибка, 400 + понятное сообщение
        conn.rollback()
        logger.warning(f"ValidationError: {ve}")
        return err(str(ve), 400)
    except KeyError as ke:
        conn.rollback()
        key = str(ke).strip("'\"")
        logger.error(f"Unexpected KeyError: {key}\n{traceback.format_exc()}")
        # Технические поля (supplier_name и т.п.) не должны блокировать пользователя
        _non_blocking = {"supplier_name", "supplier", "vendor"}
        if key in _non_blocking:
            return err("Не удалось определить поставщика. Счёт будет создан без поставщика.", 400)
        _field_labels = {
            "invoice_id":   "ID счёта",
            "invoice_date": "дата счёта",
            "unit_price":   "цена за единицу",
            "quantity":     "количество",
            "material":     "наименование материала",
        }
        label = _field_labels.get(key, f"«{key}»")
        return err(f"Не указано обязательное поле: {label}", 400)
    except Exception as e:
        conn.rollback()
        # Логируем полный traceback для диагностики
        logger.error(f"Unhandled error: {type(e).__name__}: {e}\n{traceback.format_exc()}")
        return err("Внутренняя ошибка сервера. Попробуйте позже.", 500)
    finally:
        cur.close()
        conn.close()